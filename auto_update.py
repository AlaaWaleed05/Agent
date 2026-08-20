"""
auto_update.py
سكريبت مستقل (بره Streamlit) بيشتغل من GitHub Actions على جدول زمني.
مهمته:
1. لكل مكتب فعّل "تحديث تلقائي"، لو موعد التحديث بتاعه (حسب الفترة اللي
   اختارها: 6/12/24 ساعة) قرّب بـ 10 دقايق ولسه الـ Worker المحلي مش شغال،
   يبعت إيميل تذكير للمطوّرة عشان تفتح الـ Worker لو عايزة التحديث يمشي
   عن طريقه (أأمن ضد الحظر) بدل الـ Cloud.
2. لو موعد التحديث جه فعلاً:
   - لو الـ Worker شغال → يفوّض المهمة ليه (نفس آلية app.py) ويستنى يخلص.
   - لو مش شغال → يعمل التحديث بنفسه زي المعتاد (الطريقة الأصلية، Cloud).
3. بيحدّث بس عمود "حالة الطلب الجديدة" في شيت المكتب المربوط (خلية خلية،
   بدون مسح أو إعادة كتابة أي حاجة تانية في الشيت وبدون تغيير ترتيب الصفوف).
4. بيمر على *كل* الطلاب المطلوبين في كل دورة (مفيش تقسيم لدفعات عبر أيام)،
   بترتيب معالجة بيتشافل كل دورة (في الميموري بس)، ما عدا الطلاب اللي
   وصلوا لحالة نهائية (زي "قبول نهائي") فبيتم استبعادهم.
5. بيبعت إيميل للمكتب (لو حصل تغيير فعلي في حالة أي طالب) وبيحدّث "آخر
   تحديث تلقائي" بعد ما الدورة الكاملة تخلص.

الفرق عن app.py: الملف ده مالوش أي تبعية على Streamlit، وبياخد بيانات
الاعتماد من متغيرات بيئة بدل st.secrets.
"""

import os
import io
import json
import time
import random
import requests
import openpyxl
import gspread
from datetime import datetime, timedelta
from google.oauth2.service_account import Credentials

SHEET_ID = "1BlFdtY-7ZIF1y2GwVosxlG9r7nK5xqYeW6yiIjPI_9U"
DRIVE_FOLDER_ID = "12L_qSHBnW4-tfQZRteynInWNBAML016f"

BASE_URL = "https://apiadm.study-in-egypt.gov.eg/api"
SITE_URL = "https://admission.study-in-egypt.gov.eg"

HEADERS_BASE = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "ar",
    "device": "CITIZEN",
    "origin": SITE_URL,
    "referer": SITE_URL + "/",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "content-type": "application/json",
}

# الحالات اللي لو الطالب وصلها، ما بنعيدش فحصه
FINAL_STATUSES = {
    "مقبول نهائي",
    "تم الرفض",
    "مرفوض نهائيًا",
    "مرفوض نهائيا",
}

WORKER_ONLINE_THRESHOLD_SECONDS = 150  # نفس القيمة المستخدمة في app.py
REMINDER_WINDOW_MINUTES = 10           # نبعت تذكير لو الموعد هيجي خلال كام دقيقة
JOB_MAX_WAIT_SECONDS = 60 * 90         # أقصى وقت ننتظر فيه الـ Worker يخلص المهمة

API_RETRY_MAX = 5
API_RETRY_BASE_WAIT = 20


# ==================== Retry wrapper ====================

def with_retry(fn, *args, **kwargs):
    last_err = None
    for attempt in range(1, API_RETRY_MAX + 1):
        try:
            return fn(*args, **kwargs)
        except gspread.exceptions.APIError as e:
            last_err = e
            msg = str(e)
            if "429" in msg or "Quota exceeded" in msg or "RESOURCE_EXHAUSTED" in msg:
                wait = API_RETRY_BASE_WAIT * attempt
                print(f"⚠️ Quota exceeded (محاولة {attempt}/{API_RETRY_MAX}) — استنى {wait} ثانية...")
                time.sleep(wait)
                continue
            raise
    raise last_err


# ==================== Credentials ====================

def get_creds():
    creds_json = os.environ["GCP_SERVICE_ACCOUNT_JSON"]
    creds_dict = json.loads(creds_json)
    return Credentials.from_service_account_info(
        creds_dict,
        scopes=[
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive",
        ],
    )


def get_gspread_client():
    return gspread.authorize(get_creds())


# ==================== API helpers (نفس منطق app.py) ====================

def human_delay(min_sec=2, max_sec=5):
    time.sleep(random.uniform(min_sec, max_sec))


def api_login(email, password, max_retries=3):
    try:
        session = requests.Session()
        session.headers.update(HEADERS_BASE)
        human_delay(2, 4)
        res = session.post(
            f"{BASE_URL}/student/login",
            json={"email": email, "password": password},
            timeout=30,
        )
        if res.status_code not in [200, 201]:
            return None, None, f"فشل اللوجين - كود: {res.status_code}"
        human_delay(2, 3)
        csrf_token = res.json().get("token", "") or res.headers.get("x-csrf-token", "")
        return session, csrf_token, None
    except Exception as e:
        return None, None, str(e)


def api_logout(session):
    try:
        session.post(f"{BASE_URL}/student/logout", json={"redirectUrl": SITE_URL}, timeout=15)
        human_delay(1, 2)
    except Exception:
        pass


def get_status(session, csrf_token):
    try:
        filter_obj = {
            "where": {},
            "limit": 10,
            "offset": 0,
            "order": "statusUpdatedAt DESC",
            "fields": ["serviceSlug", "ID", "createdAt", "statusUpdatedAt", "activityId", "activityName"],
        }
        headers = {}
        if csrf_token:
            headers["x-csrf-token"] = csrf_token
        human_delay(1, 3)
        res = session.get(
            f"{BASE_URL}/dynamic_services/inbox",
            params={"filter": json.dumps(filter_obj)},
            headers=headers,
            timeout=30,
        )
        if res.status_code not in [200, 304]:
            return "", f"خطأ ({res.status_code})"
        data = res.json()
        results = data.get("result", [])
        if not results:
            return "", "مفيش طلبات"
        latest = results[0]
        translations = {
            "قبول الفحص الفنى": "القبول المبدئي",
            "قبول الفحص الفني": "القبول المبدئي",
            "kb8ijfo8": "تم السداد",
            "تم السداد": "تم السداد",
            "تأكيد استلام الملف وصحة و اكتمال المستندات": "تأكيد استلام الملف وصحة واكتمال المستندات",
            "الانتظار مراجعة الطلب": "بانتظار مراجعة الطلب",
            "قبول من رئيس الادارة المركزية": "قبول من رئيس الإدارة المركزية",
        }
        raw_status = latest.get("activityName", "غير محدد")
        status = translations.get(raw_status, raw_status)
        return str(latest.get("ID", "")), status
    except Exception as e:
        return "", f"خطأ: {e}"


def find_excel_columns(ws):
    cols = {"name": None, "email": None, "password": None, "status": None}
    header_row_num = None
    header_len = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        row_values = [str(c).strip() if c else "" for c in row]
        if any("يميل" in v or "mail" in v.lower() for v in row_values):
            header_row_num = row_idx
            header_len = len(row_values)
            for i, cell in enumerate(row_values):
                if "حالة" in cell and "الجديدة" in cell:
                    cols["status"] = i
                    break
            for i, cell in enumerate(row_values):
                cell_lower = cell.lower()
                if any(k in cell for k in ["اسم", "الإسم", "الاسم"]) or "name" in cell_lower:
                    cols["name"] = i
                elif any(k in cell for k in ["يميل", "بريد"]) or "mail" in cell_lower:
                    cols["email"] = i
                elif any(k in cell for k in ["باسورد", "كلمة المرور", "password", "pass"]) or "pass" in cell_lower:
                    cols["password"] = i
            break
    if header_row_num is None:
        raise Exception("مش لاقي هيدر الإكسيل!")
    if cols["email"] is None:
        raise Exception("مش لاقي عمود الإيميل!")
    if cols["password"] is None:
        raise Exception("مش لاقي عمود الباسورد!")
    if cols["status"] is None:
        new_col_index = header_len
        ws.cell(row=header_row_num, column=new_col_index + 1, value="حالة الطلب الجديدة")
        cols["status"] = new_col_index
    return cols, header_row_num


def extract_sheet_id(link):
    import re
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", link)
    return match.group(1) if match else None


def extract_gid(link):
    import re
    match = re.search(r"[?#&]gid=(\d+)", link)
    return int(match.group(1)) if match else None


def get_target_worksheet(spreadsheet, link):
    gid = extract_gid(link)
    if gid is not None:
        try:
            for ws in spreadsheet.worksheets():
                if ws.id == gid:
                    return ws
        except Exception:
            pass
    return spreadsheet.sheet1


# ==================== Worker coordination (heartbeat / jobs) ====================

def get_heartbeat_sheet(client):
    spreadsheet = with_retry(client.open_by_key, SHEET_ID)
    try:
        return with_retry(spreadsheet.worksheet, "worker_heartbeat")
    except Exception:
        return None


def is_worker_online(client):
    sheet = get_heartbeat_sheet(client)
    if not sheet:
        return False
    try:
        last_beat_str = with_retry(sheet.cell, 1, 2).value
        if not last_beat_str:
            return False
        last_beat = datetime.strptime(last_beat_str, "%Y-%m-%d %H:%M:%S")
        return (datetime.now() - last_beat).total_seconds() <= WORKER_ONLINE_THRESHOLD_SECONDS
    except Exception:
        return False


def get_jobs_sheet(client):
    spreadsheet = with_retry(client.open_by_key, SHEET_ID)
    try:
        return with_retry(spreadsheet.worksheet, "jobs")
    except Exception:
        sheet = with_retry(spreadsheet.add_worksheet, "jobs", 500, 9)
        with_retry(sheet.append_row, ["job_id", "اسم المكتب", "نوع المصدر", "مرجع المصدر",
                                       "اسم الملف", "الحالة", "تاريخ الإنشاء", "final_drive_file_id", "خطأ"])
        return sheet


def create_job(client, office, source_type, source_ref, filename):
    sheet = get_jobs_sheet(client)
    job_id = f"{office}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{random.randint(1000,9999)}"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with_retry(sheet.append_row, [job_id, office, source_type, source_ref, filename, "pending", now, "", ""])
    return job_id


def get_job(client, job_id):
    sheet = get_jobs_sheet(client)
    records = with_retry(sheet.get_all_records)
    for r in records:
        if str(r.get("job_id", "")) == job_id:
            return r
    return None


def find_latest_drive_file_id(office):
    """بيدوّر على آخر ملف محفوظ للمكتب على Drive (نفس منطق get_latest_file_from_drive
    في app.py) وبيرجع الـ file_id بتاعه بس، عشان نبعته كمرجع للـ Worker."""
    try:
        from googleapiclient.discovery import build

        creds = get_creds()
        service = build("drive", "v3", credentials=creds)
        safe_office = str(office).replace("\\", "\\\\").replace("'", "\\'")
        results = (
            service.files()
            .list(
                q=f"'{DRIVE_FOLDER_ID}' in parents and name contains '{safe_office}' and trashed=false",
                orderBy="createdTime desc",
                pageSize=1,
                fields="files(id, name, createdTime)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            )
            .execute()
        )
        files = results.get("files", [])
        if not files:
            return None
        return files[0]["id"]
    except Exception as e:
        print(f"تعذر البحث عن آخر ملف Drive للمكتب {office}: {e}")
        return None


# ==================== Data source (Google Sheet المربوط أو آخر نسخة Drive) ====================

def get_office_source_workbook(client, office, gsheet_link):
    if gsheet_link:
        sid = extract_sheet_id(gsheet_link)
        if sid:
            try:
                spreadsheet = with_retry(client.open_by_key, sid)
                ws = get_target_worksheet(spreadsheet, gsheet_link)
                data = with_retry(ws.get_all_values)
                if data:
                    wb = openpyxl.Workbook()
                    wsheet = wb.active
                    for row in data:
                        wsheet.append(row)
                    return wb, "gsheet", gsheet_link
            except Exception as e:
                print(f"تعذر قراءة الشيت المربوط للمكتب {office}: {e}")

    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseDownload

        creds = get_creds()
        service = build("drive", "v3", credentials=creds)
        safe_office = str(office).replace("\\", "\\\\").replace("'", "\\'")
        results = (
            service.files()
            .list(
                q=f"'{DRIVE_FOLDER_ID}' in parents and name contains '{safe_office}' and trashed=false",
                orderBy="createdTime desc",
                pageSize=1,
                fields="files(id, name, createdTime)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            )
            .execute()
        )
        files = results.get("files", [])
        if not files:
            return None, None, None

        file_id = files[0]["id"]
        request = service.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        return wb, "drive", None
    except Exception as e:
        print(f"تعذر جلب آخر نسخة من Drive للمكتب {office}: {e}")
        return None, None, None


def upload_backup_to_drive(file_bytes, filename, office):
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload

        creds = get_creds()
        service = build("drive", "v3", credentials=creds)
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        drive_filename = f"{office}_{now}_{filename}"
        file_metadata = {"name": drive_filename, "parents": [DRIVE_FOLDER_ID]}
        media = MediaIoBaseUpload(
            io.BytesIO(file_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        service.files().create(body=file_metadata, media_body=media, fields="id", supportsAllDrives=True).execute()
    except Exception as e:
        print(f"فشل رفع نسخة احتياطية للمكتب {office}: {e}")


def ensure_status_header_in_gsheet(client, link, header_row_num, status_col_index, header_name="حالة الطلب الجديدة"):
    try:
        sheet_id = extract_sheet_id(link)
        if not sheet_id:
            return
        spreadsheet = with_retry(client.open_by_key, sheet_id)
        ws = get_target_worksheet(spreadsheet, link)
        col_num = status_col_index + 1
        current = with_retry(ws.cell, header_row_num, col_num).value
        if str(current or "").strip() != header_name:
            with_retry(ws.update_cell, header_row_num, col_num, header_name)
    except Exception as e:
        print(f"تعذر ضمان هيدر عمود الحالة: {e}")


def update_status_cell_in_gsheet(client, link, row_num, status_col_index, value):
    """بتكتب خلية واحدة بس في شيت المكتب — بدون مسح أو إعادة كتابة أي حاجة تانية"""
    try:
        sheet_id = extract_sheet_id(link)
        if not sheet_id:
            return False
        spreadsheet = with_retry(client.open_by_key, sheet_id)
        ws = get_target_worksheet(spreadsheet, link)
        with_retry(ws.update_cell, row_num, status_col_index + 1, value)
        return True
    except Exception as e:
        print(f"فشل تحديث خلية الحالة: {e}")
        return False


def get_previous_results(client, office):
    try:
        spreadsheet = with_retry(client.open_by_key, SHEET_ID)
        try:
            sheet = with_retry(spreadsheet.worksheet, "results")
        except Exception:
            return {}
        all_values = with_retry(sheet.get_all_values)
        if len(all_values) <= 1:
            return {}
        target = str(office).strip()
        previous = {}
        for row in all_values[1:]:
            if len(row) < 3:
                continue
            if str(row[0]).strip() == target:
                previous[str(row[1]).strip()] = str(row[2]).strip()
        return previous
    except Exception as e:
        print(f"تعذر جلب النتائج السابقة للمكتب {office}: {e}")
        return {}


def compute_changes(previous_results, new_results):
    changes = []
    for r in new_results:
        name = str(r.get("name", "")).strip()
        new_status = str(r.get("status", "")).strip()
        if not name:
            continue
        old_status = previous_results.get(name)
        if old_status is None:
            changes.append({"name": name, "new_status": new_status, "old_status": None})
        elif old_status != new_status:
            changes.append({"name": name, "new_status": new_status, "old_status": old_status})
    return changes


def build_email_message(office, changes):
    lines = [f"تحديث حالات الطلاب - {office}", ""]
    shown = changes[:200]
    for idx, c in enumerate(shown, start=1):
        lines.append(f"{idx}. {c['name']} - {c['new_status']}")
    if len(changes) > 200:
        lines.append("")
        lines.append(f"...و{len(changes) - 200} طالب/ة تاني اتغيرت حالتهم")
    lines.append("")
    lines.append(f"إجمالي التغييرات: {len(changes)}")
    return "\n".join(lines)


def send_email_notification(to_email, subject, body):
    import smtplib
    from email.mime.text import MIMEText

    sender_email = os.environ.get("SENDER_EMAIL")
    sender_password = os.environ.get("SENDER_APP_PASSWORD")

    if not sender_email or not sender_password:
        print("إعدادات الإيميل مش مكتملة على GitHub Secrets - تم تخطي الإرسال.")
        return
    if not to_email:
        print("مفيش إيميل مستقبل - تم تخطي الإرسال.")
        return

    try:
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = sender_email
        msg["To"] = to_email

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, [to_email], msg.as_string())
        print(f"تم إرسال إيميل إلى {to_email}")
    except Exception as e:
        print(f"خطأ في إرسال الإيميل: {e}")


def save_results_to_sheet(client, office, results):
    try:
        spreadsheet = with_retry(client.open_by_key, SHEET_ID)
        try:
            sheet = with_retry(spreadsheet.worksheet, "results")
        except Exception:
            sheet = with_retry(spreadsheet.add_worksheet, "results", 2000, 4)
            with_retry(sheet.append_row, ["اسم المكتب", "اسم الطالب", "الحالة", "تاريخ التحديث"])

        all_values = with_retry(sheet.get_all_values)
        headers = all_values[0] if all_values else ["اسم المكتب", "اسم الطالب", "الحالة", "تاريخ التحديث"]
        target = str(office).strip()
        kept_rows = [row for row in all_values[1:] if row and str(row[0]).strip() != target]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_rows = [[target, r.get("name", ""), r.get("status", ""), now] for r in results]
        final_data = [headers] + kept_rows + new_rows
        with_retry(sheet.clear)
        with_retry(sheet.update, final_data)
    except Exception as e:
        print(f"فشل حفظ نتائج المكتب {office}: {e}")


def log_to_sheet(client, office, action, filename=""):
    try:
        sheet = with_retry(client.open_by_key, SHEET_ID).sheet1
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        first_cell = with_retry(sheet.cell, 1, 1).value
        if sheet.row_count == 0 or first_cell != "التاريخ":
            with_retry(sheet.insert_row, ["التاريخ", "اسم المكتب", "العملية", "اسم الملف"], 1)
        with_retry(sheet.append_row, [now, office, action, filename])
    except Exception as e:
        print(f"خطأ في التسجيل: {e}")


# ==================== المسار المباشر (Cloud) — لما الـ Worker مش شغال ====================

def process_office_direct(client, office, gsheet_link, office_email):
    """بتعمل التحديث بنفسها (زي ما كانت الطريقة الأصلية) — بتمر على *كل*
    الطلاب المطلوبين في الدورة دي، بترتيب معالجة عشوائي (شافل)، وبتكتب
    خلية عمود 'حالة الطلب الجديدة' بس في شيت المكتب لكل طالب."""
    print(f"--- تحديث مباشر (Cloud) للمكتب: {office} ---")
    wb, source_type, source_link = get_office_source_workbook(client, office, gsheet_link)
    if wb is None:
        print(f"لا يوجد ملف محفوظ للمكتب {office}، تم التخطي.")
        return

    ws = wb.active
    try:
        cols, header_row_num = find_excel_columns(ws)
    except Exception as e:
        print(f"خطأ في قراءة أعمدة ملف المكتب {office}: {e}")
        return

    previous_results = get_previous_results(client, office)

    if source_type == "gsheet":
        ensure_status_header_in_gsheet(client, source_link, header_row_num, cols["status"])

    rows_data = list(ws.iter_rows(min_row=header_row_num + 1, values_only=False))
    all_valid_rows = [r for r in rows_data if r[cols["email"]].value and r[cols["password"]].value]

    pending_rows = []
    for r in all_valid_rows:
        name = str(r[cols["name"]].value or "").strip() if cols["name"] is not None else ""
        if previous_results.get(name) in FINAL_STATUSES:
            continue
        pending_rows.append(r)

    skipped_final = len(all_valid_rows) - len(pending_rows)
    if skipped_final:
        print(f"تم تخطي {skipped_final} طالب/ة وصلوا لحالة نهائية بالفعل.")

    total = len(pending_rows)
    if total == 0:
        print(f"مفيش طلاب محتاجين تحديث للمكتب {office}.")
        return

    # شافل ترتيب المعالجة (اللوجين على الموقع الحكومي) بس — الصفوف في الشيت
    # نفسها بتتكتب في مكانها الأصلي دايمًا بغض النظر عن ترتيب المعالجة
    processing_order = pending_rows[:]
    random.shuffle(processing_order)

    processed_results = []
    for row in processing_order:
        email = str(row[cols["email"]].value).strip()
        password = str(row[cols["password"]].value).strip()
        name = row[cols["name"]].value if cols["name"] is not None else ""
        display_name = name or email
        row_num_in_sheet = row[cols["email"]].row

        session, csrf_token, err = api_login(email, password)
        if err or not session:
            status_val = f"فشل تسجيل الدخول ({err})" if err else "فشل تسجيل الدخول"
        else:
            _, status_val = get_status(session, csrf_token)
            api_logout(session)

        if cols["status"] is not None:
            row[cols["status"]].value = status_val

        processed_results.append({"name": str(display_name), "status": str(status_val)})

        if source_type == "gsheet":
            update_status_cell_in_gsheet(client, source_link, row_num_in_sheet, cols["status"], status_val)

        human_delay(8, 15)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    upload_backup_to_drive(out.getvalue(), "auto_update.xlsx", office)

    # لازم نحافظ على نتائج الطلاب اللي كانوا في حالة نهائية أصلاً (اتستبعدوا من
    # المعالجة دي) عشان ما نفقدش سجلهم من شيت النتائج الداخلي
    skipped_results = [
        {"name": n, "status": s} for n, s in previous_results.items() if s in FINAL_STATUSES
    ]
    save_results_to_sheet(client, office, processed_results + skipped_results)

    if previous_results:
        changes = compute_changes(previous_results, processed_results)
        if changes:
            if office_email:
                subject = f"تحديث حالات الطلاب - {office}"
                body = build_email_message(office, changes)
                send_email_notification(office_email, subject, body)
            else:
                print(f"في {len(changes)} تغيير للمكتب {office}، بس مفيش إيميل محفوظ.")
        else:
            print(f"مفيش أي تغيير في حالات المكتب {office}.")
    else:
        print(f"أول تحديث تلقائي للمكتب {office} - تم حفظ النتائج بدون إرسال إشعار.")

    log_to_sheet(client, office, "اكتمل المعالجة التلقائية (Cloud)", "auto_update.xlsx")
    print(f"--- اكتمل تحديث المكتب: {office} ({total} طالب) ---")


# ==================== المسار المفوّض (Worker) — لما الـ Worker شغال ====================

def process_office_via_worker(client, office, gsheet_link, office_email):
    """بتنشئ Job للـ Worker المحلي وتستنى يخلصه، وبعدين تبعت إيميل التغييرات
    وتحدّث 'آخر تحديث تلقائي' زي ما المسار المباشر بيعمل بالظبط."""
    print(f"--- تحديث عن طريق Worker محلي للمكتب: {office} ---")
    previous_results = get_previous_results(client, office)

    if gsheet_link:
        job_id = create_job(client, office, "gsheet", gsheet_link, "google_sheet")
    else:
        file_id = find_latest_drive_file_id(office)
        if not file_id:
            print(f"مفيش ملف محفوظ للمكتب {office} على Drive، تم التخطي.")
            return False
        job_id = create_job(client, office, "drive", file_id, "auto_update.xlsx")

    if not job_id:
        print(f"مقدرتش أنشئ Job للمكتب {office}.")
        return False

    waited = 0
    poll_interval = 10
    job = None
    while waited < JOB_MAX_WAIT_SECONDS:
        job = get_job(client, job_id)
        status = (job or {}).get("الحالة", "")
        if status in ("done", "failed"):
            break
        time.sleep(poll_interval)
        waited += poll_interval

    if not job or job.get("الحالة") != "done":
        print(f"فشل أو انتهت مهلة انتظار Worker للمكتب {office}: {(job or {}).get('خطأ', 'مهلة انتظار')}")
        return False

    # النتائج والإيميلات اتحفظت/اتبعتت بالفعل من جوه worker.py نفسه (نفس آلية
    # save_results_to_sheet)، فبس بنبعت إيميل التغييرات هنا لو عايزين إشعار
    # منفصل، وبنعتبر الدورة خلصت
    new_results = get_previous_results(client, office)  # بعد ما الـ Worker حدّث النتائج
    changes = compute_changes(previous_results, [{"name": n, "status": s} for n, s in new_results.items()])
    if changes and office_email:
        subject = f"تحديث حالات الطلاب - {office}"
        body = build_email_message(office, changes)
        send_email_notification(office_email, subject, body)

    print(f"--- اكتمل تحديث المكتب عن طريق Worker: {office} ---")
    return True


# ==================== إيميل التذكير قبل الموعد بـ 10 دقايق ====================

def send_worker_reminder(office, minutes_left):
    """بيبعت إيميل تذكير للمطوّرة إن موعد تحديث مكتب معيّن هيجي قريب، ولسه
    الـ Worker المحلي مش شغال — عشان تفتحه لو عايزة التحديث يمشي عن طريقه."""
    developer_email = os.environ.get("DEVELOPER_EMAIL")
    if not developer_email:
        print("مفيش DEVELOPER_EMAIL محفوظ - تم تخطي إيميل التذكير.")
        return
    subject = f"⏰ تذكير: موعد تحديث المكتب '{office}' قريب"
    body = (
        f"موعد التحديث التلقائي للمكتب '{office}' هيجي خلال حوالي {minutes_left} دقيقة.\n"
        f"لو عايزة التحديث ده يمشي عن طريق الـ Worker المحلي (أأمن ضد الحظر)، "
        f"افتحي worker.py دلوقتي قبل ما الموعد يجي.\n"
        f"لو ما فتحتيهوش، التحديث هيمشي عادي عن طريق الـ Cloud زي المعتاد."
    )
    send_email_notification(developer_email, subject, body)


def maybe_send_reminder(accounts_sheet, headers, row_num, office, due_time, client):
    """بتبعت تذكير مرة واحدة بس لكل دورة قبل الموعد بـ REMINDER_WINDOW_MINUTES،
    وبتسجل إنها بعتت التذكير عشان ما تكررهوش."""
    now = datetime.now()
    minutes_left = (due_time - now).total_seconds() / 60
    if not (0 <= minutes_left <= REMINDER_WINDOW_MINUTES):
        return

    reminder_col_name = "آخر تذكير Worker"
    if reminder_col_name not in headers:
        col_num = len(headers) + 1
        with_retry(accounts_sheet.update_cell, 1, col_num, reminder_col_name)
        headers.append(reminder_col_name)
    col_num = headers.index(reminder_col_name) + 1

    last_reminder_str = with_retry(accounts_sheet.cell, row_num, col_num).value
    due_key = due_time.strftime("%Y-%m-%d %H:%M")
    if last_reminder_str == due_key:
        return  # اتبعت التذكير ده قبل كده لنفس الموعد

    if is_worker_online(client):
        return  # الـ Worker شغال بالفعل، مفيش داعي تذكير

    send_worker_reminder(office, round(minutes_left))
    with_retry(accounts_sheet.update_cell, row_num, col_num, due_key)


# ==================== Main ====================

def main():
    client = get_gspread_client()
    spreadsheet = with_retry(client.open_by_key, SHEET_ID)
    accounts_sheet = with_retry(spreadsheet.worksheet, "accounts")
    records = with_retry(accounts_sheet.get_all_records)
    headers = with_retry(accounts_sheet.row_values, 1)

    now = datetime.now()

    for i, r in enumerate(records, start=2):  # صف 1 = هيدر
        office = r.get("اسم المكتب", "")
        if not office:
            continue
        if r.get("الحالة", "") != "approved":
            continue
        if r.get("تحديث تلقائي", "لا") != "نعم":
            continue

        try:
            interval = int(r.get("كل كام ساعة", 12) or 12)
        except Exception:
            interval = 12

        last_update_str = r.get("آخر تحديث تلقائي", "")
        if last_update_str:
            try:
                last_update = datetime.strptime(last_update_str, "%Y-%m-%d %H:%M:%S")
                due_time = last_update + timedelta(hours=interval)
            except Exception:
                due_time = now  # لو التاريخ اتلخبط، اعتبريه مستحق دلوقتي
        else:
            due_time = now  # أول مرة، مستحق فورًا

        # ابعتي تذكير لو الموعد قرّب والـ Worker لسه مش شغال
        maybe_send_reminder(accounts_sheet, headers, i, office, due_time, client)

        due = now >= due_time
        if not due:
            print(f"المكتب {office}: لسه معملوش موعد التحديث (كل {interval} ساعة).")
            continue

        gsheet_link = r.get("لينك الشيت", "")
        office_email = str(r.get("الإيميل", "") or "").strip()

        try:
            success = False
            if is_worker_online(client):
                success = process_office_via_worker(client, office, gsheet_link, office_email)
                if not success:
                    print(f"فشل التفويض للـ Worker للمكتب {office} — هنعمل التحديث مباشرة (Cloud) بدلاً منه.")
                    process_office_direct(client, office, gsheet_link, office_email)
            else:
                process_office_direct(client, office, gsheet_link, office_email)

            if "آخر تحديث تلقائي" not in headers:
                col_num = len(headers) + 1
                with_retry(accounts_sheet.update_cell, 1, col_num, "آخر تحديث تلقائي")
                headers = with_retry(accounts_sheet.row_values, 1)
            col_num = headers.index("آخر تحديث تلقائي") + 1
            with_retry(accounts_sheet.update_cell, i, col_num, now.strftime("%Y-%m-%d %H:%M:%S"))
        except Exception as e:
            print(f"خطأ أثناء تحديث المكتب {office}: {e}")


if __name__ == "__main__":
    main()

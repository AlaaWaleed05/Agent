"""
Aivora - Agent
Your Smarter Support for Every Student's Application
Streamlit App
"""

import streamlit as st"""
Aivora - Agent
Your Smarter Support for Every Student's Application
"""

import io
import json
import os
import re
import threading
import time
from datetime import datetime, timezone

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

import bcrypt
import openpyxl
import pandas as pd
import requests
import streamlit as st
from cryptography.fernet import Fernet
from supabase import Client, create_client

try:
    import gspread
    from google.oauth2.service_account import Credentials
except Exception:
    gspread = None
    Credentials = None


BASE_URL = "https://apiadm.study-in-egypt.gov.eg/api"
SITE_URL = "https://admission.study-in-egypt.gov.eg"

WORKER_WAIT_SECONDS = 30

DRIVE_FOLDER_ID = os.getenv(
    "DRIVE_FOLDER_ID",
    "12L_qSHBnW4-tfQZRteynInWNBAML016f"
)

# Keep fallback pacing aligned with the Worker timing policy.
LOGIN_PAGE_DELAY_MIN, LOGIN_PAGE_DELAY_MAX = 0.8, 1.5
POST_LOGIN_DELAY_SECONDS = 1.0
INBOX_DELAY_MIN, INBOX_DELAY_MAX = 0.7, 1.5
STUDENT_DELAY_MIN, STUDENT_DELAY_MAX = 4, 8

TECH_FAILURE_STATUS = "تعذر فحص الطالب حاليًا"

FINAL_STATUSES = {
    "مقبول نهائي",
    "قبول نهائي",
    "تم الرفض",
    "مرفوض نهائيًا",
    "مرفوض نهائيا",
    "مرفوض",
    "خالص",
}

ADMIN_USERNAME = st.secrets.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "admin123")


# =========================================================
# SUPABASE
# =========================================================

@st.cache_resource(show_spinner=False)
def get_supabase() -> Client:
    url = st.secrets.get(
        "SUPABASE_URL",
        os.getenv("SUPABASE_URL")
    )

    key = st.secrets.get(
        "SUPABASE_SERVICE_ROLE_KEY",
        os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    )

    if not url or not key:
        raise RuntimeError("Supabase configuration is missing")

    return create_client(url, key)


def db():
    return get_supabase()


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def safe_log(message):
    print(f"[Aivora] {message}")


def fallback_delay(a, b):
    """
    Use the same timing ranges as worker.py
    without touching Streamlit state.
    """
    import random

    time.sleep(
        random.uniform(a, b)
    )


# =========================================================
# OFFICE
# =========================================================

def get_office_by_name(name):
    rows = (
        db()
        .table("offices")
        .select("id,name,email,status,created_at")
        .eq("name", str(name).strip())
        .limit(1)
        .execute()
        .data
        or []
    )

    return rows[0] if rows else None


def get_office_with_hash(office_id):
    rows = (
        db()
        .table("offices")
        .select(
            "id,name,email,status,password_hash,created_at"
        )
        .eq("id", office_id)
        .limit(1)
        .execute()
        .data
        or []
    )

    return rows[0] if rows else None


def get_office_by_email(email):
    rows = (
        db()
        .table("offices")
        .select("id,name,email,status")
        .eq(
            "email",
            str(email).strip().lower()
        )
        .limit(1)
        .execute()
        .data
        or []
    )

    return rows[0] if rows else None


def register_office(name, email, password):

    name = name.strip()
    email = email.strip().lower()
    password = password.strip()

    if not name or not email or not password:
        return False, "كمّلي كل البيانات الأول."

    if len(password) < 6:
        return False, "الباسورد لازم يكون 6 أحرف على الأقل."

    try:

        if get_office_by_name(name) or get_office_by_email(email):
            return False, "اسم المكتب أو الإيميل مسجل قبل كده."

        password_hash = (
            bcrypt
            .hashpw(
                password.encode(),
                bcrypt.gensalt()
            )
            .decode()
        )

        (
            db()
            .table("offices")
            .insert({
                "name": name,
                "email": email,
                "password_hash": password_hash,
                "status": "pending"
            })
            .execute()
        )

        return (
            True,
            "تم التسجيل بنجاح. الحساب في انتظار موافقة الإدارة."
        )

    except Exception:

        safe_log(
            "office registration failed"
        )

        return (
            False,
            "تعذر إنشاء الحساب حاليًا. حاولي مرة تانية."
        )


def check_login(name, password):

    try:

        office = get_office_by_name(name)

        if not office:
            return False, "اسم المكتب أو الباسورد غلط!"

        if office.get("status") == "pending":
            return False, "حسابك في انتظار موافقة الإدارة!"

        if office.get("status") != "approved":
            return False, "حسابك موقوف!"

        full = get_office_with_hash(
            office["id"]
        )

        password_hash = str(
            (full or {}).get("password_hash")
            or ""
        )

        if (
            not password_hash
            or not bcrypt.checkpw(
                password.encode(),
                password_hash.encode()
            )
        ):
            return False, "اسم المكتب أو الباسورد غلط!"

        return True, office

    except Exception:

        safe_log(
            "login check failed"
        )

        return (
            False,
            "تعذر تسجيل الدخول حاليًا. حاولي مرة تانية."
        )


def get_pending_accounts():

    try:

        return (
            db()
            .table("offices")
            .select(
                "id,name,email,status,created_at"
            )
            .eq("status", "pending")
            .order("created_at")
            .execute()
            .data
            or []
        )

    except Exception:
        return []


def set_office_status(office_id, status):

    (
        db()
        .table("offices")
        .update({
            "status": status
        })
        .eq("id", office_id)
        .execute()
    )


def log_activity(
    office_id,
    action,
    file_name="",
    details=None,
    student_id=None,
    data_source_id=None
):

    try:

        (
            db()
            .table("activity_logs")
            .insert({
                "office_id": office_id,
                "student_record_id": student_id,
                "data_source_id": data_source_id,
                "action": action,
                "file_name": file_name,
                "details": details or {},
            })
            .execute()
        )

    except Exception as exc:

        safe_log(
            f"activity log error: {exc}"
        )


# =========================================================
# GOOGLE SHEETS
# =========================================================

def get_gsheet_client():

    if gspread is None or Credentials is None:
        raise RuntimeError(
            "Google Sheets libraries unavailable"
        )

    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=[
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive",
        ],
    )

    return gspread.authorize(creds)


def extract_sheet_id(link):

    match = re.search(
        r"/spreadsheets/d/([a-zA-Z0-9-_]+)",
        str(link)
    )

    return match.group(1) if match else None


def extract_gid(link):

    match = re.search(
        r"[?#&]gid=(\d+)",
        str(link)
    )

    return int(match.group(1)) if match else None


def read_gsheet_rows(link):

    sheet_id = extract_sheet_id(link)

    if not sheet_id:
        raise ValueError(
            "رابط Google Sheets غير صحيح."
        )

    spreadsheet = (
        get_gsheet_client()
        .open_by_key(sheet_id)
    )

    gid = extract_gid(link)

    worksheet = (
        next(
            (
                w
                for w in spreadsheet.worksheets()
                if w.id == gid
            ),
            spreadsheet.sheet1
        )
        if gid is not None
        else spreadsheet.sheet1
    )

    return worksheet.get_all_values()


def get_saved_gsheet_link(office_id):

    try:

        rows = (
            db()
            .table("data_sources")
            .select(
                "source_url,created_at"
            )
            .eq(
                "office_id",
                office_id
            )
            .eq(
                "source_type",
                "google_sheet"
            )
            .not_.is_(
                "source_url",
                "null"
            )
            .order(
                "created_at",
                desc=True
            )
            .limit(1)
            .execute()
            .data
            or []
        )

        return (
            rows[0].get("source_url")
            if rows
            else None
        )

    except Exception:
        return None


def save_gsheet_link(office_id, link):

    try:

        if not extract_sheet_id(link):
            return False, "الرابط غير صحيح!"

        (
            db()
            .table("data_sources")
            .insert({
                "office_id": office_id,
                "source_type": "google_sheet",
                "source_name": "Google Sheet",
                "source_url": link,
                "column_mapping": {},
            })
            .execute()
        )

        return (
            True,
            "تم حفظ الرابط بنجاح"
        )

    except Exception:

        safe_log(
            "Google Sheet link save failed"
        )

        return (
            False,
            "تعذر حفظ الرابط حاليًا."
        )


# =========================================================
# EXCEL
# =========================================================

def find_excel_columns(ws):

    cols = {
        "name": None,
        "email": None,
        "password": None
    }

    header_row = None

    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=min(10, ws.max_row),
            values_only=True
        ),
        start=1
    ):

        values = [
            str(c).strip()
            if c is not None
            else ""
            for c in row
        ]

        if any(
            "يميل" in v
            or "mail" in v.lower()
            or "بريد" in v
            for v in values
        ):

            header_row = row_idx

            for i, cell in enumerate(values):

                low = cell.lower()

                if (
                    any(
                        k in cell
                        for k in [
                            "اسم",
                            "الإسم",
                            "الاسم"
                        ]
                    )
                    or "name" in low
                ):

                    cols["name"] = i

                elif (
                    any(
                        k in cell
                        for k in [
                            "يميل",
                            "بريد"
                        ]
                    )
                    or "mail" in low
                ):

                    cols["email"] = i

                elif any(
                    k in cell
                    for k in [
                        "باسورد",
                        "كلمة المرور",
                        "password",
                        "pass"
                    ]
                ):

                    cols["password"] = i

            break

    if header_row is None:
        raise ValueError(
            "مش لاقي هيدر الإكسيل."
        )

    if cols["email"] is None:
        raise ValueError(
            "مش لاقي عمود الإيميل."
        )

    if cols["password"] is None:
        raise ValueError(
            "مش لاقي عمود الباسورد."
        )

    if cols["name"] is None:
        cols["name"] = cols["email"]

    return cols, header_row


def parse_excel_bytes(file_bytes):

    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        data_only=False
    )

    ws = wb.active

    cols, header_row = find_excel_columns(ws)

    records = []

    seen = set()

    for excel_row, row in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            values_only=True
        ),
        start=header_row + 1
    ):

        values = list(row)

        email = (
            str(
                values[cols["email"]] or ""
            ).strip()
            if cols["email"] < len(values)
            else ""
        )

        password = (
            str(
                values[cols["password"]] or ""
            ).strip()
            if cols["password"] < len(values)
            else ""
        )

        name = (
            str(
                values[cols["name"]] or ""
            ).strip()
            if cols["name"] < len(values)
            else email
        )

        key = email.lower()

        if (
            not email
            or not password
            or key in seen
        ):
            continue

        seen.add(key)

        records.append({
            "source_row_number": excel_row,
            "student_name": name or email,
            "login_identifier": email,
            "password": password,
            "original_data": {
                f"column_{i+1}":
                    (
                        str(v)
                        if v is not None
                        else ""
                    )
                for i, v in enumerate(values)
            },
        })

    return records


def encrypt_password(password, key):

    if not key:
        raise RuntimeError(
            "Encryption key missing"
        )

    return (
        Fernet(key.encode())
        .encrypt(password.encode())
        .decode()
    )


# =========================================================
# GOOGLE DRIVE
# =========================================================

def get_google_credentials(scopes):

    if Credentials is None:
        raise RuntimeError(
            "Google credentials libraries unavailable"
        )

    creds_dict = st.secrets.get(
        "gcp_service_account"
    )

    if isinstance(creds_dict, dict):

        return (
            Credentials
            .from_service_account_info(
                creds_dict,
                scopes=scopes
            )
        )

    raw = st.secrets.get(
        "GCP_SERVICE_ACCOUNT_JSON",
        os.getenv("GCP_SERVICE_ACCOUNT_JSON")
    )

    if raw:

        return (
            Credentials
            .from_service_account_info(
                json.loads(raw),
                scopes=scopes
            )
        )

    raise RuntimeError(
        "Google service account configuration missing"
    )


def drive_service():

    return build(
        "drive",
        "v3",
        credentials=get_google_credentials(
            [
                "https://www.googleapis.com/auth/drive"
            ]
        )
    )


def upload_to_drive(
    file_bytes,
    filename,
    office
):

    service = drive_service()

    drive_filename = str(
        filename or "students.xlsx"
    )

    metadata = {
        "name": drive_filename,
        "parents": [DRIVE_FOLDER_ID]
    }

    media = MediaIoBaseUpload(
        io.BytesIO(file_bytes),
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        resumable=True,
    )

    return (
        service
        .files()
        .create(
            body=metadata,
            media_body=media,
            fields="id"
        )
        .execute()["id"]
    )


def download_drive_file_bytes(file_id):

    service = drive_service()

    buffer = io.BytesIO()

    request = (
        service
        .files()
        .get_media(
            fileId=str(file_id)
        )
    )

    downloader = MediaIoBaseDownload(
        buffer,
        request
    )

    done = False

    while not done:

        _, done = downloader.next_chunk()

    buffer.seek(0)

    return buffer.getvalue()


# =========================================================
# STATUS COLUMN
# =========================================================

def find_status_column_for_output(
    ws,
    header_row
):

    for col_idx, cell in enumerate(
        ws[header_row],
        start=1
    ):

        value = str(
            cell.value or ""
        ).strip().lower()

        if (
            value in {
                "حالة الطلب",
                "الحالة"
            }
            or (
                "حالة" in value
                and "اسم" not in value
                and "خدمة" not in value
            )
        ):

            return col_idx

    new_col = ws.max_column + 1
    ws.cell(header_row, new_col).value = "حالة الطلب"
    return new_col


# =========================================================
# OLD BULK EXCEL BUILDER
# KEPT AS REQUESTED
# =========================================================

def build_updated_excel(
    file_bytes,
    students
):

    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        data_only=False
    )

    ws = wb.active

    cols, header_row = find_excel_columns(ws)

    status_col = (
        find_status_column_for_output(
            ws,
            header_row
        )
    )

    by_login = {}

    by_row = {}

    for student in students:

        login = str(
            student.get(
                "login_identifier"
            ) or ""
        ).strip().lower()

        status = str(
            student.get(
                "application_status"
            ) or ""
        ).strip()

        if login and status:
            by_login[login] = status

        if (
            student.get(
                "source_row_number"
            )
            and status
        ):
            by_row[
                int(
                    student[
                        "source_row_number"
                    ]
                )
            ] = status

    for row_idx in range(
        header_row + 1,
        ws.max_row + 1
    ):

        status = None

        email = str(
            ws.cell(
                row_idx,
                cols["email"] + 1
            ).value or ""
        ).strip().lower()

        if email:
            status = by_login.get(email)

        if status is None:
            status = by_row.get(row_idx)

        if status is not None:

            ws.cell(
                row_idx,
                status_col
            ).value = status

    output = io.BytesIO()

    wb.save(output)

    return output.getvalue()


# =========================================================
# OLD BULK GOOGLE SHEET UPDATE
# KEPT AS REQUESTED
# =========================================================

def update_google_sheet_statuses(
    source_url,
    students,
    job_id=None
):

    sheet_id = extract_sheet_id(
        source_url
    )

    if not sheet_id:
        raise RuntimeError(
            "invalid_google_sheet_url"
        )

    spreadsheet = (
        get_gsheet_client()
        .open_by_key(sheet_id)
    )

    gid = extract_gid(
        source_url
    )

    worksheet = (
        next(
            (
                w
                for w in spreadsheet.worksheets()
                if w.id == gid
            ),
            spreadsheet.sheet1
        )
        if gid is not None
        else spreadsheet.sheet1
    )

    values = worksheet.get_all_values()

    header_idx = None
    email_idx = None
    status_idx = None

    for r_idx, row in enumerate(
        values[:10]
    ):

        normalized = [
            str(v or "")
            .strip()
            .lower()
            for v in row
        ]

        if any(
            "يميل" in v
            or "mail" in v
            or "بريد" in v
            for v in normalized
        ):

            header_idx = r_idx

            for i, value in enumerate(
                normalized
            ):

                if (
                    "يميل" in value
                    or "mail" in value
                    or "بريد" in value
                ):

                    email_idx = i

                if (
                    value in {
                        "حالة الطلب",
                        "الحالة"
                    }
                    or (
                        "حالة" in value
                        and "اسم" not in value
                        and "خدمة" not in value
                    )
                ):

                    status_idx = i

            break

    if (
        header_idx is None
        or email_idx is None
        or status_idx is None
    ):

        raise RuntimeError(
            "google_sheet_columns_missing"
        )

    by_login = {
        str(
            s.get(
                "login_identifier"
            ) or ""
        ).strip().lower():
            str(
                s.get(
                    "application_status"
                ) or ""
            ).strip()
        for s in students
    }

    for row_idx in range(
        header_idx + 1,
        len(values)
    ):

        # لو المكتب سجل خروج أثناء تحديث الشيت
        # نوقف فورًا وما نكملش باقي الصفوف.
        if (
            job_id
            and not job_is_active(job_id)
        ):

            safe_log(
                f"Job {job_id} cancelled "
                f"during Google Sheet finalization."
            )

            return

        login = str(
            values[row_idx][email_idx]
            if email_idx < len(
                values[row_idx]
            )
            else ""
        ).strip().lower()

        status = by_login.get(login)

        if status:

            worksheet.update_cell(
                row_idx + 1,
                status_idx + 1,
                status
            )


# =========================================================
# OLD FINALIZATION
# KEPT BUT NO LONGER CALLED BY THE LIVE WORKER
# =========================================================

def finalize_job_output(job):

    job_id = job["id"]

    # Never start finalization if the job was cancelled.
    if not job_is_active(job_id):

        safe_log(
            f"Job {job_id} cancelled before finalization."
        )

        return None

    students = get_students_for_job(
        job_id
    )

    source_type = str(
        job.get("source_type") or ""
    )

    if source_type == "excel":

        source_ref = str(
            job.get("source_ref") or ""
        ).strip()

        if not source_ref:
            raise RuntimeError(
                "excel_source_missing"
            )

        # Check again before downloading the original file.
        if not job_is_active(job_id):
            return None

        original = download_drive_file_bytes(
            source_ref
        )

        # Check again before building/uploading the final file.
        if not job_is_active(job_id):
            return None

        updated = build_updated_excel(
            original,
            students
        )

        if not job_is_active(job_id):
            return None

        final_id = upload_to_drive(
            updated,
            job.get("file_name")
            or "students.xlsx",
            ""
        )

        # The job could have been cancelled while upload was happening.
        if not job_is_active(job_id):

            safe_log(
                f"Job {job_id} cancelled "
                f"after final file upload."
            )

            return final_id

        (
            db()
            .table("jobs")
            .update({
                "final_drive_file_id": final_id,
                "error": None
            })
            .eq("id", job_id)
            .eq("status", "processing")
            .execute()
        )

        return final_id

    if source_type == "google_sheet":

        source_url = str(
            job.get("source_ref") or ""
        )

        if not job_is_active(job_id):
            return None

        update_google_sheet_statuses(
            source_url,
            students,
            job_id=job_id
        )

        return None

    raise RuntimeError(
        f"unsupported_source_type:{source_type}"
    )


# =========================================================
# IMPORT STUDENTS
# =========================================================

def import_students(
    office_id,
    source_type,
    source_name,
    file_bytes=None,
    source_url=None
):

    if source_type in {
        "xlsx",
        "xls",
        "excel"
    }:

        records = parse_excel_bytes(
            file_bytes
        )

    else:

        rows = read_gsheet_rows(
            source_url
        )

        if not rows:
            raise ValueError(
                "الشيت فاضي."
            )

        wb = openpyxl.Workbook()

        ws = wb.active

        for row in rows:
            ws.append(row)

        out = io.BytesIO()

        wb.save(out)

        records = parse_excel_bytes(
            out.getvalue()
        )

    if not records:
        raise ValueError(
            "مش لاقي طلاب عندهم إيميل وباسورد صالحين."
        )

    encryption_key = st.secrets.get(
        "STUDENT_PASSWORD_ENCRYPTION_KEY",
        os.getenv(
            "STUDENT_PASSWORD_ENCRYPTION_KEY"
        )
    )

    source_type = (
        "google_sheet"
        if source_type in {
            "gsheet",
            "google_sheet"
        }
        else "excel"
    )

    file_path = None

    if source_type == "excel":

        if not file_bytes:
            raise ValueError(
                "ملف Excel غير موجود."
            )

        file_path = upload_to_drive(
            file_bytes,
            source_name,
            ""
        )

    source = (
        db()
        .table("data_sources")
        .insert({
            "office_id": office_id,
            "source_type": source_type,
            "source_name": source_name,
            "source_url": source_url,
            "file_path": file_path,
            "column_mapping": {},
        })
        .execute()
        .data[0]
    )

    payload = [
        {
            "office_id": office_id,
            "data_source_id": source["id"],
            "source_row_number": r[
                "source_row_number"
            ],
            "student_name": r[
                "student_name"
            ],
            "login_identifier": r[
                "login_identifier"
            ],
            "encrypted_password":
                encrypt_password(
                    r["password"],
                    encryption_key
                ),
            "application_status": "",
            "original_data": r[
                "original_data"
            ],
            "updated_at": now_iso(),
        }
        for r in records
    ]

    (
        db()
        .table("student_records")
        .insert(payload)
        .execute()
    )

    return source, len(payload)


# =========================================================
# LIVE EXCEL UPDATE
# SAME DRIVE FILE
# =========================================================

def update_excel_student_status(
    source_ref,
    student,
    status
):

    if not source_ref:
        raise RuntimeError(
            "excel_source_missing"
        )

    file_bytes = download_drive_file_bytes(
        source_ref
    )

    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        data_only=False
    )

    ws = wb.active

    cols, header_row = find_excel_columns(
        ws
    )

    status_col = (
        find_status_column_for_output(
            ws,
            header_row
        )
    )

    source_row = student.get(
        "source_row_number"
    )

    if not source_row:

        login = str(
            student.get(
                "login_identifier"
            ) or ""
        ).strip().lower()

        for row_idx in range(
            header_row + 1,
            ws.max_row + 1
        ):

            email = str(
                ws.cell(
                    row_idx,
                    cols["email"] + 1
                ).value or ""
            ).strip().lower()

            if email == login:

                source_row = row_idx

                break

    if not source_row:
        raise RuntimeError(
            "excel_student_row_missing"
        )

    ws.cell(
        int(source_row),
        status_col
    ).value = status

    output = io.BytesIO()

    wb.save(output)

    updated_bytes = output.getvalue()

    service = drive_service()

    media = MediaIoBaseUpload(
        io.BytesIO(updated_bytes),
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        resumable=True,
    )

    (
        service
        .files()
        .update(
            fileId=str(source_ref),
            media_body=media
        )
        .execute()
    )

    safe_log(
        f"Live Excel update: "
        f"{student.get('student_name')} "
        f"-> {status}"
    )


# =========================================================
# LIVE GOOGLE SHEET UPDATE
# SAME SHEET
# =========================================================

def update_google_sheet_student_status(
    source_url,
    student,
    status
):

    sheet_id = extract_sheet_id(
        source_url
    )

    if not sheet_id:
        raise RuntimeError(
            "invalid_google_sheet_url"
        )

    spreadsheet = (
        get_gsheet_client()
        .open_by_key(sheet_id)
    )

    gid = extract_gid(
        source_url
    )

    worksheet = (
        next(
            (
                w
                for w in spreadsheet.worksheets()
                if w.id == gid
            ),
            spreadsheet.sheet1
        )
        if gid is not None
        else spreadsheet.sheet1
    )

    values = worksheet.get_all_values()

    header_idx = None
    email_idx = None
    status_idx = None

    for r_idx, row in enumerate(
        values[:10]
    ):

        normalized = [
            str(v or "")
            .strip()
            .lower()
            for v in row
        ]

        if any(
            "يميل" in v
            or "mail" in v
            or "بريد" in v
            for v in normalized
        ):

            header_idx = r_idx

            for i, value in enumerate(
                normalized
            ):

                if (
                    "يميل" in value
                    or "mail" in value
                    or "بريد" in value
                ):
                    email_idx = i

                if (
                    value in {
                        "حالة الطلب",
                        "الحالة"
                    }
                    or (
                        "حالة" in value
                        and "اسم" not in value
                        and "خدمة" not in value
                    )
                ):
                    status_idx = i

            break

    if (
        header_idx is None
        or email_idx is None
        or status_idx is None
    ):

        raise RuntimeError(
            "google_sheet_columns_missing"
        )

    target_row = None

    login = str(
        student.get(
            "login_identifier"
        ) or ""
    ).strip().lower()

    source_row = student.get(
        "source_row_number"
    )

    if source_row:

        candidate = int(
            source_row
        )

        if (
            candidate > header_idx
            and candidate <= len(values)
        ):

            candidate_login = str(
                values[candidate - 1][email_idx]
                if email_idx < len(
                    values[candidate - 1]
                )
                else ""
            ).strip().lower()

            if candidate_login == login:
                target_row = candidate

    if target_row is None:

        for row_idx in range(
            header_idx + 1,
            len(values)
        ):

            row = values[row_idx]

            row_login = str(
                row[email_idx]
                if email_idx < len(row)
                else ""
            ).strip().lower()

            if row_login == login:

                target_row = row_idx + 1

                break

    if target_row is None:
        raise RuntimeError(
            "google_sheet_student_row_missing"
        )

    worksheet.update_cell(
        target_row,
        status_idx + 1,
        status
    )

    safe_log(
        f"Live Google Sheet update: "
        f"{student.get('student_name')} "
        f"-> {status}"
    )


# =========================================================
# LIVE SOURCE UPDATE DISPATCHER
# =========================================================

def update_source_student_status(
    job,
    student,
    status
):

    source_type = str(
        job.get("source_type") or ""
    )

    if source_type == "excel":

        update_excel_student_status(
            job.get("source_ref"),
            student,
            status
        )

    elif source_type == "google_sheet":

        update_google_sheet_student_status(
            job.get("source_ref"),
            student,
            status
        )

    else:

        raise RuntimeError(
            f"unsupported_source_type:{source_type}"
        )


# =========================================================
# JOB
# =========================================================

def create_job(
    office_id,
    source,
    file_name
):

    return (
        db()
        .table("jobs")
        .insert({
            "office_id": office_id,
            "data_source_id": source["id"],
            "source_type": source["source_type"],
            "source_ref":
                source.get("file_path")
                or source.get("source_url")
                or source["id"],
            "file_name": file_name,
            "status": "pending",
        })
        .execute()
        .data[0]
    )


def get_job(job_id):

    if not job_id:
        return None

    rows = (
        db()
        .table("jobs")
        .select("*")
        .eq("id", job_id)
        .limit(1)
        .execute()
        .data
        or []
    )

    return rows[0] if rows else None


def cancel_job(
    job_id,
    office_id
):

    if not job_id or not office_id:
        return

    try:

        (
            db()
            .table("jobs")
            .update({
                "status": "cancelled",
                "finished_at": now_iso(),
                "error": "office_logout"
            })
            .eq("id", job_id)
            .eq("office_id", office_id)
            .in_(
                "status",
                [
                    "pending",
                    "processing"
                ]
            )
            .execute()
        )

        safe_log(
            f"Job {job_id} cancelled by office logout."
        )

    except Exception as exc:

        safe_log(
            f"cancel job failed: {exc}"
        )

def cancel_all_office_jobs(office_id):

    if not office_id:
        return False

    try:

        # =====================================================
        # CANCEL ALL ACTIVE JOBS FOR THIS OFFICE
        # =====================================================

        response = (
            db()
            .table("jobs")
            .update({
                "status": "cancelled",
                "finished_at": now_iso(),
                "error": "office_logout"
            })
            .eq(
                "office_id",
                office_id
            )
            .in_(
                "status",
                [
                    "pending",
                    "processing"
                ]
            )
            .select("id,status")
            .execute()
        )

        cancelled = response.data or []

        safe_log(
            f"Office {office_id}: "
            f"cancelled {len(cancelled)} active job(s) on logout."
        )

        # =====================================================
        # VERIFY:
        # Make sure there is NO pending/processing job left.
        # =====================================================

        remaining = (
            db()
            .table("jobs")
            .select("id,status")
            .eq(
                "office_id",
                office_id
            )
            .in_(
                "status",
                [
                    "pending",
                    "processing"
                ]
            )
            .execute()
            .data
            or []
        )

        if remaining:

            safe_log(
                f"WARNING: Office {office_id} still has "
                f"{len(remaining)} active job(s) after logout."
            )

            # Retry once in case another worker/request
            # changed the job during the first update.
            (
                db()
                .table("jobs")
                .update({
                    "status": "cancelled",
                    "finished_at": now_iso(),
                    "error": "office_logout"
                })
                .eq(
                    "office_id",
                    office_id
                )
                .in_(
                    "status",
                    [
                        "pending",
                        "processing"
                    ]
                )
                .execute()
            )

            # Verify one final time.
            remaining = (
                db()
                .table("jobs")
                .select("id,status")
                .eq(
                    "office_id",
                    office_id
                )
                .in_(
                    "status",
                    [
                        "pending",
                        "processing"
                    ]
                )
                .execute()
                .data
                or []
            )

        if remaining:

            safe_log(
                f"ERROR: Could not cancel all active jobs "
                f"for office {office_id}."
            )

            return False

        safe_log(
            f"Office {office_id}: "
            f"logout cancellation verified successfully."
        )

        return True

    except Exception as exc:

        safe_log(
            f"cancel all office jobs failed: "
            f"{type(exc).__name__}: {exc}"
        )

        return False
def get_job_progress_rows(job_id):

    rows = (
        db()
        .table("job_progress")
        .select(
            "student_index,total,"
            "student_name,status,created_at"
        )
        .eq(
            "job_id",
            job_id
        )
        .order("created_at")
        .execute()
        .data
        or []
    )

    latest = {}

    for row in rows:

        key = str(
            row.get(
                "student_name"
            ) or ""
        ).strip().lower()

        if key:
            latest[key] = row

    return sorted(
        latest.values(),
        key=lambda r:
            int(
                r.get(
                    "student_index"
                ) or 0
            )
    )


def get_students_for_job(job_id):

    job = get_job(job_id)

    if not job:
        return []

    rows = (
        db()
        .table("student_records")
        .select(
            "id,student_name,"
            "login_identifier,"
            "application_status,"
            "status_updated_at,"
            "source_row_number,"
            "created_at,updated_at"
        )
        .eq(
            "office_id",
            job["office_id"]
        )
        .eq(
            "data_source_id",
            job["data_source_id"]
        )
        .order(
            "source_row_number"
        )
        .execute()
        .data
        or []
    )

    latest = {}

    for row in rows:

        key = str(
            row.get(
                "login_identifier"
            )
            or row.get(
                "student_name"
            )
            or ""
        ).strip().lower()

        if not key:
            continue

        latest[key] = row

    return list(
        latest.values()
    )


def get_students(
    office_id,
    search=""
):

    rows = (
        db()
        .table("student_records")
        .select(
            "id,student_name,"
            "login_identifier,"
            "application_status,"
            "status_updated_at,"
            "source_row_number,"
            "created_at,updated_at"
        )
        .eq(
            "office_id",
            office_id
        )
        .execute()
        .data
        or []
    )

    def stamp(v):

        try:

            return datetime.fromisoformat(
                str(v or "")
                .replace(
                    "Z",
                    "+00:00"
                )
            )

        except Exception:

            return datetime.min.replace(
                tzinfo=timezone.utc
            )

    latest = {}

    for row in rows:

        key = str(
            row.get(
                "login_identifier"
            )
            or row.get(
                "student_name"
            )
            or ""
        ).strip().lower()

        if not key:
            continue

        score = (
            stamp(
                row.get(
                    "status_updated_at"
                )
            ),
            stamp(
                row.get(
                    "updated_at"
                )
            ),
            stamp(
                row.get(
                    "created_at"
                )
            )
        )

        if (
            key not in latest
            or score > latest[key][0]
        ):

            latest[key] = (
                score,
                row
            )

    rows = sorted(
        [
            x[1]
            for x in latest.values()
        ],
        key=lambda r:
            str(
                r.get(
                    "student_name"
                ) or ""
            ).lower()
    )

    q = search.strip().lower()

    return [
        r
        for r in rows
        if (
            not q
            or q in str(
                r.get(
                    "student_name"
                ) or ""
            ).lower()
        )
    ]


# =========================================================
# API FALLBACK
# =========================================================

def _legacy_api_login(
    email,
    password
):

    session = requests.Session()

    session.headers.update({
        "accept":
            "application/json, text/plain, */*",
        "accept-language":
            "ar",
        "device":
            "CITIZEN",
        "origin":
            SITE_URL,
        "referer":
            SITE_URL + "/",
        "user-agent":
            "Mozilla/5.0",
        "content-type":
            "application/json",
    })

    response = session.post(
        f"{BASE_URL}/student/login",
        json={
            "email": email,
            "password": password
        },
        timeout=30
    )

    if response.status_code not in (
        200,
        201
    ):

        return (
            None,
            None,
            "login_failed"
        )

    body = (
        response.json()
        if response.content
        else {}
    )

    return (
        session,
        body.get("token", "")
        or response.headers.get(
            "x-csrf-token",
            ""
        ),
        None
    )


def _legacy_api_get_status(
    session,
    token
):

    filt = {
        "where": {},
        "limit": 10,
        "offset": 0,
        "order":
            "statusUpdatedAt DESC",
        "fields": [
            "serviceSlug",
            "ID",
            "createdAt",
            "statusUpdatedAt",
            "activityId",
            "activityName"
        ]
    }

    headers = (
        {"x-csrf-token": token}
        if token
        else {}
    )

    response = session.get(
        f"{BASE_URL}/dynamic_services/inbox",
        params={
            "filter":
                json.dumps(filt)
        },
        headers=headers,
        timeout=30
    )

    if response.status_code not in (
        200,
        304
    ):

        raise RuntimeError(
            f"status_http_{response.status_code}"
        )

    result = (
        response
        .json()
        .get("result")
        or []
    )

    if not result:
        return "مفيش طلبات"

    activity = (
        result[0].get(
            "activityName"
        )
        or "غير محدد"
    )

    mapping = {
        "قبول الفحص الفنى":
            "القبول المبدئي",

        "قبول الفحص الفني":
            "القبول المبدئي",

        "تم السداد":
            "تم السداد",

        "تأكيد استلام الملف وصحة و اكتمال المستندات":
            "تأكيد استلام الملف وصحة واكتمال المستندات",

        "الانتظار مراجعة الطلب":
            "بانتظار مراجعة الطلب",

        "قبول من رئيس الادارة المركزية":
            "قبول من رئيس الإدارة المركزية",
    }

    return mapping.get(
        activity,
        activity
    )


def _legacy_api_logout(
    session
):

    if session is None:
        return

    try:

        session.post(
            f"{BASE_URL}/student/logout",
            json={
                "redirectUrl":
                    SITE_URL
            },
            timeout=15
        )

    except Exception:
        pass


# =========================================================
# JOB STATUS
# =========================================================

def job_is_active(job_id):

    try:

        job = get_job(job_id)

        if not job:
            return False

        return (
            str(
                job.get(
                    "status"
                ) or ""
            )
            == "processing"
        )

    except Exception as exc:

        safe_log(
            f"job activity check failed: "
            f"{type(exc).__name__}: {exc}"
        )

        return False


def job_is_cancelled(job_id):

    try:

        job = get_job(job_id)

        if not job:
            return False

        return (
            str(
                job.get(
                    "status"
                ) or ""
            )
            == "cancelled"
        )

    except Exception as exc:

        safe_log(
            f"job cancellation check failed: "
            f"{type(exc).__name__}: {exc}"
        )

        return False


def mark_job_done_if_active(
    job_id
):

    stamp = now_iso()

    response = (
        db()
        .table("jobs")
        .update({
            "status": "done",
            "finished_at": stamp,
            "error": None
        })
        .eq(
            "id",
            job_id
        )
        .eq(
            "status",
            "processing"
        )
        .select(
            "id,status"
        )
        .execute()
    )

    rows = response.data or []

    return bool(rows)


def mark_job_failed_if_active(
    job_id,
    error
):

    stamp = now_iso()

    response = (
        db()
        .table("jobs")
        .update({
            "status": "failed",
            "finished_at": stamp,
            "error": str(error)[:1000],
        })
        .eq(
            "id",
            job_id
        )
        .eq(
            "status",
            "processing"
        )
        .select(
            "id,status"
        )
        .execute()
    )

    rows = response.data or []

    return bool(rows)


# =========================================================
# FALLBACK JOB
# =========================================================

def _claim_fallback_job(
    job_id
):

    rows = (
        db()
        .table("jobs")
        .update({
            "status":
                "processing",
            "started_at":
                now_iso(),
            "claimed_by":
                "streamlit-fallback"
        })
        .eq(
            "id",
            job_id
        )
        .eq(
            "status",
            "pending"
        )
        .select("*")
        .execute()
        .data
        or []
    )

    return rows[0] if rows else None


def _run_legacy_api_fallback(
    job
):

    """
    Run the student update job.

    IMPORTANT LIVE UPDATE RULES:

    1. Students are shuffled.
    2. Every processed student is logged out.
    3. Current student is always persisted after its operation.
    4. If office logs out while current student is being processed,
       current student is saved, then the job stops.
    5. Students not yet started are never changed.
    6. Excel / Google Sheet are updated immediately per student.
    7. No finalization pass is executed.
    """

    job_id = job["id"]

    try:

        client = db()

        key = os.getenv(
            "STUDENT_PASSWORD_ENCRYPTION_KEY"
        )

        if not key:

            # Also support Streamlit secrets.
            key = st.secrets.get(
                "STUDENT_PASSWORD_ENCRYPTION_KEY",
                ""
            )

        if not key:
            raise RuntimeError(
                "encryption_key_missing"
            )

        # =====================================================
        # INITIAL JOB CHECK
        # =====================================================

        if not job_is_active(job_id):

            safe_log(
                f"Job {job_id} is no longer active. "
                f"Worker stopped."
            )

            return

        # =====================================================
        # GET STUDENTS
        # =====================================================

        students = (
            client
            .table("student_records")
            .select("*")
            .eq(
                "office_id",
                job["office_id"]
            )
            .eq(
                "data_source_id",
                job["data_source_id"]
            )
            .order(
                "source_row_number"
            )
            .execute()
            .data
            or []
        )

        unique = {}

        for student in students:

            key_id = str(
                student.get(
                    "login_identifier"
                )
                or student.get(
                    "student_name"
                )
                or ""
            ).strip().lower()

            if key_id:
                unique[key_id] = student

        students = list(
            unique.values()
        )

        # =====================================================
        # VERY IMPORTANT:
        # KEEP RANDOM SHUFFLE
        # =====================================================

        import random

        random.shuffle(students)

        total = len(students)

        # =====================================================
        # NO STUDENTS
        # =====================================================

        if not total:

            if not job_is_active(job_id):
                return

            if mark_job_done_if_active(
                job_id
            ):

                safe_log(
                    f"Job {job_id} completed: "
                    f"no students."
                )

            return

        retry_students = []

        # =========================================================
        # FIRST PASS
        # =========================================================

        for index, student in enumerate(
            students,
            1
        ):

            # -----------------------------------------------------
            # STOP BEFORE STARTING A NEW STUDENT
            #
            # This is the important boundary.
            #
            # If logout happened during the previous student,
            # that previous student has already been saved.
            #
            # We stop HERE before touching this student.
            # -----------------------------------------------------

            if not job_is_active(
                job_id
            ):

                safe_log(
                    f"Job {job_id} stopped before "
                    f"starting student {index}."
                )

                return

            name = str(
                student.get(
                    "student_name"
                )
                or student.get(
                    "login_identifier"
                )
                or "طالب"
            ).strip()

            current = str(
                student.get(
                    "application_status"
                )
                or ""
            ).strip()

            status = (
                current
                or "لم يتم الفحص بعد"
            )

            session = None

            try:

                # -------------------------------------------------
                # FINAL STATUS
                # -------------------------------------------------

                if current in FINAL_STATUSES:

                    status = current

                else:

                    password = (
                        Fernet(
                            key.encode()
                        )
                        .decrypt(
                            str(
                                student[
                                    "encrypted_password"
                                ]
                            ).encode()
                        )
                        .decode()
                    )

                    # -------------------------------------------------
                    # Check before login
                    #
                    # If office logged out BEFORE we started this
                    # student, don't touch this student.
                    # -------------------------------------------------

                    if not job_is_active(
                        job_id
                    ):
                        return

                    fallback_delay(
                        LOGIN_PAGE_DELAY_MIN,
                        LOGIN_PAGE_DELAY_MAX
                    )

                    # -------------------------------------------------
                    # If logout happened during the preparation delay,
                    # no student operation has started yet.
                    # -------------------------------------------------

                    if not job_is_active(
                        job_id
                    ):
                        return

                    session, token, error = (
                        _legacy_api_login(
                            str(
                                student[
                                    "login_identifier"
                                ]
                            ).strip(),
                            password
                        )
                    )

                    if error:

                        status = (
                            "فشل تسجيل الدخول"
                        )

                    else:

                        time.sleep(
                            POST_LOGIN_DELAY_SECONDS
                        )

                        fallback_delay(
                            INBOX_DELAY_MIN,
                            INBOX_DELAY_MAX
                        )

                        # -------------------------------------------------
                        # IMPORTANT:
                        #
                        # DO NOT return here if logout happened.
                        #
                        # The student has already been logged in.
                        # We must finish this student's cleanup,
                        # logout, save the result, and then stop.
                        # -------------------------------------------------

                        if job_is_cancelled(
                            job_id
                        ):

                            safe_log(
                                f"Job {job_id} cancelled "
                                f"while student {index} "
                                f"was logged in."
                            )

                        else:

                            status = (
                                _legacy_api_get_status(
                                    session,
                                    token
                                )
                            )

            except Exception as exc:

                status = TECH_FAILURE_STATUS

                retry_students.append(
                    (
                        index,
                        student,
                        name
                    )
                )

                safe_log(
                    f"fallback student error "
                    f"{student.get('id')}: "
                    f"{type(exc).__name__}: {exc}"
                )

            finally:

                # =====================================================
                # ALWAYS LOGOUT CURRENT STUDENT
                # =====================================================

                _legacy_api_logout(
                    session
                )

                if session is not None:

                    fallback_delay(
                        INBOX_DELAY_MIN,
                        INBOX_DELAY_MAX
                    )

            # =========================================================
            # IMPORTANT:
            #
            # DO NOT CHECK job_is_active() HERE BEFORE SAVING.
            #
            # If office logged out during this student,
            # this student MUST still be saved.
            # =========================================================

            stamp = now_iso()

            # =========================================================
            # SAVE STUDENT RESULT IN SUPABASE
            # =========================================================

            try:

                (
                    client
                    .table("student_records")
                    .update({
                        "application_status":
                            status,
                        "status_updated_at":
                            stamp,
                        "updated_at":
                            stamp
                    })
                    .eq(
                        "id",
                        student["id"]
                    )
                    .execute()
                )

            except Exception as exc:

                safe_log(
                    f"fallback student DB persistence "
                    f"error for {name}: {exc}"
                )

            # =========================================================
            # LIVE UPDATE SOURCE
            #
            # This happens immediately for the current student.
            # No finalization later.
            # =========================================================

            try:

                update_source_student_status(
                    job,
                    student,
                    status
                )

            except Exception as exc:

                # Source update should not erase the DB result.
                safe_log(
                    f"live source update failed "
                    f"for {name}: "
                    f"{type(exc).__name__}: {exc}"
                )

            # =========================================================
            # SAVE PROGRESS
            # =========================================================

            try:

                (
                    client
                    .table("job_progress")
                    .insert({
                        "job_id":
                            job_id,
                        "student_index":
                            index,
                        "total":
                            total,
                        "student_name":
                            name,
                        "status":
                            status
                    })
                    .execute()
                )

            except Exception as exc:

                safe_log(
                    f"fallback progress persistence "
                    f"error for {name}: {exc}"
                )

            # =========================================================
            # NOW CHECK CANCELLATION
            #
            # If office logged out during current student:
            #
            # current student:
            #     DONE
            #     LOGGED OUT
            #     SAVED
            #     SOURCE UPDATED
            #
            # next student:
            #     NOT TOUCHED
            # =========================================================

            if not job_is_active(
                job_id
            ):

                safe_log(
                    f"Job {job_id} stopped after "
                    f"student {index}. "
                    f"Current student was saved."
                )

                return

            # =========================================================
            # DELAY BEFORE NEXT STUDENT
            # =========================================================

            if index < total:

                if not job_is_active(
                    job_id
                ):
                    return

                fallback_delay(
                    STUDENT_DELAY_MIN,
                    STUDENT_DELAY_MAX
                )

        # =========================================================
        # RETRY TECHNICAL FAILURES
        # =========================================================

        for retry_position, (
            index,
            student,
            name
        ) in enumerate(
            retry_students
        ):

            # -----------------------------------------------------
            # Office logged out → don't retry anything.
            # -----------------------------------------------------

            if not job_is_active(
                job_id
            ):

                safe_log(
                    f"Job {job_id} cancelled "
                    f"before retry."
                )

                return

            retry_status = (
                TECH_FAILURE_STATUS
            )

            session = None

            try:

                password = (
                    Fernet(
                        key.encode()
                    )
                    .decrypt(
                        str(
                            student[
                                "encrypted_password"
                            ]
                        ).encode()
                    )
                    .decode()
                )

                fallback_delay(
                    1.0,
                    2.0
                )

                if not job_is_active(
                    job_id
                ):
                    return

                session, token, error = (
                    _legacy_api_login(
                        str(
                            student[
                                "login_identifier"
                            ]
                        ).strip(),
                        password
                    )
                )

                if error:

                    retry_status = (
                        "فشل تسجيل الدخول"
                    )

                else:

                    time.sleep(
                        POST_LOGIN_DELAY_SECONDS
                    )

                    fallback_delay(
                        INBOX_DELAY_MIN,
                        INBOX_DELAY_MAX
                    )

                    # -------------------------------------------------
                    # If cancellation happened while retry student
                    # was already logged in, DO NOT return.
                    # We still logout and save current result.
                    # -------------------------------------------------

                    if job_is_cancelled(
                        job_id
                    ):

                        safe_log(
                            f"Job {job_id} cancelled "
                            f"while retry student "
                            f"{name} was logged in."
                        )

                    else:

                        retry_status = (
                            _legacy_api_get_status(
                                session,
                                token
                            )
                        )

            except Exception as exc:

                retry_status = (
                    TECH_FAILURE_STATUS
                )

                safe_log(
                    f"fallback retry error "
                    f"{student.get('id')}: "
                    f"{type(exc).__name__}: {exc}"
                )

            finally:

                # =====================================================
                # ALWAYS LOGOUT RETRY STUDENT
                # =====================================================

                _legacy_api_logout(
                    session
                )

                if session is not None:

                    fallback_delay(
                        INBOX_DELAY_MIN,
                        INBOX_DELAY_MAX
                    )

            # =========================================================
            # SAVE RETRY RESULT
            # EVEN IF JOB WAS CANCELLED
            # =========================================================

            stamp = now_iso()

            try:

                (
                    client
                    .table("student_records")
                    .update({
                        "application_status":
                            retry_status,
                        "status_updated_at":
                            stamp,
                        "updated_at":
                            stamp
                    })
                    .eq(
                        "id",
                        student["id"]
                    )
                    .execute()
                )

            except Exception as exc:

                safe_log(
                    f"fallback retry DB persistence "
                    f"error for {name}: {exc}"
                )

            # =========================================================
            # LIVE SOURCE UPDATE FOR RETRY
            # =========================================================

            try:

                update_source_student_status(
                    job,
                    student,
                    retry_status
                )

            except Exception as exc:

                safe_log(
                    f"live retry source update failed "
                    f"for {name}: "
                    f"{type(exc).__name__}: {exc}"
                )

            # =========================================================
            # SAVE RETRY PROGRESS
            # =========================================================

            try:

                (
                    client
                    .table("job_progress")
                    .insert({
                        "job_id":
                            job_id,
                        "student_index":
                            index,
                        "total":
                            total,
                        "student_name":
                            name,
                        "status":
                            retry_status
                    })
                    .execute()
                )

            except Exception as exc:

                safe_log(
                    f"fallback retry progress "
                    f"persistence error for {name}: {exc}"
                )

            # =========================================================
            # AFTER CURRENT RETRY IS COMPLETELY SAVED
            # STOP IF CANCELLED
            # =========================================================

            if not job_is_active(
                job_id
            ):

                safe_log(
                    f"Job {job_id} stopped after "
                    f"retry student {name}."
                )

                return

        # =========================================================
        # IMPORTANT:
        #
        # THERE IS NO FINALIZATION HERE.
        #
        # Excel / Google Sheet were already updated
        # student-by-student above.
        #
        # We only mark the job done if it is STILL processing.
        # =========================================================

        if not job_is_active(
            job_id
        ):

            safe_log(
                f"Job {job_id} cancelled "
                f"before completion."
            )

            return

        marked_done = (
            mark_job_done_if_active(
                job_id
            )
        )

        if marked_done:

            safe_log(
                f"Job {job_id} completed successfully."
            )

        else:

            # This can happen if the office cancelled
            # between the last check and the conditional update.
            current_job = get_job(
                job_id
            )

            if (
                current_job
                and str(
                    current_job.get(
                        "status"
                    ) or ""
                ) == "cancelled"
            ):

                safe_log(
                    f"Job {job_id} was cancelled "
                    f"before done update."
                )

            else:

                raise RuntimeError(
                    "job_could_not_be_marked_done"
                )

    except Exception as exc:

        safe_log(
            f"fallback failed: "
            f"{type(exc).__name__}: {exc}"
        )

        try:

            # =====================================================
            # NEVER TURN CANCELLED INTO FAILED
            # =====================================================

            current_job = get_job(
                job_id
            )

            if (
                current_job
                and str(
                    current_job.get(
                        "status"
                    ) or ""
                ) == "cancelled"
            ):

                safe_log(
                    f"Job {job_id} is cancelled. "
                    f"Keeping cancelled status."
                )

                return

            # Conditional update:
            # only processing -> failed
            marked_failed = (
                mark_job_failed_if_active(
                    job_id,
                    exc
                )
            )

            if not marked_failed:

                latest_job = get_job(
                    job_id
                )

                if (
                    latest_job
                    and str(
                        latest_job.get(
                            "status"
                        ) or ""
                    ) == "cancelled"
                ):

                    safe_log(
                        f"Job {job_id} became "
                        f"cancelled before failed update."
                    )

        except Exception as db_exc:

            safe_log(
                f"fallback job finalization failed: "
                f"{db_exc}"
            )


# =========================================================
# BACKGROUND STARTER
# =========================================================

def _background_update_job(
    job_id,
    encryption_key
):

    try:

        # Keep all Streamlit state/UI calls
        # out of the background thread.
        os.environ[
            "STUDENT_PASSWORD_ENCRYPTION_KEY"
        ] = encryption_key or ""

        deadline = (
            time.monotonic()
            + WORKER_WAIT_SECONDS
        )

        while (
            time.monotonic()
            < deadline
        ):

            job = get_job(
                job_id
            )

            if not job:
                return

            status = str(
                job.get(
                    "status"
                )
                or "pending"
            )

            # Office logged out / job cancelled.
            if status != "pending":
                return

            time.sleep(2)

        # Try to claim only if the job is still pending.
        claimed = _claim_fallback_job(
            job_id
        )

        if claimed:

            _run_legacy_api_fallback(
                claimed
            )

    except Exception as exc:

        safe_log(
            f"background update failed: "
            f"{type(exc).__name__}: {exc}"
        )

        try:

            current_job = get_job(
                job_id
            )

            # VERY IMPORTANT:
            # Never change a cancelled job into failed.
            if (
                current_job
                and str(
                    current_job.get(
                        "status"
                    ) or ""
                ) != "cancelled"
            ):

                (
                    db()
                    .table("jobs")
                    .update({
                        "status":
                            "failed",
                        "finished_at":
                            now_iso(),
                        "error":
                            str(exc)[:1000]
                    })
                    .eq(
                        "id",
                        job_id
                    )
                    .eq(
                        "status",
                        "processing"
                    )
                    .execute()
                )

        except Exception as db_exc:

            safe_log(
                f"background finalization failed: "
                f"{db_exc}"
            )


# =========================================================
# STREAMLIT ACTIONS
# =========================================================

def start_update():

    if st.session_state.get(
        "update_locked"
    ):
        return

    st.session_state.update_locked = True

    st.session_state.update_start_requested = True


def reset_session_on_logout():

    office = st.session_state.get(
        "office"
    )

    # =====================================================
    # CANCEL ALL ACTIVE JOBS FOR THIS OFFICE
    # =====================================================

    if office:

        office_id = office.get("id")

        if office_id:

            cancelled = cancel_all_office_jobs(
                office_id
            )

            if not cancelled:

                st.error(
                    "تعذر إيقاف التحديث بشكل آمن. "
                    "حاولي تسجيل الخروج مرة تانية."
                )

                return

    # =====================================================
    # CLEAR THIS STREAMLIT SESSION
    # =====================================================

    st.session_state.clear()

    st.rerun()


# =========================================================
# PAGE CONFIG
# =========================================================

st.set_page_config(
    page_title="Aivora - Agent",
    page_icon="✨",
    layout="wide",
    initial_sidebar_state="collapsed"
)


# =========================================================
# STYLE
# =========================================================

st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;500;600;700;800&display=swap');

html,body,[class*="css"],.stApp{
    font-family:'Cairo',sans-serif!important;
    direction:rtl;
    translate:no
}

.stApp,[data-testid="stAppViewContainer"]>.main{
    background:#f5f7fb;
    color:#111827
}

.block-container{
    max-width:1180px;
    padding-top:1.4rem;
    padding-bottom:3rem
}

#MainMenu,footer,header,[data-testid="stToolbar"],
[data-testid="stDecoration"],[data-testid="stStatusWidget"],
[data-testid="stSidebarNav"]{
    display:none!important
}

h1{
    font-size:32px!important;
    font-weight:800!important
}

h2{
    font-size:23px!important;
    font-weight:800!important
}

h3{
    font-size:19px!important;
    font-weight:700!important
}

.topbar{
    background:#fff;
    border:1px solid #e5e7eb;
    border-radius:16px;
    padding:13px 18px;
    display:flex;
    align-items:center;
    justify-content:space-between;
    box-shadow:0 2px 10px rgba(17,24,39,.04);
    margin-bottom:22px
}

.brand{
    display:flex;
    align-items:center;
    gap:11px
}

.brand-icon{
    width:42px;
    height:42px;
    border-radius:12px;
    background:#eff6ff;
    display:flex;
    align-items:center;
    justify-content:center;
    font-size:22px
}

.brand-title{
    font-size:18px;
    font-weight:800;
    color:#111827
}

.brand-sub{
    font-size:12px;
    color:#6b7280;
    margin-top:-2px
}

.card{
    background:#fff;
    border:1px solid #e5e7eb;
    border-radius:18px;
    padding:22px;
    box-shadow:0 3px 14px rgba(17,24,39,.045);
    margin-bottom:18px
}

.hero{
    background:linear-gradient(135deg,#fff 0%,#f8fbff 100%);
    border:1px solid #dbeafe;
    border-radius:20px;
    padding:25px 28px;
    box-shadow:0 4px 18px rgba(37,99,235,.06);
    margin-bottom:20px
}

.hero-kicker{
    color:#6b7280;
    font-size:14px;
    font-weight:600
}

.hero-title{
    color:#111827;
    font-size:28px;
    font-weight:800;
    margin-top:2px
}

.hero-title strong{
    color:#2563eb
}

.hero-desc{
    color:#6b7280;
    font-size:14px;
    margin-top:3px
}

.section-title{
    font-size:18px;
    font-weight:800;
    color:#111827;
    margin:5px 0 13px
}

.section-sub{
    color:#6b7280;
    font-size:13px;
    margin-top:-8px;
    margin-bottom:14px
}

.stTextInput label,.stFileUploader label,
.stRadio>label,.stCheckbox label{
    color:#374151!important;
    font-size:14px!important;
    font-weight:700!important
}

.stTextInput input{
    background:#fff!important;
    color:#111827!important;
    border:1px solid #d1d5db!important;
    border-radius:10px!important;
    font-size:14px!important;
    min-height:44px
}

.stTextInput input:focus{
    border-color:#2563eb!important;
    box-shadow:0 0 0 3px rgba(37,99,235,.10)!important
}

.stButton>button,.stDownloadButton>button{
    width:100%;
    min-height:44px;
    border-radius:10px!important;
    border:1px solid #2563eb!important;
    background:#2563eb!important;
    color:#fff!important;
    font-weight:700!important;
    font-size:14px!important;
    box-shadow:0 3px 8px rgba(37,99,235,.16)!important
}

.stButton>button:hover,
.stDownloadButton>button:hover{
    background:#1d4ed8!important;
    border-color:#1d4ed8!important;
    transform:translateY(-1px)
}

button[data-baseweb="tab"]{
    color:#6b7280!important;
    font-weight:700!important;
    font-size:14px!important
}

button[data-baseweb="tab"][aria-selected="true"]{
    color:#2563eb!important
}

[data-baseweb="tab-highlight"]{
    background:#2563eb!important;
    height:2px!important
}

.stRadio div[role="radiogroup"]{
    gap:10px
}

.stRadio div[role="radiogroup"] label{
    background:#fff;
    border:1px solid #e5e7eb;
    border-radius:12px;
    padding:10px 14px
}

[data-testid="stFileUploaderDropzone"]{
    background:#f8fafc!important;
    border:1.5px dashed #cbd5e1!important;
    border-radius:14px!important
}

[data-testid="InputInstructions"]{
    display:none!important
}

[data-testid="stTextInput"] button{
    display:none!important
}

.connected-box{
    background:#f0fdf4;
    border:1px solid #bbf7d0;
    border-radius:11px;
    padding:11px 14px;
    color:#166534;
    font-size:13px;
    margin-bottom:12px
}

.wait-box{
    background:#eff6ff;
    border:1px solid #bfdbfe;
    border-radius:12px;
    padding:13px 15px;
    color:#1d4ed8;
    font-weight:700
}

.success-box{
    background:#ecfdf5;
    border:1px solid #bbf7d0;
    border-radius:12px;
    padding:14px 15px;
    color:#166534;
    font-weight:800
}

.lock-box{
    background:#fff7ed;
    border:1px solid #fed7aa;
    border-radius:12px;
    padding:13px 15px;
    color:#9a3412
}

.progress-table{
    width:100%;
    border-collapse:separate;
    border-spacing:0;
    overflow:hidden;
    border:1px solid #e5e7eb;
    border-radius:12px
}

.progress-table th,
.progress-table td{
    padding:10px 12px;
    border-bottom:1px solid #eef0f4;
    font-size:13px
}

.progress-table th{
    background:#f8fafc;
    color:#6b7280;
    font-weight:800
}

.progress-table tr:last-child td{
    border-bottom:0
}

.progress-table .num{
    width:48px;
    text-align:center;
    direction:ltr
}

.progress-table .status{
    text-align:left;
    direction:rtl
}

.progress-table .name{
    text-align:right
}
</style>
""",
    unsafe_allow_html=True
)


# =========================================================
# SESSION DEFAULTS
# =========================================================

for key, default in [
    (
        "logged_in",
        False
    ),
    (
        "is_admin",
        False
    ),
    (
        "office",
        None
    ),
    (
        "update_locked",
        False
    ),
    (
        "active_job_id",
        None
    ),
    (
        "update_start_requested",
        False
    ),
    (
        "job_preparing",
        False
    ),
    (
        "pending_file_bytes",
        None
    ),
    (
        "pending_filename",
        ""
    ),
    (
        "final_file_bytes_cache",
        None
    ),
    (
        "final_file_cache_id",
        None
    ),
]:

    if key not in st.session_state:

        st.session_state[key] = default


# =========================================================
# LOGIN
# =========================================================

if (
    not st.session_state.logged_in
    and not st.session_state.is_admin
):

    st.markdown(
        """
<div style='text-align:center;margin:36px 0 22px'>
    <div style='font-size:46px'>✨</div>
    <div style='font-size:31px;font-weight:800'>Aivora</div>
    <div style='font-size:14px;color:#6b7280'>
        Your Smarter Support for Every Student's Application
    </div>
</div>
""",
        unsafe_allow_html=True
    )

    left, right = st.columns(
        [1.15, 1],
        gap="large"
    )

    with left:

        st.markdown(
            """
<div style="padding:40px 20px 20px 10px;">
    <div style="font-size:14px;color:#2563eb;font-weight:800;margin-bottom:8px;">
        حل بسيط لإدارة المتابعة
    </div>

    <div style="font-size:31px;font-weight:800;color:#111827;line-height:1.35;">
        تابع طلبات طلابك<br>من مكان واحد.
    </div>

    <div style="font-size:15px;color:#6b7280;line-height:1.9;margin-top:12px;max-width:480px;">
        حدّث حالات الطلبات، اربط Google Sheets، وابحث عن أي طالب بسرعة بدون متابعة يدوية.
    </div>

    <div style="margin-top:22px;color:#374151;font-size:14px;line-height:2.2;">
        ✓ تحديث حالات الطلاب بشكل منظم<br>
        ✓ حفظ مصدر البيانات للمكتب<br>
        ✓ بحث سريع عن حالة أي طالب
    </div>
</div>
""",
            unsafe_allow_html=True
        )

    with right:

        st.markdown(
            '<div class="card">',
            unsafe_allow_html=True
        )

        tab1, tab2 = st.tabs(
            [
                "تسجيل الدخول",
                "حساب جديد"
            ]
        )

        with tab1:

            st.markdown(
                """
<div style='font-size:22px;font-weight:800;color:#111827;margin:8px 0 3px;'>
    مرحبًا بعودتك 👋
</div>

<div style='color:#6b7280;font-size:13px;margin-bottom:18px;'>
    سجل دخولك لإدارة طلبات الطلاب
</div>
""",
                unsafe_allow_html=True
            )

            username = st.text_input(
                "اسم المكتب",
                key="login_user",
                placeholder="اكتب اسم المكتب"
            )

            password = st.text_input(
                "كلمة المرور",
                type="password",
                key="login_pass",
                placeholder="اكتب كلمة المرور"
            )

            if st.button(
                "تسجيل الدخول",
                key="login_btn"
            ):

                if (
                    username
                    == ADMIN_USERNAME
                    and password
                    == ADMIN_PASSWORD
                ):

                    st.session_state.is_admin = True

                    st.rerun()

                else:

                    ok, result = check_login(
                        username,
                        password
                    )

                    if ok:
                        
                        st.session_state.logged_in = True

                        st.session_state.office = result

                        st.session_state.update_locked = False

                        st.session_state.active_job_id = None

                        st.session_state.update_start_requested = False

                        st.session_state.job_preparing = False

                        st.session_state.pending_file_bytes = None

                        st.session_state.pending_filename = ""

                        st.session_state.final_file_bytes_cache = None

                        st.session_state.final_file_cache_id = None

                        st.rerun()

                    st.error(
                        result
                    )

        with tab2:

            st.markdown(
                """
<div style='font-size:22px;font-weight:800;color:#111827;margin:8px 0 3px;'>
    إنشاء حساب
</div>

<div style='color:#6b7280;font-size:13px;margin-bottom:18px;'>
    سجل مكتبك لبدء استخدام الخدمة
</div>
""",
                unsafe_allow_html=True
            )

            new_office = st.text_input(
                "اسم المكتب",
                key="reg_office",
                placeholder="اسم المكتب"
            )

            new_email = st.text_input(
                "الإيميل",
                key="reg_email",
                placeholder="example@email.com"
            )

            new_pass = st.text_input(
                "كلمة المرور",
                type="password",
                key="reg_pass",
                placeholder="كلمة المرور"
            )

            new_pass2 = st.text_input(
                "تأكيد كلمة المرور",
                type="password",
                key="reg_pass2",
                placeholder="أعد كتابة كلمة المرور"
            )

            if st.button(
                "إنشاء الحساب",
                key="reg_btn"
            ):

                if new_pass != new_pass2:

                    st.error(
                        "كلمة المرور مش متطابقة!"
                    )

                else:

                    ok, msg = register_office(
                        new_office,
                        new_email,
                        new_pass
                    )

                    (
                        st.success
                        if ok
                        else st.error
                    )(msg)

        st.markdown(
            '</div>',
            unsafe_allow_html=True
        )

    st.stop()


# =========================================================
# ADMIN
# =========================================================

if st.session_state.is_admin:

    st.markdown(
        """
<div class='hero'>
    <div class='hero-kicker'>الإدارة</div>
    <div class='hero-title'>لوحة الإدارة</div>
    <div class='hero-desc'>
        إدارة الحسابات الجديدة ومتابعة حالة المكاتب.
    </div>
</div>
""",
        unsafe_allow_html=True
    )

    pending = get_pending_accounts()

    st.markdown(
        f"""
<div class='section-title'>
    الحسابات المعلقة
    <span style='color:#2563eb'>
        ({len(pending)})
    </span>
</div>
""",
        unsafe_allow_html=True
    )

    for acc in get_pending_accounts():

        c1, c2, c3, c4 = st.columns(
            [3, 3, 1, 1]
        )

        c1.write(
            acc.get("name", "")
        )

        c2.write(
            acc.get("email", "")
        )

        if c3.button(
            "قبول",
            key=f"approve_{acc['id']}"
        ):

            set_office_status(
                acc["id"],
                "approved"
            )

            st.rerun()

        if c4.button(
            "رفض",
            key=f"reject_{acc['id']}"
        ):

            set_office_status(
                acc["id"],
                "rejected"
            )

            st.rerun()

    st.markdown(
        """
<div class='section-title'
     style='margin-top:28px;'>
    كل الحسابات
</div>
""",
        unsafe_allow_html=True
    )

    try:

        all_accounts = (
            db()
            .table("offices")
            .select(
                "name,email,status,created_at"
            )
            .order(
                "created_at",
                desc=True
            )
            .execute()
            .data
            or []
        )

        if all_accounts:

            st.dataframe(
                pd.DataFrame(
                    all_accounts
                ),
                use_container_width=True,
                hide_index=True
            )

    except Exception:
        pass

    if st.button(
        "تسجيل الخروج من الإدارة",
        key="admin_logout"
    ):

        reset_session_on_logout()

    st.stop()


# =========================================================
# OFFICE
# =========================================================

office = st.session_state.office

if not office:

    st.session_state.logged_in = False

    st.rerun()

office_id = office["id"]


# =========================================================
# TOP BAR
# =========================================================

st.markdown(
    """
<div class="topbar">
    <div class="brand">
        <div class="brand-icon">✨</div>
        <div>
            <div class="brand-title">Aivora</div>
            <div class="brand-sub">
                Your Smarter Support for Every Student's Application
            </div>
        </div>
    </div>

    <div style="font-size:13px;color:#6b7280;">
        نظام متابعة المكاتب
    </div>
</div>
""",
    unsafe_allow_html=True
)


# =========================================================
# HERO
# =========================================================

_hour = datetime.now().hour

_greeting = (
    "صباح الخير"
    if _hour < 12
    else "مساء الخير"
)

st.markdown(
    f"""
<div class="hero">
    <div class="hero-kicker">
        {_greeting} 👋
    </div>

    <div class="hero-title">
        أهلاً بيك،
        <strong>
            {office.get('name', '')}
        </strong>
    </div>

    <div class="hero-desc">
        تابع طلبات طلابك وحدّث الحالات من مكان واحد.
    </div>
</div>
""",
    unsafe_allow_html=True
)


# =========================================================
# SOURCE SELECTION
# =========================================================

st.markdown(
    '<div class="card">',
    unsafe_allow_html=True
)

st.markdown(
    """
<div class="section-title">
    مصدر بيانات الطلاب
</div>

<div class="section-sub">
    اختر الطريقة التي يحتوي بها ملف الطلاب.
</div>
""",
    unsafe_allow_html=True
)

source = st.radio(
    "",
    [
        "📂 رفع ملف Excel",
        "🔗 ربط Google Sheets"
    ],
    horizontal=True,
    label_visibility="collapsed",
    key="source_mode"
)

saved_link = get_saved_gsheet_link(
    office_id
)

file_bytes = None
filename = ""
source_url = None


# =========================================================
# EXCEL SOURCE
# =========================================================

if source == "📂 رفع ملف Excel":

    uploaded = st.file_uploader(
        "ارفع ملف Excel",
        type=[
            "xlsx",
            "xls"
        ],
        label_visibility="collapsed",
        key="excel_upload"
    )

    if uploaded:

        file_bytes = uploaded.getvalue()

        filename = uploaded.name

        st.session_state.pending_file_bytes = (
            file_bytes
        )

        st.session_state.pending_filename = (
            filename
        )

        st.success(
            f"تم اختيار الملف: {filename}"
        )

    elif st.session_state.pending_file_bytes:

        file_bytes = (
            st.session_state.pending_file_bytes
        )

        filename = (
            st.session_state.pending_filename
        )


# =========================================================
# GOOGLE SHEET SOURCE
# =========================================================

else:

    if saved_link:

        st.markdown(
            """
<div class="connected-box">
    ✓ Google Sheets متصل بالفعل لهذا المكتب
</div>
""",
            unsafe_allow_html=True
        )

        change = st.checkbox(
            "تغيير رابط الشيت",
            key="change_link"
        )

        if change:

            new_link = st.text_input(
                "رابط Google Sheets",
                key="new_link"
            )

            if st.button(
                "حفظ الرابط",
                key="save_link"
            ):

                ok, msg = save_gsheet_link(
                    office_id,
                    new_link
                )

                (
                    st.success
                    if ok
                    else st.error
                )(msg)

                if ok:
                    st.rerun()

        source_url = saved_link

    else:

        new_link = st.text_input(
            "رابط Google Sheets",
            key="first_link"
        )

        if st.button(
            "حفظ وربط الشيت",
            key="save_first_link"
        ):

            ok, msg = save_gsheet_link(
                office_id,
                new_link
            )

            (
                st.success
                if ok
                else st.error
            )(msg)

            if ok:
                st.rerun()

    if (
        source_url
        and st.button(
            "تحميل بيانات الشيت",
            key="load_sheet"
        )
    ):

        try:

            rows = read_gsheet_rows(
                source_url
            )

            wb = openpyxl.Workbook()

            ws = wb.active

            for row in rows:
                ws.append(row)

            out = io.BytesIO()

            wb.save(out)

            st.session_state.pending_file_bytes = (
                out.getvalue()
            )

            st.session_state.pending_filename = (
                "google_sheet"
            )

            st.success(
                "تم جلب البيانات. جاهزة للتحديث."
            )

        except Exception:

            st.error(
                "تعذر جلب بيانات الشيت حاليًا."
            )

    file_bytes = (
        st.session_state.pending_file_bytes
    )

    filename = (
        st.session_state.pending_filename
    )


st.markdown(
    '</div>',
    unsafe_allow_html=True
)


# =========================================================
# START UPDATE
# =========================================================

if (
    file_bytes
    and not st.session_state.update_locked
    and not st.session_state.active_job_id
):

    st.button(
        "▶ تحديث حالات الطلاب",
        key="start_update",
        on_click=start_update
    )


# =========================================================
# PREPARE NEW JOB
# =========================================================

if (
    st.session_state.update_start_requested
    and not st.session_state.active_job_id
    and not st.session_state.job_preparing
):
    running = (
    db()
    .table("jobs")
    .select("id,status,created_at,started_at")
    .eq(
        "office_id",
        office_id
    )
    .in_(
        "status",
        [
            "pending",
            "processing"
        ]
    )
    .order(
        "created_at",
        desc=True
    )
    .limit(1)
    .execute()
    .data
    or []
)


    if running:

        st.session_state.update_start_requested = False

        st.session_state.update_locked = True

        st.warning(
            "في تحديث شغال بالفعل لهذا المكتب. "
            "استني لحد ما يخلص."
        )

    else:

        st.session_state.job_preparing = True

        try:

            if source == "🔗 ربط Google Sheets":

                source_type = "google_sheet"

                source_name = "Google Sheet"

            else:

                source_type = "excel"

                source_name = (
                    filename
                    or "students.xlsx"
                )

            src, _ = import_students(
                office_id,
                source_type,
                source_name,
                file_bytes=file_bytes,
                source_url=source_url
            )

            job = create_job(
                office_id,
                src,
                source_name
            )

            st.session_state.active_job_id = (
                job["id"]
            )

            encryption_key = st.secrets.get(
                "STUDENT_PASSWORD_ENCRYPTION_KEY",
                os.getenv(
                    "STUDENT_PASSWORD_ENCRYPTION_KEY",
                    ""
                )
            )

            threading.Thread(
                target=_background_update_job,
                args=(
                    job["id"],
                    encryption_key
                ),
                daemon=True
            ).start()

        except Exception as exc:

            safe_log(
                f"job preparation failed: "
                f"{type(exc).__name__}: {exc}"
            )

            st.session_state.update_start_requested = False

            st.session_state.update_locked = False

            st.error(
                "تعذر تجهيز التحديث حاليًا. "
                "حاولي مرة تانية."
            )

        finally:

            st.session_state.job_preparing = False


# =========================================================
# PROCESSING UI
# =========================================================

@st.fragment(run_every=2)
def render_processing():

    job_id = st.session_state.get(
        "active_job_id"
    )

    if (
        not job_id
        and not st.session_state.get(
            "update_locked"
        )
    ):
        return

    st.markdown(
        '<div class="card">',
        unsafe_allow_html=True
    )

    st.markdown(
        """
<div class="section-title">
    تحديث حالات الطلاب
</div>
""",
        unsafe_allow_html=True
    )

    job = (
        get_job(job_id)
        if job_id
        else None
    )

    if not job:

        if st.session_state.get(
            "update_locked"
        ):

            st.markdown(
                """
<div class="lock-box">
    🔒 تم تشغيل تحديث بالفعل في هذه الجلسة.
    لو عايزة تبدئي تحديث جديد، سجّلي خروج وادخلي تاني.
</div>
""",
                unsafe_allow_html=True
            )

        st.markdown(
            '</div>',
            unsafe_allow_html=True
        )

        return

    status = str(
        job.get(
            "status"
        )
        or "pending"
    )

    if status == "pending":

        st.markdown(
            """
<div class="wait-box">
    ⏳ سيبدأ التحديث خلال ثواني…
</div>
""",
            unsafe_allow_html=True
        )

    elif status == "processing":

        st.markdown(
            """
<div class="wait-box">
    ▶️ بدأ التحديث.
    جاري فحص الطلاب وتحديث الحالات…
</div>
""",
            unsafe_allow_html=True
        )

    elif status == "cancelled":

        st.markdown(
            """
<div class="lock-box">
    🛑 تم إيقاف التحديث.
    الطلاب الذين تم فحصهم قبل الخروج تم حفظ نتائجهم،
    والطلاب الذين لم يبدأ فحصهم لم يتم تغييرهم.
</div>
""",
            unsafe_allow_html=True
        )

    students = get_students_for_job(
        job_id
    )

    progress_rows = get_job_progress_rows(
        job_id
    )

    progress_map = {
        str(
            r.get(
                "student_name"
            ) or ""
        ).strip().lower(): r
        for r in progress_rows
    }

    if students:

        table_rows = []

        checked = 0

        for index, student in enumerate(
            students,
            1
        ):

            name = str(
                student.get(
                    "student_name"
                )
                or student.get(
                    "login_identifier"
                )
                or "طالب"
            )

            row = progress_map.get(
                name.strip().lower()
            )

            current = (
                row.get("status")
                if row
                else str(
                    student.get(
                        "application_status"
                    )
                    or "لم يتم الفحص بعد"
                )
            )

            if row:
                checked += 1

            table_rows.append(
                (
                    index,
                    name,
                    current
                )
            )

        total = len(
            table_rows
        )

        st.progress(
            min(
                checked / max(
                    total,
                    1
                ),
                1.0
            )
        )

        st.caption(
            f"تم فحص {min(checked, total)} من {total} طالب"
        )

        if progress_rows:

            last = progress_rows[-1]

            st.info(
                f"🔄 آخر طالب تم فحصه: "
                f"**{last.get('student_name') or 'طالب'}** "
                f"— الحالة: "
                f"**{last.get('status') or ''}**"
            )

        html = [
            """
<table class="progress-table">
<thead>
<tr>
<th class="num">#</th>
<th class="name">اسم الطالب</th>
<th class="status">الحالة</th>
</tr>
</thead>
<tbody>
"""
        ]

        for (
            number,
            name,
            current
        ) in table_rows:

            html.append(
                f"""
<tr>
<td class="num">{number}</td>
<td class="name">{name}</td>
<td class="status">{current}</td>
</tr>
"""
            )

        html.append(
            """
</tbody>
</table>
"""
        )

        st.markdown(
            "".join(html),
            unsafe_allow_html=True
        )

    if status == "done":

        st.markdown(
            """
<div class="success-box"
     style="margin-top:14px">
    اكتمل التحديث 🎉
</div>
""",
            unsafe_allow_html=True
        )

        # =====================================================
        # LIVE EXCEL FILE
        #
        # There is no finalization file anymore.
        # The same source Drive file was updated live.
        # So download the same source file.
        # =====================================================

        if (
            str(
                job.get(
                    "source_type"
                ) or ""
            )
            == "excel"
        ):

            live_file_id = str(
                job.get(
                    "source_ref"
                ) or ""
            ).strip()

            if live_file_id:

                try:

                    if (
                        st.session_state.get(
                            "final_file_cache_id"
                        )
                        != live_file_id
                    ):

                        st.session_state.final_file_bytes_cache = (
                            download_drive_file_bytes(
                                live_file_id
                            )
                        )

                        st.session_state.final_file_cache_id = (
                            live_file_id
                        )

                    if st.session_state.get(
                        "final_file_bytes_cache"
                    ):

                        st.download_button(
                            "⬇️ تحميل ملف Excel المحدث",
                            data=(
                                st.session_state
                                .final_file_bytes_cache
                            ),
                            file_name=(
                                job.get(
                                    "file_name"
                                )
                                or "students_updated.xlsx"
                            ),
                            mime=(
                                "application/vnd.openxmlformats-"
                                "officedocument.spreadsheetml.sheet"
                            ),
                            key=(
                                f"download_final_{job_id}"
                            ),
                        )

                except Exception as exc:

                    safe_log(
                        f"live file download failed: "
                        f"{type(exc).__name__}: {exc}"
                    )

                    st.info(
                        "تعذر تجهيز الملف للتحميل حاليًا. "
                        "حاولي مرة تانية بعد قليل."
                    )

    elif status == "failed":

        # Technical details remain in Supabase logs/job.error;
        # office sees only this safe message.

        st.info(
            "تعذر إكمال التحديث حاليًا. "
            "حاولي مرة تانية بعد قليل."
        )

    elif not job_id and st.session_state.get(
        "update_locked"
    ):

        st.markdown(
            """
<div class="lock-box">
    🔒 تم تشغيل تحديث بالفعل في هذه الجلسة.
    لو عايزة تبدئي تحديث جديد، سجّلي خروج وادخلي تاني.
</div>
""",
            unsafe_allow_html=True
        )

    st.markdown(
        '</div>',
        unsafe_allow_html=True
    )


render_processing()


# =========================================================
# SEARCH
# =========================================================

st.markdown(
    '<div class="card">',
    unsafe_allow_html=True
)

st.markdown(
    """
<div class="section-title">
    البحث عن طالب
</div>

<div class="section-sub">
    اكتب اسم الطالب لمعرفة آخر حالة محفوظة.
</div>
""",
    unsafe_allow_html=True
)

search_query = st.text_input(
    "اسم الطالب",
    label_visibility="collapsed",
    key="student_search"
)

if search_query:

    found = get_students(
        office_id,
        search_query
    )

    if found:

        for student in found:

            status = (
                student.get(
                    "application_status"
                )
                or "لم يتم الفحص بعد"
            )

            st.markdown(
                f"""
<div style="
background:#f8fafc;
border:1px solid #e5e7eb;
border-radius:11px;
padding:12px;
margin:7px 0
">
    <b>
        👤 {student.get("student_name", "")}
    </b>
    <br>
    <span style="
        color:#1d4ed8;
        font-size:13px
    ">
        {status}
    </span>
</div>
""",
                unsafe_allow_html=True
            )

    else:

        st.info(
            "مفيش طالب بالاسم ده."
        )

st.markdown(
    '</div>',
    unsafe_allow_html=True
)


# =========================================================
# LOGOUT
# =========================================================

# Logout is intentionally at the bottom.
# Logging out cancels the active job,
# but the worker saves the CURRENT student
# before stopping.

if st.button(
    "تسجيل الخروج",
    key="logout_main"
):

    reset_session_on_logout()
import gspread
import plotly.express as px
import pandas as pd
from google.oauth2.service_account import Credentials
from datetime import datetime
import requests
import openpyxl
import json
import io
import time
import random

# ==================== إعدادات ====================
ADMIN_USERNAME = "admin"   # ← اسم المسؤول
ADMIN_PASSWORD = "admin123"  # ← باسورد المسؤول

SHEET_ID = "1BlFdtY-7ZIF1y2GwVosxlG9r7nK5xqYeW6yiIjPI_9U"
DRIVE_FOLDER_ID = "12L_qSHBnW4-tfQZRteynInWNBAML016f"

# ==================== Google Sheets client (كاش + retry لتفادي 429) ====================
# بنكاش الاتصال وشيت النظام الأساسي بدل ما نعمل authorize + open_by_key من جديد
# مع كل نداء - ده اللي كان بيستهلك quota بسرعة (خصوصًا في لوب متابعة تقدّم
# الـ Worker اللي بيسأل كل كذا ثانية).

@st.cache_resource(show_spinner=False)
def _get_gspread_client():
    creds_dict = st.secrets["gcp_service_account"]
    creds = Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/drive"]
    )
    return gspread.authorize(creds)


@st.cache_resource(show_spinner=False)
def _get_main_spreadsheet():
    return _get_gspread_client().open_by_key(SHEET_ID)


def with_retry(fn, *args, max_retries=6, **kwargs):
    """بينفذ أي نداء لـ gspread، ولو حصل 429 (quota exceeded) بيستنى وبيعيد المحاولة
    بدل ما يوقف الصفحة بخطأ فورًا."""
    delay = 5
    for attempt in range(1, max_retries + 1):
        try:
            return fn(*args, **kwargs)
        except gspread.exceptions.APIError as e:
            msg = str(e)
            if "429" in msg or "Quota exceeded" in msg or "RESOURCE_EXHAUSTED" in msg:
                if attempt == max_retries:
                    raise
                time.sleep(delay)
                delay = min(delay * 2, 60)
                continue
            raise


_worksheet_cache = {}


def _get_cached_worksheet(name, create_fn):
    if name in _worksheet_cache:
        return _worksheet_cache[name]
    spreadsheet = _get_main_spreadsheet()
    try:
        ws = with_retry(spreadsheet.worksheet, name)
    except Exception:
        ws = create_fn(spreadsheet)
    _worksheet_cache[name] = ws
    return ws


_office_ws_cache = {}


def _get_office_worksheet(link):
    """بنكاش تبويب شيت المكتب بعد أول فتح ليه في نفس الجلسة، بدل ما نفتحه من
    جديد كل مرة."""
    sheet_id = extract_sheet_id(link)
    if not sheet_id:
        return None
    if link in _office_ws_cache:
        return _office_ws_cache[link]
    spreadsheet = with_retry(_get_gspread_client().open_by_key, sheet_id)
    ws = get_target_worksheet(spreadsheet, link)
    _office_ws_cache[link] = ws
    return ws


# ==================== Google Sheets Logging ====================

def get_sheet():
    """اتصل بـ Google Sheets"""
    try:
        return _get_main_spreadsheet().sheet1
    except Exception as e:
        print(f"خطأ في Google Sheets: {e}")
        return None

def log_to_sheet(office, action, filename=""):
    """بيسجل كل عملية في Google Sheets"""
    try:
        sheet = get_sheet()
        if sheet:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # تأكد إن الهيدر موجود
        if sheet.row_count == 0 or with_retry(sheet.cell, 1, 1).value != "التاريخ":
            with_retry(sheet.insert_row, ["التاريخ", "اسم المكتب", "العملية", "اسم الملف"], 1)
        with_retry(sheet.append_row, [now, office, action, filename])
    except Exception as e:
        print(f"خطأ في التسجيل: {e}")

def get_accounts_sheet():
    """جيب شيت الحسابات"""
    try:
        def create(spreadsheet):
            sheet = spreadsheet.add_worksheet("accounts", 1000, 5)
            sheet.append_row(["اسم المكتب", "الإيميل", "الباسورد", "الحالة", "تاريخ التسجيل"])
            return sheet
        return _get_cached_worksheet("accounts", create)
    except Exception as e:
        return None

def register_office(office_name, email, password):
    """تسجيل مكتب جديد"""
    try:
        sheet = get_accounts_sheet()
        if not sheet:
            return False, "خطأ في الاتصال"
        
        # تأكد مش مسجل قبل كده
        records = sheet.get_all_records()
        for r in records:
            if r.get("اسم المكتب") == office_name or r.get("الإيميل") == email:
                return False, "اسم المكتب أو الإيميل مسجل قبل كده!"
        
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append_row([office_name, email, password, "pending", now])
        return True, "تم التسجيل! في انتظار موافقة الإدارة."
    except Exception as e:
        return False, str(e)

def check_login(office_name, password):
    """تحقق من اللوجين"""
    try:
        sheet = get_accounts_sheet()
        if not sheet:
            return False, "خطأ في الاتصال"
        
        records = sheet.get_all_records()
        for r in records:
            if r.get("اسم المكتب") == office_name and r.get("الباسورد") == password:
                if r.get("الحالة") == "approved":
                    return True, "ok"
                elif r.get("الحالة") == "pending":
                    return False, "حسابك في انتظار موافقة الإدارة!"
                else:
                    return False, "حسابك موقوف!"
        return False, "اسم المكتب أو الباسورد غلط!"
    except Exception as e:
        return False, str(e)

def get_pending_accounts():
    """جيب الحسابات في انتظار الموافقة"""
    try:
        sheet = get_accounts_sheet()
        if not sheet:
            return []
        records = sheet.get_all_records()
        return [r for r in records if r.get("الحالة") == "pending"]
    except:
        return []

def approve_account(office_name):
    """وافقي على حساب"""
    try:
        sheet = get_accounts_sheet()
        records = sheet.get_all_records()
        for i, r in enumerate(records, start=2):
            if r.get("اسم المكتب") == office_name:
                sheet.update_cell(i, 4, "approved")
                return True
        return False
    except:
        return False

def reject_account(office_name):
    """ارفضي حساب"""
    try:
        sheet = get_accounts_sheet()
        records = sheet.get_all_records()
        for i, r in enumerate(records, start=2):
            if r.get("اسم المكتب") == office_name:
                sheet.update_cell(i, 4, "rejected")
                return True
        return False
    except:
        return False

def get_latest_file_from_drive(office):
    """بيجيب آخر إكسيل محدث للمكتب من Drive"""
    try:
        from googleapiclient.discovery import build
        creds_dict = st.secrets["gcp_service_account"]
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        service = build("drive", "v3", credentials=creds)

        # اسم المكتب ممكن يكون فيه علامات تكسر الكويري (زي ' أو \) — نعمل escape ليها
        safe_office = str(office).replace("\\", "\\\\").replace("'", "\\'")

        # ابحث عن آخر ملف للمكتب ده في الفولدر (مع دعم Shared Drives لو الفولدر جواها)
        results = service.files().list(
            q=f"'{DRIVE_FOLDER_ID}' in parents and name contains '{safe_office}' and trashed=false",
            orderBy="createdTime desc",
            pageSize=1,
            fields="files(id, name, createdTime)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()

        files = results.get("files", [])
        if not files:
            # نتأكد الفولدر أصلاً متاح ومقدر نشوف اللي فيه، عشان نميز
            # بين "مفيش ملفات بالاسم ده" و"مش قادرين نوصل للفولدر خالص"
            try:
                probe = service.files().list(
                    q=f"'{DRIVE_FOLDER_ID}' in parents and trashed=false",
                    pageSize=5,
                    fields="files(id, name)",
                    supportsAllDrives=True,
                    includeItemsFromAllDrives=True,
                ).execute()
                existing = [f["name"] for f in probe.get("files", [])]
                if not existing:
                    return None, "الفولدر على Drive فاضي أو الـ service account مالوش صلاحية يشوف محتوياته — تأكدي إن الفولدر متشيير مع إيميل الـ service account."
                return None, f"مفيش ملفات باسم فيه '{office}'. أسماء ملفات موجودة فعلاً في الفولدر: {existing}"
            except Exception as probe_err:
                return None, f"مفيش ملفات محفوظة للمكتب ده، وكمان فشل فحص الفولدر: {probe_err}"

        # حمل الملف
        from googleapiclient.http import MediaIoBaseDownload
        import io as _io

        file_id = files[0]["id"]
        request = service.files().get_media(fileId=file_id)
        buf = _io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)
        return buf, None

    except Exception as e:
        return None, str(e)


def search_in_excel(file_bytes, name_query):
    """بيبحث في الإكسيل عن طالب بالاسم ويرجع آخر حالة"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        # لاقي الهيدر
        cols, header_row_num = find_excel_columns(ws)

        results = []
        for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
            name   = str(row[cols["name"]]   if cols["name"]   is not None else "").strip()
            status = str(row[cols["status"]] if cols["status"] is not None else "").strip()

            if name_query.strip() in name:
                results.append({"name": name, "status": status})

        return results
    except Exception as e:
        return []


def save_gsheet_link(office, link):
    """بيحفظ لينك Google Sheets بتاع المكتب — بترجع (ok, message) عشان تبان المشكلة الحقيقية"""
    try:
        sheet = get_accounts_sheet()
        if not sheet:
            return False, "مش قادر أوصل لشيت الحسابات (accounts sheet) — تأكدي من صلاحيات الـ service account."

        headers = with_retry(sheet.row_values, 1)
        target = str(office).strip()

        # بندور صف صف على عمود "اسم المكتب" (العمود الأول) بدل get_all_records
        # عشان نتجنب مشاكل الهيدرز الفاضية/المكررة
        office_col_values = with_retry(sheet.col_values, 1)
        row_num = None
        for i, val in enumerate(office_col_values[1:], start=2):  # صف 1 = هيدر
            if str(val).strip() == target:
                row_num = i
                break

        if row_num is None:
            all_names = [str(v).strip() for v in office_col_values[1:]]
            return False, f"مش لاقي مكتب اسمه '{target}' في شيت الحسابات. الأسماء الموجودة فعلاً: {all_names}"

        # لو مفيش عمود للينك، ضيفيه (مع توسيع الشيت لو لازم)
        if "لينك الشيت" not in headers:
            col_num = len(headers) + 1
            if col_num > sheet.col_count:
                sheet.add_cols(col_num - sheet.col_count)
            with_retry(sheet.update_cell, 1, col_num, "لينك الشيت")
        else:
            col_num = headers.index("لينك الشيت") + 1

        with_retry(sheet.update_cell, row_num, col_num, link)
        return True, "تم الحفظ بنجاح"

    except Exception as e:
        print(f"خطأ في حفظ اللينك: {e}")
        return False, f"خطأ: {e}"


def get_gsheet_link(office):
    """جيب لينك Google Sheets بتاع المكتب"""
    try:
        sheet = get_accounts_sheet()
        if not sheet:
            return None
        headers = with_retry(sheet.row_values, 1)
        if "لينك الشيت" not in headers:
            return None
        link_col = headers.index("لينك الشيت") + 1

        # بندور صف صف على عمود "اسم المكتب" (العمود الأول) بدل get_all_records
        # عشان نتجنب مشاكل الهيدرز الفاضية/المكررة اللي ممكن تطلع استثناء صامت
        office_col_values = with_retry(sheet.col_values, 1)
        target = str(office).strip()
        for i, val in enumerate(office_col_values[1:], start=2):  # صف 1 = هيدر
            if str(val).strip() == target:
                link = with_retry(sheet.cell, i, link_col).value
                return link if link else None
        return None
    except Exception as e:
        print(f"خطأ في get_gsheet_link: {e}")
        return None


def _find_office_row(sheet, office):
    """بيرجع رقم صف المكتب في شيت الحسابات (أو None لو مش موجود)"""
    office_col_values = with_retry(sheet.col_values, 1)
    target = str(office).strip()
    for i, val in enumerate(office_col_values[1:], start=2):  # صف 1 = هيدر
        if str(val).strip() == target:
            return i
    return None


def extract_sheet_id(link):
    """بيستخرج الـ Sheet ID من اللينك"""
    import re
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', link)
    if match:
        return match.group(1)
    return None


def extract_gid(link):
    """بيستخرج رقم الـ gid (بيحدد التبويب/الشيت بالظبط جوه الملف) من اللينك لو موجود"""
    import re
    match = re.search(r'[?#&]gid=(\d+)', link)
    if match:
        return int(match.group(1))
    return None


def get_target_worksheet(spreadsheet, link):
    """بيرجع التبويب المطابق بالظبط للينك المحفوظ (حسب gid)، أو أول تبويب لو مفيش gid في اللينك"""
    gid = extract_gid(link)
    if gid is not None:
        try:
            for ws in spreadsheet.worksheets():
                if ws.id == gid:
                    return ws
        except Exception:
            pass
    return spreadsheet.sheet1


def read_gsheet_as_excel(link):
    """بيقرا التبويب المحدد في اللينك (حسب gid) من Google Sheet ويرجعه كـ BytesIO زي إكسيل"""
    try:
        ws = _get_office_worksheet(link)
        if ws is None:
            return None, "الرابط غير صحيح!"
        data = with_retry(ws.get_all_values)

        if not data:
            return None, "الشيت فاضي!"

        wb = openpyxl.Workbook()
        wsheet = wb.active
        for row in data:
            wsheet.append(row)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out, None
    except Exception as e:
        return None, str(e)


def _get_gsheet_worksheet(link):
    """بترجع (worksheet, error) للتبويب المحدد في اللينك (بنكاش الاتصال بدل ما
    نفتحه من جديد مع كل نداء - بيتنادى مرة لكل طالب)."""
    try:
        ws = _get_office_worksheet(link)
        if ws is None:
            return None, "الرابط غير صحيح"
        return ws, None
    except Exception as e:
        return None, str(e)


def ensure_status_header_in_gsheet(link, header_row_num, status_col_index, header_name="حالة الطلب الجديدة"):
    """بتتأكد إن هيدر عمود 'حالة الطلب الجديدة' مكتوب في شيت المكتب — وده العمود
    الوحيد اللي الأجنت بيضيفه/يكتب فيه. لو موجود بالفعل، مبيعملش أي حاجة."""
    try:
        ws, err = _get_gsheet_worksheet(link)
        if err or not ws:
            return False
        col_num = status_col_index + 1
        current = with_retry(ws.cell, header_row_num, col_num).value
        if str(current or "").strip() != header_name:
            with_retry(ws.update_cell, header_row_num, col_num, header_name)
        return True
    except Exception:
        return False


def update_status_cell_in_gsheet(link, row_num, status_col_index, value):
    """بتكتب قيمة الحالة في خلية واحدة بس (عمود 'حالة الطلب الجديدة') لصف طالب
    معيّن في شيت المكتب — من غير ما تلمس أي عمود أو صف تاني، ومن غير أي مسح
    أو إعادة كتابة لباقي الشيت."""
    try:
        ws, err = _get_gsheet_worksheet(link)
        if err or not ws:
            return False, err or "تعذر الوصول للشيت"
        with_retry(ws.update_cell, row_num, status_col_index + 1, value)
        return True, None
    except Exception as e:
        return False, str(e)


def get_results_sheet():
    """جيب أو أنشئ ورقة النتائج الموحدة (بديل رفع الملفات على Drive اللي مش شغال مع service account)"""
    try:
        def create(spreadsheet):
            sheet = spreadsheet.add_worksheet("results", 2000, 4)
            sheet.append_row(["اسم المكتب", "اسم الطالب", "الحالة", "تاريخ التحديث"])
            return sheet
        return _get_cached_worksheet("results", create)
    except Exception as e:
        print(f"خطأ في ورقة النتائج: {e}")
        return None


def save_results_to_sheet(office, results):
    """بيحفظ آخر نتائج المكتب في ورقة 'results' — بيمسح القديم بتاع المكتب ده ويكتب الجديد.
    بتتنادى بعد كل طالب أثناء التحديث (مش بس في الآخر) عشان لو الصفحة اتقفلت أو
    التحديث اتوقف فجأة، النتائج اللي خلصت فعلاً تكون محفوظة أولاً بأول."""
    try:
        sheet = get_results_sheet()
        if not sheet:
            return False, "مش قادر أوصل لورقة النتائج"

        all_values = with_retry(sheet.get_all_values)
        headers = all_values[0] if all_values else ["اسم المكتب", "اسم الطالب", "الحالة", "تاريخ التحديث"]

        target = str(office).strip()
        # سيبي باقي المكاتب زي ما هي، وامسحي بس صفوف المكتب ده القديمة
        kept_rows = [row for row in all_values[1:] if row and str(row[0]).strip() != target]

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_rows = [[target, r.get("name", ""), r.get("status", ""), now] for r in results]

        final_data = [headers] + kept_rows + new_rows
        with_retry(sheet.clear)
        with_retry(sheet.update, final_data)
        return True, "تم حفظ النتائج"
    except Exception as e:
        print(f"خطأ في حفظ النتائج: {e}")
        return False, f"خطأ: {e}"


def search_results_in_sheet(office, name_query):
    """بيبحث عن طالب باسمه جوه آخر نتائج محفوظة للمكتب ده"""
    try:
        sheet = get_results_sheet()
        if not sheet:
            return [], "مش قادر أوصل لورقة النتائج"

        all_values = with_retry(sheet.get_all_values)
        if len(all_values) <= 1:
            return [], "مفيش نتائج محفوظة لأي مكتب لسه — لازم تعملي '▶ ابدأ' مرة واحدة الأول."

        target_office = str(office).strip()
        query = str(name_query).strip()
        found = []
        office_has_any = False
        for row in all_values[1:]:
            if len(row) < 3:
                continue
            row_office, row_name, row_status = row[0], row[1], row[2]
            if str(row_office).strip() != target_office:
                continue
            office_has_any = True
            if query in str(row_name):
                found.append({"name": row_name, "status": row_status})

        if not office_has_any:
            return [], "مفيش نتائج محفوظة للمكتب ده لسه — لازم تعملي '▶ ابدأ' مرة واحدة الأول."
        return found, None
    except Exception as e:
        return [], f"خطأ: {e}"


def upload_to_drive(file_bytes, filename, office):
    """بيرفع الإكسيل على Google Drive — بترجع (ok, message) عشان تبان أي مشكلة فعلية"""
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
        import io as _io

        creds_dict = st.secrets["gcp_service_account"]
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://www.googleapis.com/auth/drive"]
        )

        service = build("drive", "v3", credentials=creds)

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        drive_filename = f"{office}_{now}_{filename}"

        file_metadata = {
            "name": drive_filename,
            "parents": [DRIVE_FOLDER_ID]
        }

        media = MediaIoBaseUpload(
            _io.BytesIO(file_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id",
            supportsAllDrives=True,
        ).execute()

        return True, "تم الرفع بنجاح"
    except Exception as e:
        print(f"خطأ في رفع الملف: {e}")
        return False, f"فشل الرفع على Drive: {e}"


# ==================== Worker (تحديث من جهاز محلي بدل Cloud) ====================
# الفكرة: لو فيه Worker شغال على جهاز محلي (بيبعت "نبضة" كل شوية ثواني)، بنحوّل
# شغل تسجيل الدخول للموقع الحكومي ليه (IP منزلي أقل عرضة للحظر من IP بتاع الـ Cloud)،
# بدل ما الـ Streamlit App اللي شغال على Streamlit Cloud يعمل اللوجينز بنفسه.
# التواصل بين الاتنين بيتم بالكامل عن طريق Google Sheets (tabs: worker_heartbeat, jobs, job_progress).

# مفيش heartbeat خالص. بنعرف إن فيه Worker شغال أو لأ بطريقة واحدة بس: بننشئ
# مهمة (job) ونستنى لحد 45 ثانية نشوف هل حالتها اتحولت لـ "processing" (يعني
# حد شغال قاعد ياخدها) - لو حصل، نكمل معاه؛ لو لأ، نلغي المهمة ونشتغل Cloud.
JOB_PICKUP_WAIT_SECONDS = 45
JOB_PICKUP_POLL_INTERVAL = 3

# الحالات اللي لو الطالب وصلها، ما بنعيدش فحصه في التحديثات الجاية
FINAL_STATUSES = {
    "مقبول نهائي",
    "قبول نهائي",
    "تم الرفض",
    "مرفوض نهائيًا",
    "مرفوض نهائيا",
    "مرفوض",
    "خالص",
}

# علامات على إن فشل تسجيل الدخول كان بسبب مشكلة اتصال مؤقتة (نت/SSL/تايم آوت)
# مش بسبب إيميل أو باسورد غلط - الحالات دي بس بتتعاد محاولتها تلقائيًا في آخر
# التحديث بدل ما تفضل مسجلة "فشل" نهائي من أول محاولة واحدة اتعطلت.
TRANSIENT_ERROR_MARKERS = [
    "SSLError", "SSLEOFError", "SSLZeroReturnError",
    "ConnectionError", "ConnectionResetError", "RemoteDisconnected",
    "Max retries exceeded", "Timeout", "ReadTimeout", "ConnectTimeout",
    "EOF occurred",
]


def _is_transient_login_error(status_text):
    text = str(status_text or "")
    return text.startswith("فشل تسجيل الدخول") and any(marker in text for marker in TRANSIENT_ERROR_MARKERS)


def _get_gs_client():
    return _get_gspread_client()


def get_jobs_sheet():
    try:
        def create(spreadsheet):
            sheet = spreadsheet.add_worksheet("jobs", 500, 9)
            sheet.append_row(["job_id", "اسم المكتب", "نوع المصدر", "مرجع المصدر",
                               "اسم الملف", "الحالة", "تاريخ الإنشاء", "final_drive_file_id", "خطأ"])
            return sheet
        return _get_cached_worksheet("jobs", create)
    except Exception as e:
        print(f"خطأ في get_jobs_sheet: {e}")
        return None


def get_progress_sheet():
    try:
        def create(spreadsheet):
            sheet = spreadsheet.add_worksheet("job_progress", 5000, 5)
            sheet.append_row(["job_id", "index", "total", "اسم الطالب", "الحالة"])
            return sheet
        return _get_cached_worksheet("job_progress", create)
    except Exception as e:
        print(f"خطأ في get_progress_sheet: {e}")
        return None


def upload_pending_file(file_bytes, filename, office):
    """بيرفع ملف الطلاب على Drive مؤقتًا عشان الـ Worker يقدر يحمّله ويشتغل عليه.
    بترجع (file_id, error)."""
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
        import io as _io

        creds_dict = st.secrets["gcp_service_account"]
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        service = build("drive", "v3", credentials=creds)

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        drive_filename = f"PENDING_{office}_{now}_{filename}"
        file_metadata = {"name": drive_filename, "parents": [DRIVE_FOLDER_ID]}
        media = MediaIoBaseUpload(
            _io.BytesIO(file_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        created = service.files().create(
            body=file_metadata, media_body=media, fields="id", supportsAllDrives=True
        ).execute()
        return created.get("id"), None
    except Exception as e:
        return None, f"فشل رفع الملف المؤقت على Drive: {e}"


def download_drive_file_bytes(file_id):
    """بيحمل ملف من Drive بالـ ID ويرجعه كـ bytes"""
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseDownload
        import io as _io

        creds_dict = st.secrets["gcp_service_account"]
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        service = build("drive", "v3", credentials=creds)
        request = service.files().get_media(fileId=file_id)
        buf = _io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)
        return buf.read(), None
    except Exception as e:
        return None, str(e)


def create_job(office, source_type, source_ref, filename):
    """بينشئ مهمة جديدة في شيت jobs عشان الـ Worker يلتقطها. بيرجع job_id."""
    sheet = get_jobs_sheet()
    if not sheet:
        return None
    job_id = f"{office}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{random.randint(1000,9999)}"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append_row([job_id, office, source_type, source_ref, filename, "pending", now, "", ""])
    return job_id


def get_job(job_id):
    """بيرجع dict فيه بيانات المهمة الحالية (الحالة، الأخطاء، إلخ)"""
    sheet = get_jobs_sheet()
    if not sheet:
        return None
    records = sheet.get_all_records()
    for r in records:
        if str(r.get("job_id", "")) == job_id:
            return r
    return None


def cancel_job(job_id):
    """بيلغي مهمة لسه pending (محدش شافها) - بنستخدمها لما محدش من الـ Worker
    ياخدها خلال مهلة الانتظار، عشان نتأكد إنه لو الـ Worker اشتغل متأخر مايبقاش
    فيه تعارض ويشتغل عليها من غير داعي."""
    try:
        sheet = get_jobs_sheet()
        if not sheet:
            return False
        cell = sheet.find(job_id)
        if not cell:
            return False
        row_num = cell.row
        headers = sheet.row_values(1)
        status_col = headers.index("الحالة") + 1
        current_status = sheet.cell(row_num, status_col).value
        if current_status == "pending":
            sheet.update_cell(row_num, status_col, "cancelled")
        return True
    except Exception as e:
        print(f"خطأ في cancel_job: {e}")
        return False


def wait_for_job_pickup(job_id):
    """بيستنى لحد JOB_PICKUP_WAIT_SECONDS نشوف هل فيه Worker هياخد المهمة دي.
    بيرجع (picked_up, last_job) - picked_up = True لو الحالة اتحولت processing
    (أو خلصت أصلاً) خلال المهلة."""
    waited = 0
    job = None
    while waited < JOB_PICKUP_WAIT_SECONDS:
        job = get_job(job_id)
        status = (job or {}).get("الحالة", "")
        if status in ("processing", "done", "failed"):
            return True, job
        time.sleep(JOB_PICKUP_POLL_INTERVAL)
        waited += JOB_PICKUP_POLL_INTERVAL
    return False, job


def get_job_progress_rows(job_id):
    """بيرجع كل صفوف التقدّم الخاصة بمهمة معينة، مرتبة حسب index"""
    sheet = get_progress_sheet()
    if not sheet:
        return []
    records = sheet.get_all_records()
    rows = [r for r in records if str(r.get("job_id", "")) == job_id]
    try:
        rows.sort(key=lambda r: int(r.get("index", 0)))
    except Exception:
        pass
    return rows


# ==================== API ====================
BASE_URL = "https://apiadm.study-in-egypt.gov.eg/api"
SITE_URL = "https://admission.study-in-egypt.gov.eg"

HEADERS_BASE = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "ar",
    "device": "CITIZEN",
    "origin": SITE_URL,
    "referer": SITE_URL + "/",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "content-type": "application/json",
}


def human_delay(min_sec=2, max_sec=5):
    """استنى وقت عشوائي زي إنسان"""
    time.sleep(random.uniform(min_sec, max_sec))


def api_login(email, password, max_retries=3):
    """بيعمل لوجين ويرجع session جديدة فيها الكوكي.
    لو حصل خطأ اتصال/SSL مؤقت (زي الموقع بيقفل الكونكشن فجأة)، بيعيد المحاولة
    كذا مرة بفاصل زمني متزايد قبل ما يعتبرها فشل نهائي."""
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            session = requests.Session()
            session.headers.update(HEADERS_BASE)

            # استنى شوية قبل اللوجين
            human_delay(3, 6)

            res = session.post(
                f"{BASE_URL}/student/login",
                json={"email": email, "password": password},
                timeout=30
            )

            if res.status_code not in [200, 201]:
                last_err = f"فشل اللوجين - كود: {res.status_code}"
                # لو كود 429 (طلبات كتير) أو 5xx (مشكلة سيرفر)، يستاهل إعادة محاولة
                if res.status_code == 429 or res.status_code >= 500:
                    human_delay(6 * attempt, 10 * attempt)
                    continue
                return None, None, last_err

            # استنى شوية بعد اللوجين
            human_delay(2, 4)

            csrf_token = res.json().get("token", "") or res.headers.get("x-csrf-token", "")

            return session, csrf_token, None

        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
            # أخطاء اتصال/SSL مؤقتة (الموقع بيقفل الكونكشن فجأة) - نعيد المحاولة
            last_err = str(e)
            if attempt < max_retries:
                human_delay(6 * attempt, 10 * attempt)
                continue
        except Exception as e:
            return None, None, str(e)

    # خلصت كل المحاولات ولسه فاشل
    return None, None, last_err


def api_logout(session):
    """تسجيل خروج"""
    try:
        session.post(
            f"{BASE_URL}/student/logout",
            json={"redirectUrl": SITE_URL},
            timeout=15
        )
        human_delay(1, 2)
    except:
        pass


def get_status(session, csrf_token):
    """بيجيب حالة أحدث طلب"""
    try:
        filter_obj = {
            "where": {},
            "limit": 10,
            "offset": 0,
            "order": "statusUpdatedAt DESC",
            "fields": ["serviceSlug", "ID", "createdAt", "statusUpdatedAt", "activityId", "activityName"]
        }

        headers = {}
        if csrf_token:
            headers["x-csrf-token"] = csrf_token

        # استنى شوية قبل الـ request
        human_delay(1, 3)

        res = session.get(
            f"{BASE_URL}/dynamic_services/inbox",
            params={"filter": json.dumps(filter_obj)},
            headers=headers,
            timeout=30
        )

        if res.status_code not in [200, 304]:
            return "", f"خطأ ({res.status_code})"

        data = res.json()
        results = data.get("result", [])

        if not results:
            return "", "مفيش طلبات"

        # أحدث طلب (مرتبين من الأحدث للأقدم)
        latest = results[0]

        # جدول ترجمة activityName للعربي الصح
        translations = {
            "قبول الفحص الفنى": "القبول المبدئي",
            "قبول الفحص الفني": "القبول المبدئي",
            "kb8ijfo8": "تم السداد",
            "تم السداد": "تم السداد",
            "تأكيد استلام الملف وصحة و اكتمال المستندات": "تأكيد استلام الملف وصحة واكتمال المستندات",
            "الانتظار مراجعة الطلب": "بانتظار مراجعة الطلب",
            "قبول من رئيس الادارة المركزية": "قبول من رئيس الإدارة المركزية",
        }
        raw_status = latest.get("activityName", "غير محدد")
        status = translations.get(raw_status, raw_status)
        app_id = str(latest.get("ID", ""))
        return app_id, status

    except Exception as e:
        return "", f"خطأ: {e}"


def find_excel_columns(ws):
    cols = {"name": None, "email": None, "password": None, "status": None}
    header_row_num = None
    header_len = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        row_values = [str(c).strip() if c else "" for c in row]
        if any("يميل" in v or "mail" in v.lower() for v in row_values):
            header_row_num = row_idx
            header_len = len(row_values)

            # أولوية: عمود اسمه "حالة الطلب الجديدة" تحديدًا - ده العمود اللي هنكتب فيه التحديث،
            # عشان ميتلخبطش مع أي عمود "حالة الطلب" أصلي موجود أصلاً في ملف المكتب
            for i, cell in enumerate(row_values):
                if "حالة" in cell and "الجديدة" in cell:
                    cols["status"] = i
                    break

            for i, cell in enumerate(row_values):
                cell_lower = cell.lower()
                if any(k in cell for k in ["اسم", "الإسم", "الاسم"]) or "name" in cell_lower:
                    cols["name"] = i
                elif any(k in cell for k in ["يميل", "بريد"]) or "mail" in cell_lower:
                    cols["email"] = i
                elif any(k in cell for k in ["باسورد", "كلمة المرور", "password", "pass"]) or "pass" in cell_lower:
                    cols["password"] = i
            break
    if header_row_num is None:
        raise Exception("مش لاقي هيدر الإكسيل!")
    if cols["email"] is None:
        raise Exception("مش لاقي عمود الإيميل!")
    if cols["password"] is None:
        raise Exception("مش لاقي عمود الباسورد!")

    # لو عمود "حالة الطلب الجديدة" مش موجود أصلاً، نضيفه تلقائيًا في آخر الشيت
    if cols["status"] is None:
        new_col_index = header_len
        ws.cell(row=header_row_num, column=new_col_index + 1, value="حالة الطلب الجديدة")
        cols["status"] = new_col_index

    return cols, header_row_num


# ==================== الواجهة ====================
st.set_page_config(
    menu_items={"Get help": None, "Report a bug": None, "About": None},
    page_title="Aivora - Agent",
    page_icon="✨",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ===== منع ترجمة المتصفح التلقائية (كانت بتلخبط عناصر الصفحة زي حقل كلمة المرور وزرار الرفع) =====
# st.markdown بيحقن الكود جوه الـ body، لكن Google Translate بيحترم meta tag بس لو كانت
# فعليًا جوه الـ <head>. فبنستخدم JS عشان نحقنها في مستند الصفحة الأصلي (مش iframe المكوّن).
import streamlit.components.v1 as components

components.html("""
<script>
(function() {
    try {
        var doc = window.parent.document;
        if (!doc.querySelector('meta[name="google"][content="notranslate"]')) {
            var meta = doc.createElement('meta');
            meta.name = 'google';
            meta.content = 'notranslate';
            doc.head.appendChild(meta);
        }
        doc.documentElement.classList.add('notranslate');
        doc.documentElement.setAttribute('translate', 'no');
    } catch (e) {}
})();
</script>
""", height=0, width=0)

st.markdown("""
<style>
/* ===== Global ===== */
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"], .stApp {
    font-family: 'Cairo', sans-serif !important;
    direction: rtl;
    translate: no; /* منع ترجمة المتصفح التلقائية اللي بتلخبط الـ DOM */
}
.stApp { background: #f5f7fb; color: #111827; }
[data-testid="stAppViewContainer"] > .main { background: #f5f7fb; }
.block-container { max-width: 1180px; padding-top: 1.4rem; padding-bottom: 3rem; }
#MainMenu, footer, header, [data-testid="stToolbar"], [data-testid="stDecoration"],
[data-testid="stStatusWidget"], [data-testid="stSidebarNav"] { display:none !important; }

/* ===== Typography ===== */
h1, h2, h3, h4, p, label, span, div { font-family: 'Cairo', sans-serif !important; }
h1 { color:#111827 !important; font-size:32px !important; font-weight:800 !important; }
h2 { color:#111827 !important; font-size:23px !important; font-weight:800 !important; }
h3 { color:#111827 !important; font-size:19px !important; font-weight:700 !important; }
.stCaption, [data-testid="stCaptionContainer"] p { color:#6b7280 !important; }

/* ===== Top bar ===== */
.topbar {
    background:#ffffff;
    border:1px solid #e5e7eb;
    border-radius:16px;
    padding:13px 18px;
    display:flex;
    align-items:center;
    justify-content:space-between;
    box-shadow:0 2px 10px rgba(17,24,39,.04);
    margin-bottom:22px;
}
.brand { display:flex; align-items:center; gap:11px; }
.brand-icon {
    width:42px; height:42px; border-radius:12px; background:#eff6ff;
    display:flex; align-items:center; justify-content:center; font-size:22px;
}
.brand-title { font-size:18px; font-weight:800; color:#111827; }
.brand-sub { font-size:12px; color:#6b7280; margin-top:-2px; }

/* ===== Cards ===== */
.card {
    background:#fff; border:1px solid #e5e7eb; border-radius:18px;
    padding:22px; box-shadow:0 3px 14px rgba(17,24,39,.045);
    margin-bottom:18px;
}
.hero {
    background:linear-gradient(135deg,#ffffff 0%,#f8fbff 100%);
    border:1px solid #dbeafe; border-radius:20px; padding:25px 28px;
    box-shadow:0 4px 18px rgba(37,99,235,.06); margin-bottom:20px;
}
.hero-kicker { color:#6b7280; font-size:14px; font-weight:600; }
.hero-title { color:#111827; font-size:28px; font-weight:800; margin-top:2px; }
.hero-title strong { color:#2563eb; }
.hero-desc { color:#6b7280; font-size:14px; margin-top:3px; }

.stat-card { background:#fff; border:1px solid #e5e7eb; border-radius:16px; padding:17px 18px; }
.stat-label { color:#6b7280; font-size:12px; font-weight:600; }
.stat-value { color:#111827; font-size:21px; font-weight:800; margin-top:2px; }

.section-title { font-size:18px; font-weight:800; color:#111827; margin:5px 0 13px; }
.section-sub { color:#6b7280; font-size:13px; margin-top:-8px; margin-bottom:14px; }

/* ===== Inputs ===== */
.stTextInput label, .stFileUploader label, .stRadio > label, .stCheckbox label {
    color:#374151 !important; font-size:14px !important; font-weight:700 !important;
}
.stTextInput input {
    background:#fff !important; color:#111827 !important; border:1px solid #d1d5db !important;
    border-radius:10px !important; font-size:14px !important; min-height:44px;
}
.stTextInput input:focus { border-color:#2563eb !important; box-shadow:0 0 0 3px rgba(37,99,235,.10) !important; }
.stTextInput input::placeholder { color:#9ca3af !important; }

/* ===== Buttons ===== */
.stButton > button, .stDownloadButton > button {
    width:100%; min-height:44px; border-radius:10px !important; border:1px solid #2563eb !important;
    background:#2563eb !important; color:#fff !important; font-weight:700 !important;
    font-size:14px !important; box-shadow:0 3px 8px rgba(37,99,235,.16) !important;
    transition:.15s ease;
}
.stButton > button:hover, .stDownloadButton > button:hover {
    background:#1d4ed8 !important; border-color:#1d4ed8 !important; transform:translateY(-1px);
}

/* Secondary buttons */
button[kind="secondary"] { background:#fff !important; color:#2563eb !important; }

/* ===== Tabs ===== */
button[data-baseweb="tab"] {
    color:#6b7280 !important; font-weight:700 !important; font-size:14px !important;
}
button[data-baseweb="tab"][aria-selected="true"] { color:#2563eb !important; }
[data-baseweb="tab-highlight"] { background:#2563eb !important; height:2px !important; }

/* ===== Radio / uploader ===== */
.stRadio div[role="radiogroup"] { gap:10px; }
.stRadio div[role="radiogroup"] label {
    background:#fff; border:1px solid #e5e7eb; border-radius:12px; padding:10px 14px;
}
[data-testid="stFileUploaderDropzone"] {
    background:#f8fafc !important; border:1.5px dashed #cbd5e1 !important; border-radius:14px !important;
}
[data-testid="stFileUploaderDropzone"] button {
    background:#fff !important; border:1px solid #bfdbfe !important;
    border-radius:8px !important;
    position:relative !important;
    /* بنخفي أي نص أصلي (حتى لو بيتكرر بصريًا لأي سبب) ونعرض تسمية واحدة ثابتة بدلاً منه */
    font-size:0 !important;
    color:transparent !important;
    text-shadow:none !important;
    min-width:110px;
}
[data-testid="stFileUploaderDropzone"] button::after {
    content: "اختيار ملف";
    position:absolute;
    inset:0;
    display:flex;
    align-items:center;
    justify-content:center;
    font-family:'Cairo', sans-serif !important;
    font-size:14px !important;
    font-weight:700 !important;
    color:#2563eb !important;
}
[data-testid="stFileUploaderDropzone"] button * { display:none !important; }

/* ===== Alerts ===== */
div[data-testid="stAlert"] { border-radius:12px !important; border:1px solid #e5e7eb !important; }
div[data-testid="stAlert"] p { font-size:13px !important; font-weight:600 !important; }

/* ===== Search results ===== */
.result-card { background:#fff; border:1px solid #e5e7eb; border-radius:13px; padding:14px 16px; margin:8px 0; }
.result-name { color:#111827; font-size:15px; font-weight:800; }
.result-status { color:#2563eb; font-size:13px; font-weight:700; margin-top:2px; }

/* ===== Status badges ===== */
.status-badge { display:inline-block; padding:4px 10px; border-radius:999px; font-size:12px; font-weight:700; }
.status-ok { background:#ecfdf5; color:#15803d; }
.status-warn { background:#fffbeb; color:#b45309; }
.status-error { background:#fef2f2; color:#b91c1c; }
.status-info { background:#eff6ff; color:#1d4ed8; }

/* ===== Hide Streamlit input instructions / password visibility toggle =====
   استخدمنا selector مش معتمد على :has() عشان يشتغل على كل المتصفحات
   وعشان نتجنب ظهور زرار "Upload" مكرر بسبب ترجمة المتصفح أو أي DOM إضافي غريب */
[data-testid="InputInstructions"] { display:none !important; }
[data-testid="stTextInput"] button {
    display:none !important;
    visibility:hidden !important;
    width:0 !important;
    height:0 !important;
    padding:0 !important;
    margin:0 !important;
}

/* ===== Mobile ===== */
@media (max-width: 700px) {
    .block-container { padding: .8rem .7rem 2rem; }
    .hero-title { font-size:23px; }
    .topbar { padding:11px 13px; }
}
</style>
""", unsafe_allow_html=True)

# ===== Session state =====
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "is_admin" not in st.session_state:
    st.session_state.is_admin = False
if "update_locked" not in st.session_state:
    st.session_state.update_locked = False

# ===== Login =====
if not st.session_state.logged_in and not st.session_state.is_admin:
    st.markdown("""
    <div style="text-align:center; margin:42px 0 24px;">
        <div style="font-size:48px;">✨</div>
        <div style="font-size:31px;font-weight:800;color:#111827;">Aivora</div>
        <div style="font-size:14px;color:#6b7280;margin-top:3px;">Your Smarter Support for Every Student's Application</div>
    </div>
    """, unsafe_allow_html=True)

    left, right = st.columns([1.15, 1], gap="large")
    with left:
        st.markdown("""
        <div style="padding:40px 20px 20px 10px;">
            <div style="font-size:14px;color:#2563eb;font-weight:800;margin-bottom:8px;">حل بسيط لإدارة المتابعة</div>
            <div style="font-size:31px;font-weight:800;color:#111827;line-height:1.35;">تابع طلبات طلابك<br>من مكان واحد.</div>
            <div style="font-size:15px;color:#6b7280;line-height:1.9;margin-top:12px;max-width:480px;">
                حدّث حالات الطلبات، اربط Google Sheets، وابحث عن أي طالب بسرعة بدون متابعة يدوية.
            </div>
            <div style="margin-top:22px;color:#374151;font-size:14px;line-height:2.2;">
                ✓ تحديث حالات الطلاب بشكل منظم<br>
                ✓ حفظ مصدر البيانات للمكتب<br>
                ✓ بحث سريع عن حالة أي طالب
            </div>
        </div>
        """, unsafe_allow_html=True)

    with right:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        tab1, tab2 = st.tabs(["تسجيل الدخول", "حساب جديد"])
        with tab1:
            st.markdown("<div style='font-size:22px;font-weight:800;color:#111827;margin:8px 0 3px;'>مرحبًا بعودتك 👋</div><div style='color:#6b7280;font-size:13px;margin-bottom:18px;'>سجل دخولك لإدارة طلبات الطلاب</div>", unsafe_allow_html=True)
            username = st.text_input("اسم المكتب", key="login_user", placeholder="اكتب اسم المكتب")
            password = st.text_input("كلمة المرور", type="password", key="login_pass", placeholder="اكتب كلمة المرور")
            if st.button("تسجيل الدخول", key="login_btn"):
                if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
                    st.session_state.is_admin = True
                    st.rerun()
                else:
                    ok, msg = check_login(username, password)
                    if ok:
                        st.session_state.logged_in = True
                        st.session_state.office = username
                        log_to_sheet(username, "تسجيل دخول")
                        st.rerun()
                    else:
                        st.error(msg)
        with tab2:
            st.markdown("<div style='font-size:22px;font-weight:800;color:#111827;margin:8px 0 3px;'>إنشاء حساب</div><div style='color:#6b7280;font-size:13px;margin-bottom:18px;'>سجل مكتبك لبدء استخدام الخدمة</div>", unsafe_allow_html=True)
            new_office = st.text_input("اسم المكتب", key="reg_office", placeholder="اسم المكتب")
            new_email = st.text_input("الإيميل", key="reg_email", placeholder="example@email.com")
            new_pass = st.text_input("كلمة المرور", type="password", key="reg_pass", placeholder="كلمة المرور")
            new_pass2 = st.text_input("تأكيد كلمة المرور", type="password", key="reg_pass2", placeholder="أعد كتابة كلمة المرور")
            if st.button("إنشاء الحساب", key="reg_btn"):
                if not new_office or not new_email or not new_pass:
                    st.error("املأ كل الحقول!")
                elif new_pass != new_pass2:
                    st.error("كلمة المرور مش متطابقة!")
                else:
                    ok, msg = register_office(new_office, new_email, new_pass)
                    if ok:
                        st.success(msg)
                    else:
                        st.error(msg)
        st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# ===== Top bar =====
st.markdown("""
<div class="topbar">
    <div class="brand">
        <div class="brand-icon">✨</div>
        <div><div class="brand-title">Aivora</div><div class="brand-sub">Your Smarter Support for Every Student's Application</div></div>
    </div>
    <div style="font-size:13px;color:#6b7280;">نظام متابعة المكاتب</div>
</div>
""", unsafe_allow_html=True)

# ===== Admin =====
if st.session_state.is_admin:
    st.markdown("<div class='hero'><div class='hero-kicker'>الإدارة</div><div class='hero-title'>لوحة الإدارة</div><div class='hero-desc'>إدارة الحسابات الجديدة ومتابعة حالة المكاتب.</div></div>", unsafe_allow_html=True)
    pending = get_pending_accounts()
    st.markdown(f"<div class='section-title'>الحسابات المعلقة <span style='color:#2563eb'>({len(pending)})</span></div>", unsafe_allow_html=True)
    if pending:
        for acc in pending:
            c1, c2, c3, c4 = st.columns([3, 3, 1, 1])
            c1.write(acc.get("اسم المكتب", ""))
            c2.write(acc.get("الإيميل", ""))
            if c3.button("قبول", key=f"approve_{acc.get('اسم المكتب')}"):
                approve_account(acc.get("اسم المكتب")); st.rerun()
            if c4.button("رفض", key=f"reject_{acc.get('اسم المكتب')}"):
                reject_account(acc.get("اسم المكتب")); st.rerun()
    else:
        st.info("لا توجد حسابات في انتظار الموافقة.")

    st.markdown("<div class='section-title' style='margin-top:28px;'>كل الحسابات</div>", unsafe_allow_html=True)
    try:
        sheet = get_accounts_sheet()
        if sheet:
            all_accounts = sheet.get_all_records()
            if all_accounts:
                df = pd.DataFrame(all_accounts).drop(columns=["الباسورد"], errors="ignore")
                st.dataframe(df, use_container_width=True, hide_index=True)
    except Exception:
        pass
    if st.button("تسجيل الخروج من الإدارة"):
        st.session_state.is_admin = False
        st.rerun()
    st.stop()

# ===== Main dashboard =====
_hour = datetime.now().hour
_greeting = "صباح الخير" if _hour < 12 else "مساء الخير"
office = st.session_state.office

st.markdown(f"""
<div class="hero">
    <div class="hero-kicker">{_greeting} 👋</div>
    <div class="hero-title">أهلاً بيك، <strong>{office}</strong></div>
    <div class="hero-desc">تابع طلبات طلابك وحدّث الحالات من مكان واحد.</div>
</div>
""", unsafe_allow_html=True)

saved_link = get_gsheet_link(office)

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)

# ===== Data source card =====
st.markdown("<div class='card'>", unsafe_allow_html=True)
st.markdown("<div class='section-title'>مصدر بيانات الطلاب</div><div class='section-sub'>اختر الطريقة التي يحتوي بها ملف الطلاب.</div>", unsafe_allow_html=True)
source_options = ["📂 رفع ملف Excel", "🔗 ربط Google Sheets"]
source = st.radio("", source_options, horizontal=True, label_visibility="collapsed")

file_bytes = None
sheet_id_source = None
filename = ""

if source == "📂 رفع ملف Excel":
    uploaded = st.file_uploader("ارفع ملف Excel", type=["xlsx", "xls"], label_visibility="collapsed")
    if uploaded:
        file_bytes = uploaded.read()
        filename = uploaded.name
        st.success(f"تم اختيار الملف: {uploaded.name}")

else:
    if saved_link:
        st.markdown(f"<div style='background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:11px 14px;color:#166534;font-size:13px;'>✓ Google Sheets متصل بالفعل لهذا المكتب</div>", unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        change = st.checkbox("تغيير رابط الشيت")
        if change:
            new_link = st.text_input("رابط Google Sheets", placeholder="الصق الرابط هنا")
            if st.button("حفظ الرابط"):
                sid_check = extract_sheet_id(new_link) if new_link else None
                if not new_link:
                    st.error("أدخل الرابط أولاً!")
                elif not sid_check:
                    st.error("الرابط غير صحيح!")
                else:
                    ok, msg = save_gsheet_link(office, new_link)
                    if ok:
                        st.success("تم حفظ الرابط!"); st.rerun()
                    else:
                        st.error(msg)
        sheet_id_source = extract_sheet_id(saved_link)
    else:
        new_link = st.text_input("رابط Google Sheets", placeholder="https://docs.google.com/spreadsheets/d/...")
        if st.button("حفظ وربط الشيت"):
            sid = extract_sheet_id(new_link) if new_link else None
            if sid:
                ok, msg = save_gsheet_link(office, new_link)
                if ok:
                    st.success("تم حفظ الرابط بنجاح!"); st.rerun()
                else:
                    st.error(msg)
            else:
                st.error("الرابط غير صحيح!")

    if sheet_id_source:
        update_option = st.radio("طريقة التحديث", ["🔄 تحديث من الشيت الأونلاين", "📂 رفع Excel جديد بدلاً عنه"], horizontal=True)
        if update_option == "🔄 تحديث من الشيت الأونلاين":
            if st.button("تحميل بيانات الشيت"):
                with st.spinner("بيقرأ الشيت..."):
                    result, err = read_gsheet_as_excel(saved_link)
                    if err:
                        st.error(f"خطأ: {err}"); st.stop()
                    file_bytes = result.read(); filename = "google_sheet"
                    st.session_state.pending_file_bytes = file_bytes
                    st.session_state.pending_filename = filename
                    st.success("تم جلب بيانات الشيت. البيانات جاهزة للتحديث.")
        else:
            uploaded2 = st.file_uploader("ارفع ملف Excel الجديد", type=["xlsx", "xls"], label_visibility="collapsed")
            if uploaded2:
                file_bytes = uploaded2.read(); filename = uploaded2.name
                st.session_state.pending_file_bytes = file_bytes
                st.session_state.pending_filename = filename
st.markdown('</div>', unsafe_allow_html=True)

if not file_bytes and st.session_state.get("pending_file_bytes"):
    file_bytes = st.session_state.pending_file_bytes
    filename = st.session_state.get("pending_filename", filename)

# ===== Processing =====
if file_bytes:
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>تحديث حالات الطلاب</div><div class='section-sub'>اضغط الزر لبدء فحص الطلبات وتحديث النتائج.</div>", unsafe_allow_html=True)
    if st.session_state.update_locked:
        st.info("🔒 تم تشغيل تحديث بالفعل في هذه الجلسة. لو عايزة تبدئي تحديث جديد، سجّلي خروج وادخلي تاني.")
    elif st.button("▶ تحديث حالات الطلاب", key="start_main"):
        st.session_state.update_locked = True
        log_to_sheet(office, "رفع ملف", filename)

        # =====================================================================
        # بننشئ المهمة على طول من غير ما نتأكد الأول إن فيه Worker شغال، وبعدين
        # نستنى مهلة قصيرة (45 ثانية) نشوف هل حد هياخدها. لو محدش أخدها خلال
        # المهلة، نلغيها ونشتغل بالطريقة العادية (Cloud) على طول.
        # =====================================================================
        if sheet_id_source and source == "🔗 ربط Google Sheets":
            job_id = create_job(office, "gsheet", saved_link, filename or "google_sheet")
        else:
            pending_file_id, up_err = upload_pending_file(file_bytes, filename, office)
            if up_err:
                st.error(up_err); st.stop()
            job_id = create_job(office, "drive_pending", pending_file_id, filename)

        worker_online = False
        if job_id:
            with st.spinner("بنتأكد هل فيه جهاز محلي متصل هياخد التحديث..."):
                worker_online, _ = wait_for_job_pickup(job_id)
            if not worker_online:
                cancel_job(job_id)

        # =====================================================================
        # المسار الأول: فيه Worker شغال على جهاز محلي → نحوّل الشغل ليه
        # (نفس الشكل بالظبط قدام المكتب: اسم الطالب + حالته الجديدة تظهر أول
        # بأول، والشيت بيتحدث تلقائيًا بنفس الطريقة، مفيش أي فرق ظاهر للمكتب)
        # =====================================================================
        if worker_online:
            st.markdown(
                "<div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;"
                "padding:10px 14px;color:#1d4ed8;font-size:13px;margin-bottom:10px;'>"
                "🖥️ التحديث شغال دلوقتي عن طريق جهاز محلي متصل (أسرع وأأمن ضد الحظر)</div>",
                unsafe_allow_html=True,
            )

            progress = st.progress(0)
            status_caption = st.empty()
            table_placeholder = st.empty()
            seen_count = 0
            job = None
            max_wait_seconds = 60 * 60  # حد أقصى للانتظار (ساعة) كحماية من التعليق الأبدي
            waited = 0
            poll_interval = 5  # كل 5 ثواني — توفير للـ API Quota
            while waited < max_wait_seconds:
                job = get_job(job_id)
                rows = get_job_progress_rows(job_id)
                if rows:
                    latest = rows[-1]
                    try:
                        total_from_worker = int(latest.get("total", 0)) or len(rows)
                    except Exception:
                        total_from_worker = len(rows)
                    status_caption.markdown(
                        f"طالب {len(rows)} من {total_from_worker}"
                    )
                    # جدول بيتزود سطر بسطر - كل طالب اتفحص وحالته بيظهروا فورًا
                    # عشان المكتب يكون مطمن إن التحديث شغال فعليًا
                    table_df = pd.DataFrame(
                        [{"اسم الطالب": r.get("اسم الطالب", ""), "الحالة": r.get("الحالة", "")} for r in reversed(rows)]
                    )
                    table_placeholder.dataframe(table_df, use_container_width=True, hide_index=True)
                    progress.progress(min(len(rows) / max(total_from_worker, 1), 1.0))
                    seen_count = len(rows)

                job_status = (job or {}).get("الحالة", "")
                if job_status in ("done", "failed"):
                    break
                time.sleep(poll_interval)
                waited += poll_interval

            if job and job.get("الحالة") == "done":
                st.session_state.pending_file_bytes = None
                st.session_state.pending_filename = ""
                st.markdown(
                    f"<div style='background:#ecfdf5;border:1px solid #bbf7d0;border-radius:14px;padding:16px;margin-top:15px;'>"
                    f"<div style='font-size:17px;font-weight:800;color:#166534;'>اكتمل التحديث 🎉</div>"
                    f"<div style='color:#166534;font-size:13px;margin-top:4px;'>إجمالي {seen_count} طالب اتحدث عن طريق الجهاز المحلي</div></div>",
                    unsafe_allow_html=True,
                )
                final_file_id = job.get("final_drive_file_id", "")
                if final_file_id:
                    final_bytes, dl_err = download_drive_file_bytes(final_file_id)
                    if final_bytes:
                        st.download_button(
                            label="⬇ تحميل ملف Excel المحدث",
                            data=final_bytes,
                            file_name="students_updated.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                log_to_sheet(office, "اكتمل المعالجة عبر Worker", filename)
            elif job and job.get("الحالة") == "failed":
                st.error(f"فشل التحديث عن طريق الجهاز المحلي: {job.get('خطأ','')}")
            else:
                st.warning("استنينا فترة طويلة ومفيش رد من الجهاز المحلي — جربي تاني أو استخدمي التحديث العادي.")

        # =====================================================================
        # المسار الثاني (الأصلي): مفيش Worker شغال → نفس الطريقة القديمة بالظبط
        # =====================================================================
        if not worker_online:
            is_gsheet_source = bool(sheet_id_source and source == "🔗 ربط Google Sheets")

            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            try:
                cols, header_row_num = find_excel_columns(ws)
            except Exception as e:
                st.error(str(e)); st.stop()

            # لو المصدر Google Sheets، وعمود "حالة الطلب الجديدة" ده كان جديد (مش
            # موجود أصلاً)، سجّلي هيدره في شيت المكتب نفسه مرة واحدة بس. ده العمود
            # الوحيد اللي بنكتب فيه، ومفيش أي عمود أو صف تاني بيتلمس أو يتغير ترتيبه.
            if is_gsheet_source:
                ensure_status_header_in_gsheet(saved_link, header_row_num, cols["status"])

            # استبعاد الطلاب اللي حالتهم المحفوظة "نهائية" (زي قبول نهائي) —
            # مفيش داعي نعيد فحصهم كل مرة
            previous_by_name = {}
            try:
                _all_found, _e = search_results_in_sheet(office, "")
                for r in (_all_found or []):
                    previous_by_name[str(r.get("name", "")).strip()] = str(r.get("status", "")).strip()
            except Exception:
                previous_by_name = {}

            rows_data = list(ws.iter_rows(min_row=header_row_num + 1, values_only=False))
            all_valid_rows = [r for r in rows_data if r[cols["email"]].value and r[cols["password"]].value]

            valid_rows = []
            for r in all_valid_rows:
                nm = str(r[cols["name"]].value or "").strip() if cols["name"] is not None else ""
                if previous_by_name.get(nm) in FINAL_STATUSES:
                    continue
                valid_rows.append(r)

            total = len(valid_rows)

            # ترتيب المعالجة (اللوجين على موقع الحكومة) بيتشافل في كل مرة — ده بس
            # ترتيب تسجيل الدخول في دماغ الأجنت، وما بيغيرش ترتيب أو مكان أي صف في
            # شيت المكتب؛ كل طالب بيتكتب في نفس صفه الأصلي بالظبط.
            processing_order = list(range(total))
            random.shuffle(processing_order)

            progress = st.progress(0)
            status_caption = st.empty()
            table_placeholder = st.empty()
            live_table_rows = []
            success = failed = 0
            # النتائج اللي خلصت فعلاً — بتتحفظ أول بأول في شيت "results" بعد كل طالب،
            # فلو التحديث اتوقف فجأة (قفل التاب / نت وقع)، اللي اتحدث فعلاً يفضل محفوظ ومتسجل.
            processed_results = []

            for done_count, pos in enumerate(processing_order):
                row = valid_rows[pos]
                email = str(row[cols["email"]].value).strip()
                password = str(row[cols["password"]].value).strip()
                name = row[cols["name"]].value if cols["name"] is not None else ""
                display_name = name or email
                row_num_in_sheet = row[cols["email"]].row  # الصف الحقيقي في الشيت (ثابت، مش بيتغير)

                status_caption.markdown(f"جاري فحص: **{display_name}** — طالب {done_count+1} من {total}")

                session, csrf_token, err = api_login(email, password)
                if err or not session:
                    # بنسيب سبب الفشل الحقيقي جوه الحالة (كود الخطأ من الموقع) عشان تقدري تفرقي
                    # بين فشل باسورد/إيميل غلط وبين حظر/rate-limit مؤقت من موقع الحكومة نفسه
                    current_status = f"فشل تسجيل الدخول ({err})" if err else "فشل تسجيل الدخول"
                    if cols["status"] is not None: row[cols["status"]].value = current_status
                    failed += 1
                else:
                    app_num, current_status = get_status(session, csrf_token)
                    if cols["status"] is not None: row[cols["status"]].value = current_status
                    api_logout(session)
                    success += 1

                # ضيفي سطر الطالب ده لأعلى الجدول المتزايد قدام المكتب فورًا
                live_table_rows.insert(0, {"اسم الطالب": str(display_name), "الحالة": str(current_status)})
                table_placeholder.dataframe(pd.DataFrame(live_table_rows), use_container_width=True, hide_index=True)
                status_caption.markdown(f"طالب {done_count+1} من {total}")

                # سجليه في قايمة النتائج المتراكمة واحفظيها فورًا في شيت النتائج
                processed_results.append({"name": str(display_name), "status": str(current_status)})
                save_results_to_sheet(office, processed_results)

                # لو المصدر شيت Google متربط، اكتبي بس خلية عمود "حالة الطلب الجديدة"
                # بتاعة الصف ده — من غير ما تمسحي أو تعيدي كتابة أي حاجة تانية في الشيت
                if is_gsheet_source:
                    update_status_cell_in_gsheet(saved_link, row_num_in_sheet, cols["status"], current_status)

                progress.progress((done_count + 1) / max(total, 1))
                if done_count < total - 1: human_delay(10, 18)

            # =================================================================
            # إعادة محاولة تلقائية لمين فشل بسبب خطأ اتصال مؤقت (SSL/Timeout/...)
            # مش بسبب إيميل أو باسورد غلط - عشان دول مش فشل حقيقي، مجرد
            # الكونكشن اتقطع لحظتها.
            # =================================================================
            retry_indices = [
                pos for pos in processing_order
                if _is_transient_login_error(valid_rows[pos][cols["status"]].value if cols["status"] is not None else "")
            ]

            if retry_indices:
                status_caption.markdown(
                    f"🔁 بيعيد محاولة **{len(retry_indices)}** طالب فشلوا بسبب مشكلة اتصال مؤقتة..."
                )
                for r_i, pos in enumerate(retry_indices):
                    row = valid_rows[pos]
                    email = str(row[cols["email"]].value).strip()
                    password = str(row[cols["password"]].value).strip()
                    name = row[cols["name"]].value if cols["name"] is not None else ""
                    display_name = name or email
                    row_num_in_sheet = row[cols["email"]].row

                    human_delay(5, 10)
                    session, csrf_token, err = api_login(email, password)
                    if err or not session:
                        current_status = f"فشل تسجيل الدخول ({err})" if err else "فشل تسجيل الدخول"
                    else:
                        app_num, current_status = get_status(session, csrf_token)
                        api_logout(session)

                    if cols["status"] is not None:
                        row[cols["status"]].value = current_status

                    # حدّثي سطر الطالب ده في الجدول المعروض (مش إضافة سطر جديد،
                    # تحديث نفس السطر بحالته الجديدة بعد إعادة المحاولة)
                    for lr in live_table_rows:
                        if lr["اسم الطالب"] == str(display_name):
                            lr["الحالة"] = str(current_status)
                            break
                    table_placeholder.dataframe(pd.DataFrame(live_table_rows), use_container_width=True, hide_index=True)
                    status_caption.markdown(
                        f"🔁 إعادة محاولة {r_i+1} من {len(retry_indices)}"
                    )

                    # حدّثي النتيجة المتراكمة لنفس الطالب بدل ما نضيف سطر مكرر
                    for r in processed_results:
                        if r["name"] == str(display_name):
                            r["status"] = str(current_status)
                            break
                    save_results_to_sheet(office, processed_results)

                    if is_gsheet_source:
                        update_status_cell_in_gsheet(saved_link, row_num_in_sheet, cols["status"], current_status)

            out = io.BytesIO(); wb.save(out); out.seek(0)
            if is_gsheet_source:
                st.success("تم تحديث عمود 'حالة الطلب الجديدة' في شيت المكتب أول بأول أثناء التحديث!")
            up_ok, up_msg = upload_to_drive(out.getvalue(), filename, office)
            if up_ok: st.caption("تم حفظ نسخة احتياطية على Drive")
            log_to_sheet(office, "اكتمل المعالجة", filename)
            st.session_state.pending_file_bytes = None
            st.session_state.pending_filename = ""
            # النتائج اتحفظت أول بأول أثناء الحلقة، فمفيش داعي نحفظها تاني هنا
            st.session_state.last_results = processed_results
            st.markdown(f"<div style='background:#ecfdf5;border:1px solid #bbf7d0;border-radius:14px;padding:16px;margin-top:15px;'><div style='font-size:17px;font-weight:800;color:#166534;'>اكتمل التحديث 🎉</div><div style='color:#166534;font-size:13px;margin-top:4px;'>إجمالي {total} طالب · نجح {success} · فشل {failed}</div></div>", unsafe_allow_html=True)
            st.download_button(label="⬇ تحميل ملف Excel المحدث", data=out, file_name="students_updated.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.markdown('</div>', unsafe_allow_html=True)

# ===== Search =====
st.markdown("<div class='card'>", unsafe_allow_html=True)
st.markdown("<div class='section-title'>البحث عن طالب</div><div class='section-sub'>اكتب اسم الطالب لمعرفة آخر حالة محفوظة.</div>", unsafe_allow_html=True)
search_query = st.text_input("اسم الطالب", placeholder="مثال: Ahmed Mohamed", label_visibility="collapsed")
if search_query:
    with st.spinner("بيبحث..."):
        found, err = search_results_in_sheet(office, search_query)
    if err:
        st.warning(err)
    elif found:
        for r in found:
            status = r.get("status", "")
            st.markdown(f"<div class='result-card'><div class='result-name'>👤 {r.get('name','')}</div><div class='result-status'>● {status}</div></div>", unsafe_allow_html=True)
    else:
        st.info("مفيش طالب بالاسم ده.")
st.markdown('</div>', unsafe_allow_html=True)

# ===== Footer / logout =====
st.markdown("<div style='height:5px'></div>", unsafe_allow_html=True)
if st.button("تسجيل الخروج", key="logout_main"):
    log_to_sheet(st.session_state.get("office", ""), "تسجيل خروج")
    st.session_state.clear()
    st.rerun()

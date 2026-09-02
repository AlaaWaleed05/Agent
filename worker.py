"""Aivora Selenium Worker.

Claims pending jobs from Supabase, checks student application status with Selenium,
and writes one live progress row per student.

Excel files and Google Sheets are updated live after each processed student.

If the office cancels/logs out while a student is being processed, the current
student is allowed to finish, gets logged out, its result is saved to Supabase,
the source is updated, and then the job stops before starting the next student.

Cancelled jobs are never marked as done or failed.
"""

import io
import json
import os
import random
import re
import socket
import smtplib
import time
import traceback
import uuid
from datetime import datetime, timezone
from email.mime.text import MIMEText
from pathlib import Path

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

import openpyxl
from cryptography.fernet import Fernet
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from supabase import Client, create_client


SITE_URL = "https://admission.study-in-egypt.gov.eg"
LOGIN_URL = f"{SITE_URL}/login"
INBOX_URL = f"{SITE_URL}/inbox"

WAIT_TIME = 20

JOB_POLL_INTERVAL_SECONDS = 10

STUDENT_DELAY_MIN, STUDENT_DELAY_MAX = 4, 8

FINAL_STATUSES = {
    "مقبول نهائي",
    "قبول نهائي",
    "تم الرفض",
    "مرفوض نهائيًا",
    "مرفوض نهائيا",
    "مرفوض",
    "خالص",
}

TECH_FAILURE_STATUS = "تعذر فحص الطالب حاليًا"

WORKER_ID = f"{socket.gethostname()}-{uuid.uuid4().hex[:6]}"

CONFIG_FILE = Path(__file__).with_name("service_account.json")

DRIVE_FOLDER_ID = (
    os.environ.get("DRIVE_FOLDER_ID")
    or "12L_qSHBnW4-tfQZRteynInWNBAML016f"
)


# ============================================================
# CONFIG
# ============================================================

def get_config():
    if not CONFIG_FILE.exists():
        return {}

    try:
        return json.loads(
            CONFIG_FILE.read_text(encoding="utf-8")
        )

    except Exception:
        return {}


CONFIG = get_config()


def get_setting(name, required=True):
    value = os.environ.get(name) or CONFIG.get(name)

    if required and not value:
        raise RuntimeError(f"{name} missing")

    return value


def get_supabase() -> Client:
    return create_client(
        get_setting("SUPABASE_URL"),
        get_setting("SUPABASE_SERVICE_ROLE_KEY"),
    )


def get_google_credentials(scopes):
    data = (
        CONFIG.get("gcp_service_account")
        if isinstance(CONFIG.get("gcp_service_account"), dict)
        else None
    )

    if (
        data is None
        and isinstance(CONFIG, dict)
        and CONFIG.get("client_email")
        and CONFIG.get("private_key")
    ):
        data = CONFIG

    raw = (
        os.environ.get("GCP_SERVICE_ACCOUNT_JSON")
        or os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    )

    if data is None and raw:
        data = json.loads(raw)

    if data:
        return Credentials.from_service_account_info(
            data,
            scopes=scopes,
        )

    credentials_file = os.environ.get(
        "GOOGLE_APPLICATION_CREDENTIALS"
    )

    if credentials_file:
        return Credentials.from_service_account_file(
            credentials_file,
            scopes=scopes,
        )

    raise RuntimeError(
        "Google service account configuration missing"
    )


def drive_service():
    return build(
        "drive",
        "v3",
        credentials=get_google_credentials(
            ["https://www.googleapis.com/auth/drive"]
        ),
    )


# ============================================================
# GOOGLE DRIVE / EXCEL
# ============================================================

def download_drive_file_bytes(file_id):
    service = drive_service()

    buffer = io.BytesIO()

    request = service.files().get_media(
        fileId=str(file_id)
    )

    downloader = MediaIoBaseDownload(
        buffer,
        request
    )

    done = False

    while not done:
        _, done = downloader.next_chunk()

    buffer.seek(0)

    return buffer.getvalue()


def upload_drive_file_bytes(file_bytes, filename):
    service = drive_service()

    metadata = {
        "name": str(filename),
        "parents": [DRIVE_FOLDER_ID],
    }

    media = MediaIoBaseUpload(
        io.BytesIO(file_bytes),
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        resumable=True,
    )

    return (
        service.files()
        .create(
            body=metadata,
            media_body=media,
            fields="id",
        )
        .execute()["id"]
    )


def find_status_column(ws, header_row):
    for col_idx, cell in enumerate(
        ws[header_row],
        start=1
    ):
        value = str(
            cell.value or ""
        ).strip().lower()

        if (
            value in {"حالة الطلب", "الحالة"}
            or (
                "حالة" in value
                and "اسم" not in value
                and "خدمة" not in value
            )
        ):
            return col_idx

    new_col = ws.max_column + 1
    ws.cell(header_row, new_col).value = "حالة الطلب"
    return new_col


def build_updated_excel(file_bytes, students):
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        data_only=False
    )

    ws = wb.active

    cols, header_row = find_excel_columns_for_output(ws)

    status_col = find_status_column(
        ws,
        header_row
    )

    by_login = {}
    by_row = {}

    for student in students:

        login = str(
            student.get("login_identifier")
            or ""
        ).strip().lower()

        status = str(
            student.get("application_status")
            or ""
        ).strip()

        if login and status:
            by_login[login] = status

        if (
            student.get("source_row_number")
            and status
        ):
            by_row[
                int(student["source_row_number"])
            ] = status

    email_col = cols.get("email")

    for row_idx in range(
        header_row + 1,
        ws.max_row + 1
    ):

        status = None

        if email_col is not None:

            email = str(
                ws.cell(
                    row_idx,
                    email_col + 1
                ).value or ""
            ).strip().lower()

            status = by_login.get(email)

        if status is None:
            status = by_row.get(row_idx)

        if status is not None:
            ws.cell(
                row_idx,
                status_col
            ).value = status

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    return output.getvalue()


def find_excel_columns_for_output(ws):
    cols = {
        "name": None,
        "email": None,
        "password": None,
    }

    header_row = None

    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=min(10, ws.max_row),
            values_only=True
        ),
        start=1
    ):

        values = [
            str(c).strip()
            if c is not None
            else ""
            for c in row
        ]

        if any(
            "يميل" in v
            or "mail" in v.lower()
            or "بريد" in v
            for v in values
        ):

            header_row = row_idx

            for i, cell in enumerate(values):

                low = cell.lower()

                if (
                    any(
                        k in cell
                        for k in [
                            "اسم",
                            "الإسم",
                            "الاسم"
                        ]
                    )
                    or "name" in low
                ):
                    cols["name"] = i

                elif (
                    any(
                        k in cell
                        for k in [
                            "يميل",
                            "بريد"
                        ]
                    )
                    or "mail" in low
                ):
                    cols["email"] = i

                elif any(
                    k in cell
                    for k in [
                        "باسورد",
                        "كلمة المرور",
                        "password",
                        "pass"
                    ]
                ):
                    cols["password"] = i

            break

    if (
        header_row is None
        or cols["email"] is None
        or cols["password"] is None
    ):
        raise RuntimeError(
            "excel_columns_missing"
        )

    if cols["name"] is None:
        cols["name"] = cols["email"]

    return cols, header_row


# ============================================================
# GOOGLE SHEET - OLD BULK FUNCTION KEPT
# ============================================================

def update_google_sheet_statuses(source_url, students):
    import gspread

    credentials = get_google_credentials([
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ])

    client = gspread.authorize(credentials)

    match = re.search(
        r"/spreadsheets/d/([a-zA-Z0-9-_]+)",
        str(source_url)
    )

    if not match:
        raise RuntimeError(
            "invalid_google_sheet_url"
        )

    spreadsheet = client.open_by_key(
        match.group(1)
    )

    gid_match = re.search(
        r"[?#&]gid=(\d+)",
        str(source_url)
    )

    gid = (
        int(gid_match.group(1))
        if gid_match
        else None
    )

    worksheet = (
        next(
            (
                w
                for w in spreadsheet.worksheets()
                if w.id == gid
            ),
            spreadsheet.sheet1
        )
        if gid is not None
        else spreadsheet.sheet1
    )

    values = worksheet.get_all_values()

    if not values:
        raise RuntimeError(
            "google_sheet_empty"
        )

    header_idx = None
    email_idx = None
    status_idx = None

    for r_idx, row in enumerate(
        values[:10]
    ):

        normalized = [
            str(v or "").strip().lower()
            for v in row
        ]

        if any(
            "يميل" in v
            or "mail" in v
            or "بريد" in v
            for v in normalized
        ):

            header_idx = r_idx

            for i, value in enumerate(
                normalized
            ):

                if (
                    "يميل" in value
                    or "mail" in value
                    or "بريد" in value
                ):
                    email_idx = i

                if (
                    value in {
                        "حالة الطلب",
                        "الحالة"
                    }
                    or (
                        "حالة" in value
                        and "اسم" not in value
                        and "خدمة" not in value
                    )
                ):
                    status_idx = i

            break

    if (
        header_idx is None
        or email_idx is None
        or status_idx is None
    ):
        raise RuntimeError(
            "google_sheet_columns_missing"
        )

    by_login = {
        str(
            s.get("login_identifier")
            or ""
        ).strip().lower():
            str(
                s.get("application_status")
                or ""
            ).strip()
        for s in students
    }

    for row_idx in range(
        header_idx + 1,
        len(values)
    ):

        login = str(
            values[row_idx][email_idx]
            if email_idx < len(values[row_idx])
            else ""
        ).strip().lower()

        if (
            login in by_login
            and by_login[login]
        ):
            worksheet.update_cell(
                row_idx + 1,
                status_idx + 1,
                by_login[login]
            )


# ============================================================
# OLD FINALIZE FUNCTION KEPT
#
# IMPORTANT:
# process_job() DOES NOT CALL THIS ANYMORE.
# ============================================================

def finalize_job_output(job):
    students = get_students_for_job(job)

    source_type = str(
        job.get("source_type")
        or ""
    )

    if source_type == "excel":

        source_ref = str(
            job.get("source_ref")
            or ""
        ).strip()

        if not source_ref:
            raise RuntimeError(
                "excel_source_missing"
            )

        original = download_drive_file_bytes(
            source_ref
        )

        updated = build_updated_excel(
            original,
            students
        )

        final_id = upload_drive_file_bytes(
            updated,
            job.get("file_name")
            or "students_updated.xlsx"
        )

        db.table("jobs").update({
            "final_drive_file_id": final_id,
            "error": None
        }).eq(
            "id",
            job["id"]
        ).execute()

        print(
            f"Final Excel uploaded for job "
            f"{job['id']}: {final_id}"
        )

        return final_id

    if source_type == "google_sheet":

        source_url = str(
            job.get("source_ref")
            or ""
        ).strip()

        if not source_url:
            raise RuntimeError(
                "google_sheet_source_missing"
            )

        update_google_sheet_statuses(
            source_url,
            students
        )

        print(
            f"Google Sheet updated for job "
            f"{job['id']}"
        )

        return None

    raise RuntimeError(
        f"unsupported_source_type:{source_type}"
    )


# ============================================================
# DATABASE
# ============================================================

db = get_supabase()


def now_iso():
    return datetime.now(
        timezone.utc
    ).isoformat()


def decrypt_student_password(value):
    key = get_setting(
        "STUDENT_PASSWORD_ENCRYPTION_KEY"
    )

    return Fernet(
        key.encode()
    ).decrypt(
        str(value).encode()
    ).decode()


def claim_next_pending_job():
    # This is the actual RPC created for
    # Aivora's atomic queue.

    response = db.rpc(
        "claim_next_pending_job",
        {
            "p_worker_id": WORKER_ID
        }
    ).execute()

    data = response.data

    if isinstance(data, dict):

        return (
            data
            if data.get("id")
            else None
        )

    if isinstance(data, list):

        for row in data:

            if (
                isinstance(row, dict)
                and row.get("id")
            ):
                return row

    return None


def set_job_status(
    job_id,
    status,
    error=None
):
    payload = {
        "status": status
    }

    if status == "processing":
        payload["started_at"] = now_iso()

    if status in {
        "done",
        "failed"
    }:
        payload["finished_at"] = now_iso()

    if error is not None:
        payload["error"] = str(error)[:1000]

    db.table("jobs").update(
        payload
    ).eq(
        "id",
        job_id
    ).execute()


# ============================================================
# JOB STATUS CHECK
# ============================================================

def get_job_status(job_id):
    try:

        rows = (
            db.table("jobs")
            .select("status")
            .eq("id", job_id)
            .limit(1)
            .execute()
            .data
            or []
        )

        if not rows:
            return None

        return str(
            rows[0].get("status")
            or ""
        ).strip().lower()

    except Exception as exc:

        print(
            f"⚠️ Could not check job status "
            f"{job_id}: {exc}"
        )

        # If the status check itself fails,
        # do not assume cancellation.
        return None


def job_is_cancelled(job_id):
    status = get_job_status(job_id)

    return status == "cancelled"


def job_is_active(job_id):
    status = get_job_status(job_id)

    return status == "processing"


def mark_job_done_if_active(job_id):
    """
    Mark the job done ONLY if it is still processing.

    This prevents a race where Streamlit cancels the job
    between our status check and the final update.
    """

    stamp = now_iso()

    response = (
        db.table("jobs")
        .update({
            "status": "done",
            "finished_at": stamp,
        })
        .eq("id", job_id)
        .eq("status", "processing")
        .select("id,status")
        .execute()
    )

    rows = response.data or []

    return bool(rows)


def mark_job_failed_if_active(
    job_id,
    error
):
    """
    Mark failed ONLY if the job is still processing.

    If Streamlit already changed it to cancelled,
    this update does nothing.
    """

    stamp = now_iso()

    response = (
        db.table("jobs")
        .update({
            "status": "failed",
            "finished_at": stamp,
            "error": str(error)[:1000],
        })
        .eq("id", job_id)
        .eq("status", "processing")
        .select("id,status")
        .execute()
    )

    rows = response.data or []

    return bool(rows)


def append_progress(
    job_id,
    index,
    total,
    student_name,
    status
):
    db.table("job_progress").insert({
        "job_id": job_id,
        "student_index": index,
        "total": total,
        "student_name": str(student_name),
        "status": str(status),
    }).execute()


def get_office(office_id):
    rows = (
        db.table("offices")
        .select(
            "id,name,email,status"
        )
        .eq(
            "id",
            office_id
        )
        .limit(1)
        .execute()
        .data
        or []
    )

    return rows[0] if rows else None


def get_students_for_job(job):
    rows = (
        db.table("student_records")
        .select("*")
        .eq(
            "office_id",
            job["office_id"]
        )
        .eq(
            "data_source_id",
            job["data_source_id"]
        )
        .order(
            "source_row_number"
        )
        .execute()
        .data
        or []
    )

    latest = {}

    for row in rows:

        key = str(
            row.get("login_identifier")
            or row.get("student_name")
            or ""
        ).strip().lower()

        if key:
            latest[key] = row

    return list(
        latest.values()
    )


def update_student_status(
    student_id,
    status
):
    stamp = now_iso()

    db.table("student_records").update({
        "application_status": str(status),
        "status_updated_at": stamp,
        "updated_at": stamp,
    }).eq(
        "id",
        student_id
    ).execute()


# ============================================================
# LIVE SOURCE UPDATE
# ============================================================

def update_source_student_status(
    job,
    student,
    status,
    name=""
):
    """
    Update the student's source immediately.

    Excel:
        updates the SAME Drive file.

    Google Sheet:
        updates the SAME Google Sheet.

    This happens after EVERY processed student.
    """

    source_type = str(
        job.get("source_type")
        or ""
    ).strip().lower()

    if source_type == "excel":

        source_ref = str(
            job.get("source_ref")
            or ""
        ).strip()

        if not source_ref:
            raise RuntimeError(
                "excel_source_missing"
            )

        update_excel_student_status(
            source_ref,
            student,
            status
        )

        print(
            f"    ✓ Excel live update: {name}"
        )

        return

    if source_type == "google_sheet":

        source_url = str(
            job.get("source_ref")
            or ""
        ).strip()

        if not source_url:
            raise RuntimeError(
                "google_sheet_source_missing"
            )

        update_google_sheet_student_status(
            source_url,
            student,
            status
        )

        print(
            f"    ✓ Google Sheet live update: {name}"
        )

        return

    raise RuntimeError(
        f"unsupported_source_type:{source_type}"
    )


# ============================================================
# ACTIVITY / EMAIL
# ============================================================

def log_activity(
    office_id,
    action,
    file_name="",
    details=None,
    student_id=None,
    data_source_id=None
):
    try:

        db.table(
            "activity_logs"
        ).insert({
            "office_id": office_id,
            "student_record_id": student_id,
            "data_source_id": data_source_id,
            "action": action,
            "file_name": file_name,
            "details": details or {},
        }).execute()

    except Exception as exc:

        print(
            f"Activity log error: {exc}"
        )


def send_email_notification(
    to_email,
    subject,
    body
):
    sender = get_setting(
        "SENDER_EMAIL",
        required=False
    )

    password = get_setting(
        "SENDER_APP_PASSWORD",
        required=False
    )

    if (
        not sender
        or not password
        or not to_email
    ):
        return

    try:

        msg = MIMEText(
            body,
            "plain",
            "utf-8"
        )

        msg["Subject"] = subject
        msg["From"] = sender
        msg["To"] = to_email

        with smtplib.SMTP_SSL(
            "smtp.gmail.com",
            465
        ) as server:

            server.login(
                sender,
                password
            )

            server.sendmail(
                sender,
                [to_email],
                msg.as_string()
            )

    except Exception as exc:

        print(
            f"Email error: {exc}"
        )


def notify_office_status_changes(
    office,
    previous,
    processed
):
    if (
        not office
        or not office.get("email")
    ):
        return

    changes = []

    for item in processed:

        name = str(
            item.get("name")
            or ""
        ).strip()

        status = str(
            item.get("status")
            or ""
        ).strip()

        if (
            name
            and previous.get(name) != status
        ):
            changes.append(
                (name, status)
            )

    if not changes:
        return

    body = [
        f"تحديث حالات الطلاب - "
        f"{office.get('name', '')}",
        ""
    ]

    body.extend(
        f"{i}. {name} - {status}"
        for i, (name, status)
        in enumerate(
            changes[:200],
            1
        )
    )

    send_email_notification(
        office["email"],
        f"تحديث حالات الطلاب - "
        f"{office.get('name', '')}",
        "\n".join(body)
    )


def notify_developer_error(
    office,
    job_id,
    error
):
    email = get_setting(
        "DEVELOPER_EMAIL",
        required=False
    )

    if email:

        send_email_notification(
            email,
            "Aivora Worker job failed",
            (
                f"job_id: {job_id}\n"
                f"office: "
                f"{office.get('name') if office else ''}\n"
                f"error: {error}"
            )
        )


# ============================================================
# HUMAN BEHAVIOR
# ============================================================

def human_delay(
    a,
    b,
    msg=""
):
    seconds = random.uniform(
        a,
        b
    )

    if msg:

        print(
            f"    {msg} "
            f"({seconds:.1f}s)..."
        )

    time.sleep(seconds)


def human_type(
    element,
    text
):
    element.clear()

    human_delay(
        0.3,
        0.7
    )

    for char in str(text):

        element.send_keys(char)

        time.sleep(
            random.uniform(
                0.05,
                0.15
            )
        )




# ============================================================
# SELENIUM
# ============================================================

def setup_browser():
    options = Options()

    options.add_argument(
        "--start-maximized"
    )

    return webdriver.Chrome(
        options=options
    )


def clear_session(driver):
    try:

        driver.delete_all_cookies()

    except Exception:
        pass

    try:

        driver.execute_script(
            "window.localStorage.clear(); "
            "window.sessionStorage.clear();"
        )

    except Exception:
        pass


def selenium_login(
    driver,
    email,
    password
):
    try:

        clear_session(driver)

        driver.get(
            LOGIN_URL
        )

        wait = WebDriverWait(
            driver,
            WAIT_TIME
        )

        human_delay(
            0.8,
            1.5
        )

        email_field = wait.until(
            EC.presence_of_element_located(
                (
                    By.CSS_SELECTOR,
                    "input[type='email'], "
                    "input[name='email']"
                )
            )
        )

        human_type(
            email_field,
            email
        )

        password_field = wait.until(
            EC.presence_of_element_located(
                (
                    By.CSS_SELECTOR,
                    "input[type='password']"
                )
            )
        )

        human_type(
            password_field,
            password
        )

        login_btn = wait.until(
            EC.element_to_be_clickable(
                (
                    By.CSS_SELECTOR,
                    "button[type='submit'], "
                    "input[type='submit']"
                )
            )
        )

        login_btn.click()

        for _ in range(15):

            time.sleep(1)

            try:

                driver.find_element(
                    By.CSS_SELECTOR,
                    "input[type='email'], "
                    "input[name='email']"
                )

            except Exception:

                return (
                    True,
                    False,
                    None
                )

            if (
                "login"
                not in driver.current_url
            ):

                return (
                    True,
                    False,
                    None
                )

        return (
            False,
            False,
            "login_failed"
        )

    except (
        TimeoutException,
        WebDriverException
    ) as exc:

        return (
            False,
            True,
            str(exc)
        )

    except Exception as exc:

        return (
            False,
            True,
            str(exc)
        )


def selenium_go_to_inbox(driver):
    try:

        wait = WebDriverWait(
            driver,
            WAIT_TIME
        )

        # افتح قائمة المستخدم
        user_menu = wait.until(
            EC.element_to_be_clickable(
                (
                    By.CSS_SELECTOR,
                    "[class*='user'], "
                    "[class*='profile'], "
                    "[class*='avatar'], "
                    "[class*='account']"
                )
            )
        )

        user_menu.click()

        # اضغط "طلباتي"
        my_requests = wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//*[contains("
                    "normalize-space(), "
                    "'طلباتي')]"
                )
            )
        )

        my_requests.click()

        # انتظر الصفحة الجديدة
        human_delay(
            0.7,
            1.5
        )

        return (
            True,
            None
        )

    except (
        TimeoutException,
        WebDriverException
    ) as exc:

        return (
            False,
            str(exc)
        )


def selenium_get_status(driver):
    try:

        wait = WebDriverWait(
            driver,
            WAIT_TIME
        )

        wait.until(
            EC.presence_of_element_located(
                (
                    By.CSS_SELECTOR,
                    "table"
                )
            )
        )

        headers = driver.find_elements(
            By.CSS_SELECTOR,
            "table thead th, table tr th"
        )

        header_texts = [
            h.text.strip()
            for h in headers
        ]

        status_index = next(
            (
                i
                for i, h in enumerate(header_texts)
                if (
                    h in (
                        "حالة الطلب",
                        "الحالة"
                    )
                    or (
                        "حالة" in h
                        and "اسم" not in h
                    )
                )
            ),
            None
        )

        rows = driver.find_elements(
            By.CSS_SELECTOR,
            "table tbody tr"
        )

        statuses = []

        for row in rows:

            cells = row.find_elements(
                By.CSS_SELECTOR,
                "td"
            )

            if not cells:
                continue

            status = (
                cells[status_index].text.strip()
                if (
                    status_index is not None
                    and status_index < len(cells)
                )
                else ""
            )

            if status:
                statuses.append(status)

        return (
            statuses[0]
            if statuses
            else "مفيش طلبات",
            False,
            None
        )

    except (
        TimeoutException,
        WebDriverException
    ) as exc:

        return (
            "",
            True,
            str(exc)
        )

    except Exception as exc:

        return (
            "",
            True,
            str(exc)
        )


def selenium_logout(driver):
    try:

        wait = WebDriverWait(
            driver,
            WAIT_TIME
        )

        user_menu = wait.until(
            EC.element_to_be_clickable(
                (
                    By.CSS_SELECTOR,
                    "[class*='user'], "
                    "[class*='profile'], "
                    "[class*='avatar'], "
                    "[class*='account']"
                )
            )
        )

        user_menu.click()

        logout = wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//*[contains("
                    "text(), "
                    "'تسجيل خروج') "
                    "or contains("
                    "text(), "
                    "'خروج')]"
                )
            )
        )

        logout.click()

        human_delay(
            0.7,
            1.5
        )

    except Exception as exc:

        print(
            f"Logout warning: {exc}"
        )

    finally:

        clear_session(
            driver
        )


def safe_quit(driver):
    if driver is not None:

        try:

            driver.quit()

        except Exception:
            pass


# ============================================================
# LIVE EXCEL UPDATE
# ============================================================

def update_excel_student_status(
    source_ref,
    student,
    status
):
    source_ref = str(
        source_ref or ""
    ).strip()

    if not source_ref:
        raise RuntimeError(
            "excel_source_missing"
        )

    # Download the CURRENT version of the same file.
    file_bytes = download_drive_file_bytes(
        source_ref
    )

    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        data_only=False
    )

    ws = wb.active

    cols, header_row = (
        find_excel_columns_for_output(
            ws
        )
    )

    status_col = find_status_column(
        ws,
        header_row
    )

    email_col = cols.get(
        "email"
    )

    target_row = None

    login = str(
        student.get(
            "login_identifier"
        )
        or ""
    ).strip().lower()

    # First try matching by email/login
    if (
        email_col is not None
        and login
    ):

        for row_idx in range(
            header_row + 1,
            ws.max_row + 1
        ):

            email = str(
                ws.cell(
                    row_idx,
                    email_col + 1
                ).value or ""
            ).strip().lower()

            if email == login:

                target_row = row_idx

                break

    # Fallback to original source row
    if (
        target_row is None
        and student.get(
            "source_row_number"
        )
    ):

        target_row = int(
            student[
                "source_row_number"
            ]
        )

    if target_row is None:

        raise RuntimeError(
            f"excel_student_row_not_found:"
            f"{login}"
        )

    # IMPORTANT:
    # Only update THIS student's status cell.
    # All other students remain untouched.

    ws.cell(
        target_row,
        status_col
    ).value = str(status)

    output = io.BytesIO()

    wb.save(
        output
    )

    output.seek(0)

    service = drive_service()

    media = MediaIoBaseUpload(
        output,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        resumable=True,
    )

    # IMPORTANT:
    # Update the SAME Drive file.
    # Do NOT create a new file.
    (
        service.files()
        .update(
            fileId=source_ref,
            media_body=media,
        )
        .execute()
    )


# ============================================================
# LIVE GOOGLE SHEET UPDATE
# ============================================================

def update_google_sheet_student_status(
    source_url,
    student,
    status
):
    import gspread

    credentials = get_google_credentials([
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ])

    client = gspread.authorize(
        credentials
    )

    match = re.search(
        r"/spreadsheets/d/([a-zA-Z0-9-_]+)",
        str(source_url)
    )

    if not match:

        raise RuntimeError(
            "invalid_google_sheet_url"
        )

    spreadsheet = client.open_by_key(
        match.group(1)
    )

    gid_match = re.search(
        r"[?#&]gid=(\d+)",
        str(source_url)
    )

    gid = (
        int(gid_match.group(1))
        if gid_match
        else None
    )

    worksheet = (
        next(
            (
                w
                for w in spreadsheet.worksheets()
                if w.id == gid
            ),
            spreadsheet.sheet1
        )
        if gid is not None
        else spreadsheet.sheet1
    )

    values = worksheet.get_all_values()

    if not values:

        raise RuntimeError(
            "google_sheet_empty"
        )

    header_idx = None
    email_idx = None
    status_idx = None

    for r_idx, row in enumerate(
        values[:10]
    ):

        normalized = [
            str(v or "").strip().lower()
            for v in row
        ]

        if any(
            "يميل" in v
            or "mail" in v
            or "بريد" in v
            for v in normalized
        ):

            header_idx = r_idx

            for i, value in enumerate(
                normalized
            ):

                if (
                    "يميل" in value
                    or "mail" in value
                    or "بريد" in value
                ):
                    email_idx = i

                if (
                    value in {
                        "حالة الطلب",
                        "الحالة"
                    }
                    or (
                        "حالة" in value
                        and "اسم" not in value
                        and "خدمة" not in value
                    )
                ):
                    status_idx = i

            break

    if (
        header_idx is None
        or email_idx is None
        or status_idx is None
    ):

        raise RuntimeError(
            "google_sheet_columns_missing"
        )

    login = str(
        student.get(
            "login_identifier"
        )
        or ""
    ).strip().lower()

    target_row = None

    for row_idx in range(
        header_idx + 1,
        len(values)
    ):

        current_login = str(
            values[row_idx][email_idx]
            if email_idx < len(values[row_idx])
            else ""
        ).strip().lower()

        if current_login == login:

            target_row = row_idx + 1

            break

    if target_row is None:

        raise RuntimeError(
            f"google_sheet_student_not_found:"
            f"{login}"
        )

    # IMPORTANT:
    # Only update THIS student's status cell.

    worksheet.update_cell(
        target_row,
        status_idx + 1,
        str(status)
    )


# ============================================================
# PROCESS JOB
# ============================================================

def process_job(job):

    job_id = str(
        job["id"]
    )

    office = get_office(
        job.get("office_id")
    )

    driver = None

    processed = []

    try:

        # ========================================================
        # LOAD STUDENTS
        # ========================================================

        students = get_students_for_job(
            job
        )

        # ========================================================
        # IMPORTANT:
        # KEEP RANDOM SHUFFLE.
        #
        # The worker intentionally does NOT process students
        # according to Excel/Google Sheet order.
        # ========================================================

        random.shuffle(
            students
        )

        total = len(
            students
        )

        # ========================================================
        # EMPTY JOB
        # ========================================================

        if not total:

            if job_is_active(
                job_id
            ):

                mark_job_done_if_active(
                    job_id
                )

            return

        # ========================================================
        # PREVIOUS STATUSES
        # ========================================================

        previous = {
            str(
                s.get(
                    "student_name"
                )
                or ""
            ).strip():
                str(
                    s.get(
                        "application_status"
                    )
                    or ""
                ).strip()
            for s in students
            if str(
                s.get(
                    "student_name"
                )
                or ""
            ).strip()
        }

        print(
            f"=== Starting job "
            f"{job_id}: "
            f"{total} students ==="
        )

        # ========================================================
        # BROWSER
        # ========================================================

        driver = setup_browser()

        retry_students = []

        # ========================================================
        # FIRST PASS
        # ========================================================

        for index, student in enumerate(
            students,
            1
        ):

            # ====================================================
            # CHECK BEFORE STARTING A NEW STUDENT
            # ====================================================

            if job_is_cancelled(
                job_id
            ):

                print(
                    f"🛑 Job {job_id} "
                    f"was cancelled before "
                    f"student {index}."
                )

                return

            name = str(
                student.get(
                    "student_name"
                )
                or student.get(
                    "login_identifier"
                )
                or "طالب"
            ).strip()

            current = str(
                student.get(
                    "application_status"
                )
                or ""
            ).strip()

            status = (
                current
                or "لم يتم الفحص بعد"
            )

            technical_error = None

            # ====================================================
            # IMPORTANT:
            #
            # Browser restart is delayed until AFTER:
            #
            # 1. Logout
            # 2. Supabase update
            # 3. Excel/Google Sheet update
            # 4. Progress update
            # 5. Cancellation check
            # ====================================================

            needs_browser_restart = False

            logged_in = False

            # ====================================================
            # PROCESS CURRENT STUDENT
            # ====================================================

            try:

                if True:

                    password = decrypt_student_password(
                        student[
                            "encrypted_password"
                        ]
                    )

                    ok, technical, error = (
                        selenium_login(
                            driver,
                            str(
                                student[
                                    "login_identifier"
                                ]
                            ).strip(),
                            password
                        )
                    )

                    if (
                        not ok
                        and not technical
                    ):

                        status = (
                            "فشل تسجيل الدخول"
                        )

                    elif not ok:

                        technical_error = error

                    else:

                        logged_in = True

                        ok2, error2 = (
                            selenium_go_to_inbox(
                                driver
                            )
                        )

                        if not ok2:

                            technical_error = error2

                        else:

                            (
                                status_text,
                                technical2,
                                error3
                            ) = selenium_get_status(
                                driver
                            )

                            if technical2:

                                technical_error = error3

                            else:

                                status = status_text

            except Exception as exc:

                technical_error = str(
                    exc
                )

            finally:

                # =================================================
                # ALWAYS LOGOUT CURRENT STUDENT
                # =================================================

                if logged_in:

                    try:

                        print(
                            f"    Logging out: "
                            f"{name}"
                        )

                        selenium_logout(
                            driver
                        )

                    except Exception as logout_exc:

                        print(
                            f"⚠️ Logout warning "
                            f"for {name}: "
                            f"{logout_exc}"
                        )

            # ====================================================
            # TECHNICAL FAILURE
            #
            # A Selenium failure is NEVER sent to the API here.
            # Queue the student for one Selenium-only retry pass
            # at the end of this job.
            # ====================================================

            if technical_error:

                status = TECH_FAILURE_STATUS
                retry_students.append((index, student, name))

                print(
                    f"⚠️ Technical error for {name}: "
                    f"{technical_error}; queued for Selenium retry pass"
                )

                needs_browser_restart = False

            # ====================================================
            # SAVE TO SUPABASE
            # ====================================================

            try:

                update_student_status(
                    student["id"],
                    status
                )

            except Exception as exc:

                print(
                    f"❌ Could not update "
                    f"student record "
                    f"{name}: {exc}"
                )

            # ====================================================
            # LIVE SOURCE UPDATE
            # ====================================================

            try:

                update_source_student_status(
                    job,
                    student,
                    status,
                    name
                )

            except Exception as exc:

                print(
                    f"❌ Could not live-update "
                    f"source for {name}: "
                    f"{exc}"
                )

            # ====================================================
            # LIVE PROGRESS
            # ====================================================

            try:

                append_progress(
                    job_id,
                    index,
                    total,
                    name,
                    status
                )

            except Exception as exc:

                print(
                    f"❌ Could not append "
                    f"progress for {name}: "
                    f"{exc}"
                )

            processed.append({
                "name": name,
                "status": status
            })

            print(
                f"    {index}/{total} | "
                f"{name} | {status}"
            )

            # ====================================================
            # CANCELLATION CHECK
            #
            # Current student is now finished.
            # ====================================================

            if job_is_cancelled(
                job_id
            ):

                print(
                    f"🛑 Job {job_id} "
                    f"cancelled after "
                    f"student {index}: "
                    f"{name}"
                )

                print(
                    f"    Processed: "
                    f"{index}/{total}"
                )

                return

            # ====================================================
            # RESTART CHROME ONLY NOW
            #
            # The current student's result is already saved.
            # ====================================================

            if needs_browser_restart:

                print(
                    f"🔄 Restarting Chrome "
                    f"after saving "
                    f"{name}..."
                )

                safe_quit(
                    driver
                )

                driver = None

                # Cancellation must be checked BEFORE
                # creating a new browser.

                if job_is_cancelled(
                    job_id
                ):

                    print(
                        f"🛑 Job {job_id} "
                        f"was cancelled after "
                        f"saving {name}, "
                        f"before Chrome restart."
                    )

                    return

                try:

                    driver = setup_browser()

                    print(
                        f"    ✓ Chrome restarted "
                        f"after {name}"
                    )

                except Exception as restart_exc:

                    print(
                        f"❌ Chrome restart "
                        f"failed after "
                        f"{name}: "
                        f"{restart_exc}"
                    )

                    # Leave driver=None.
                    #
                    # Do NOT touch any unprocessed student.
                    # Retry pass will attempt recovery.

            # ====================================================
            # DELAY BEFORE NEXT STUDENT
            # ====================================================

            if index < total:

                human_delay(
                    STUDENT_DELAY_MIN,
                    STUDENT_DELAY_MAX,
                    "Pause before next student"
                )

                if job_is_cancelled(
                    job_id
                ):

                    print(
                        f"🛑 Job {job_id} "
                        f"cancelled during "
                        f"pause before next student."
                    )

                    return

        # ========================================================
        # RETRY TECHNICAL FAILURES ONCE
        # ========================================================

        if retry_students:

            if job_is_cancelled(
                job_id
            ):

                print(
                    f"🛑 Job {job_id} "
                    f"cancelled before retry pass."
                )

                return

            print(
                f"=== Retrying "
                f"{len(retry_students)} "
                f"technical failure(s) ==="
            )

            if driver is None:

                try:

                    driver = setup_browser()

                except Exception as exc:

                    print(
                        f"❌ Could not start "
                        f"Chrome for retry "
                        f"pass: {exc}"
                    )

                    raise RuntimeError(
                        "chrome_unavailable_for_retry_pass"
                    )

            # ====================================================
            # RETRY EACH STUDENT
            # ====================================================

            for retry_position, (
                index,
                student,
                name
            ) in enumerate(
                retry_students,
                1
            ):

                # =================================================
                # CHECK BEFORE STARTING RETRY
                # =================================================

                if job_is_cancelled(
                    job_id
                ):

                    print(
                        f"🛑 Job {job_id} "
                        f"cancelled before "
                        f"retrying {name}."
                    )

                    return

                retry_status = (
                    TECH_FAILURE_STATUS
                )

                logged_in = False

                # IMPORTANT:
                # This flag controls Chrome restart.
                # It is intentionally set only after the retry
                # has failed technically.
                needs_browser_restart = False

                # =================================================
                # MAKE SURE CHROME EXISTS
                # =================================================

                if driver is None:

                    try:

                        driver = setup_browser()

                    except Exception as exc:

                        print(
                            f"❌ Could not create "
                            f"Chrome before retrying "
                            f"{name}: {exc}"
                        )

                        raise RuntimeError(
                            "chrome_unavailable_during_retry"
                        )

                human_delay(
                    1.0,
                    2.0,
                    f"Retrying {name}"
                )

                # =================================================
                # RETRY CURRENT STUDENT
                # =================================================

                try:

                    password = (
                        decrypt_student_password(
                            student[
                                "encrypted_password"
                            ]
                        )
                    )

                    # -----------------------------------------
                    # RETRY LOGIN
                    # -----------------------------------------

                    ok, technical, error = (
                        selenium_login(
                            driver,
                            str(
                                student[
                                    "login_identifier"
                                ]
                            ).strip(),
                            password
                        )
                    )

                    if (
                        not ok
                        and not technical
                    ):

                        retry_status = (
                            "فشل تسجيل الدخول"
                        )

                    elif not ok:

                        raise RuntimeError(
                            error
                            or "retry_login_failed"
                        )

                    else:

                        logged_in = True

                        # -------------------------------------
                        # RETRY -> MY REQUESTS
                        # -------------------------------------

                        ok2, error2 = (
                            selenium_go_to_inbox(
                                driver
                            )
                        )

                        if not ok2:

                            raise RuntimeError(
                                error2
                                or "retry_inbox_failed"
                            )

                        # -------------------------------------
                        # RETRY -> STATUS
                        # -------------------------------------

                        (
                            retry_status,
                            technical2,
                            error3
                        ) = selenium_get_status(
                            driver
                        )

                        if technical2:

                            raise RuntimeError(
                                error3
                                or "retry_status_failed"
                            )

                except Exception as exc:

                    retry_status = (
                        TECH_FAILURE_STATUS
                    )

                    print(
                        f"⚠️ Retry failed "
                        f"for {name}: "
                        f"{exc}"
                    )

                    # IMPORTANT:
                    # DO NOT restart Chrome here.
                    #
                    # We first save this retry result
                    # everywhere.

                    needs_browser_restart = False

                finally:

                    # =========================================
                    # ALWAYS LOGOUT RETRY STUDENT
                    # =========================================

                    if logged_in:

                        try:

                            print(
                                f"    Logging out "
                                f"after retry: "
                                f"{name}"
                            )

                            selenium_logout(
                                driver
                            )

                        except Exception as logout_exc:

                            print(
                                f"⚠️ Retry logout "
                                f"warning for "
                                f"{name}: "
                                f"{logout_exc}"
                            )

                # =================================================
                # IMPORTANT:
                #
                # NO CHROME RESTART HERE.
                #
                # The retry result must first be saved:
                #
                # 1. Supabase
                # 2. Excel / Google Sheet
                # 3. Progress
                # 4. processed
                # 5. Cancellation check
                #
                # ONLY THEN restart Chrome.
                # =================================================

                # =================================================
                # SAVE RETRY RESULT TO DB
                # =================================================

                try:

                    update_student_status(
                        student["id"],
                        retry_status
                    )

                except Exception as exc:

                    print(
                        f"❌ Retry student "
                        f"update error "
                        f"for {name}: "
                        f"{exc}"
                    )

                # =================================================
                # LIVE SOURCE UPDATE AFTER RETRY
                # =================================================

                try:

                    update_source_student_status(
                        job,
                        student,
                        retry_status,
                        name
                    )

                except Exception as exc:

                    print(
                        f"❌ Retry live source "
                        f"update error "
                        f"for {name}: "
                        f"{exc}"
                    )

                # =================================================
                # RETRY PROGRESS
                # =================================================

                try:

                    append_progress(
                        job_id,
                        index,
                        total,
                        name,
                        retry_status
                    )

                except Exception as exc:

                    print(
                        f"❌ Retry progress "
                        f"error for {name}: "
                        f"{exc}"
                    )

                # =================================================
                # REPLACE OLD RESULT
                # =================================================

                processed = [
                    item
                    for item in processed
                    if item.get("name")
                    != name
                ]

                processed.append({
                    "name": name,
                    "status": retry_status
                })

                print(
                    f"    Retry result: "
                    f"{name} | "
                    f"{retry_status}"
                )

                # =================================================
                # CANCELLATION CHECK
                #
                # Retry student is now COMPLETELY saved.
                # =================================================

                if job_is_cancelled(
                    job_id
                ):

                    print(
                        f"🛑 Job {job_id} "
                        f"cancelled after "
                        f"retrying {name}."
                    )

                    return

                # =================================================
                # NOW RESTART CHROME IF RETRY FAILED TECHNICALLY
                #
                # IMPORTANT:
                # This is AFTER DB + source + progress.
                # =================================================

                if needs_browser_restart:

                    print(
                        f"🔄 Restarting Chrome "
                        f"after saving retry "
                        f"result for {name}..."
                    )

                    safe_quit(
                        driver
                    )

                    driver = None

                    # Check cancellation BEFORE
                    # creating another browser.

                    if job_is_cancelled(
                        job_id
                    ):

                        print(
                            f"🛑 Job {job_id} "
                            f"was cancelled after "
                            f"saving retry result "
                            f"for {name}, "
                            f"before Chrome restart."
                        )

                        return

                    try:

                        driver = setup_browser()

                        print(
                            f"    ✓ Chrome restarted "
                            f"after retry of "
                            f"{name}"
                        )

                    except Exception as restart_exc:

                        print(
                            f"❌ Chrome restart "
                            f"after retry failure "
                            f"failed: "
                            f"{restart_exc}"
                        )

                        # Keep driver=None.
                        #
                        # If there are more retry students,
                        # the job will stop WITHOUT touching them.

                # =================================================
                # IF CHROME FAILED TO RESTART
                #
                # DO NOT TOUCH ANY OTHER STUDENT.
                # =================================================

                if (
                    driver is None
                    and retry_position
                    < len(retry_students)
                ):

                    print(
                        f"❌ Chrome unavailable "
                        f"after retrying "
                        f"{name}."
                    )

                    raise RuntimeError(
                        "chrome_restart_failed_during_retry"
                    )

                # =================================================
                # DELAY BETWEEN RETRIES
                # =================================================

                if (
                    retry_position
                    < len(retry_students)
                ):

                    human_delay(
                        1.0,
                        2.0,
                        "Pause before next retry"
                    )

                    if job_is_cancelled(
                        job_id
                    ):

                        print(
                            f"🛑 Job {job_id} "
                            f"cancelled during "
                            f"retry pause."
                        )

                        return

        # ========================================================
        # IMPORTANT:
        #
        # NO finalize_job_output(job)
        #
        # Excel / Google Sheet have already been updated
        # student-by-student.
        # ========================================================

        # ========================================================
        # FINAL CANCELLATION CHECK
        # ========================================================

        if job_is_cancelled(
            job_id
        ):

            print(
                f"🛑 Job {job_id} "
                f"was cancelled before completion."
            )

            return

        # ========================================================
        # NOTIFY OFFICE
        # ========================================================

        try:

            notify_office_status_changes(
                office,
                previous,
                processed
            )

        except Exception as exc:

            print(
                f"Status email warning: "
                f"{exc}"
            )

        # ========================================================
        # CHECK AGAIN BEFORE ACTIVITY / DONE
        # ========================================================

        if job_is_cancelled(
            job_id
        ):

            print(
                f"🛑 Job {job_id} "
                f"was cancelled before "
                f"finalization."
            )

            return

        # ========================================================
        # ACTIVITY LOG
        # ========================================================

        log_activity(
            job.get("office_id"),
            "اكتمل تحديث حالات الطلاب",
            job.get("file_name")
            or "",
            {
                "job_id": job_id,
                "students_processed": len(
                    processed
                ),
                "worker_id": WORKER_ID
            },
            data_source_id=job.get(
                "data_source_id"
            )
        )

        # ========================================================
        # MARK DONE ONLY IF STILL PROCESSING
        # ========================================================

        marked_done = (
            mark_job_done_if_active(
                job_id
            )
        )

        if not marked_done:

            current_job_status = (
                get_job_status(
                    job_id
                )
            )

            if current_job_status == "cancelled":

                print(
                    f"🛑 Job {job_id} "
                    f"was cancelled before "
                    f"it could be marked done."
                )

                return

            raise RuntimeError(
                "job_could_not_be_marked_done"
            )

        print(
            f"=== Job {job_id} finished "
            f"({len(processed)}/{total}) ==="
        )

    # ============================================================
    # JOB-LEVEL ERROR
    # ============================================================

    except Exception as exc:

        error_text = (
            f"{type(exc).__name__}: "
            f"{exc}"
        )

        print(
            f"❌ Job {job_id} "
            f"error: "
            f"{error_text}"
        )

        traceback.print_exc()

        # ========================================================
        # VERY IMPORTANT:
        #
        # If the office already cancelled the job,
        # DO NOT change it to failed.
        # ========================================================

        current_job_status = (
            get_job_status(
                job_id
            )
        )

        if current_job_status == "cancelled":

            print(
                f"🛑 Job {job_id} "
                f"is already cancelled. "
                f"Keeping status as cancelled."
            )

            return

        # ========================================================
        # Otherwise mark FAILED ONLY IF STILL PROCESSING.
        # ========================================================

        try:

            marked_failed = (
                mark_job_failed_if_active(
                    job_id,
                    error_text
                )
            )

            if marked_failed:

                print(
                    f"❌ Job {job_id} "
                    f"marked as failed."
                )

            else:

                latest_status = (
                    get_job_status(
                        job_id
                    )
                )

                if latest_status == "cancelled":

                    print(
                        f"🛑 Job {job_id} "
                        f"became cancelled "
                        f"while handling "
                        f"the error."
                    )

        except Exception as status_exc:

            print(
                f"❌ Could not mark "
                f"job failed: "
                f"{status_exc}"
            )

        # ========================================================
        # DEVELOPER NOTIFICATION
        # ========================================================

        try:

            latest_status = (
                get_job_status(
                    job_id
                )
            )

            if latest_status != "cancelled":

                notify_developer_error(
                    office,
                    job_id,
                    error_text
                )

        except Exception as notify_exc:

            print(
                f"Developer notification "
                f"warning: "
                f"{notify_exc}"
            )

    # ============================================================
    # ALWAYS CLOSE BROWSER
    # ============================================================

    finally:

        safe_quit(
            driver
        )


# ============================================================
# MAIN WORKER LOOP
# ============================================================

def main():

    print(
        f"Worker (Selenium + Supabase) "
        f"running... ID: {WORKER_ID}"
    )

    while True:

        try:

            job = claim_next_pending_job()

            if job:

                process_job(
                    job
                )

            else:

                time.sleep(
                    random.uniform(
                        JOB_POLL_INTERVAL_SECONDS * 0.7,
                        JOB_POLL_INTERVAL_SECONDS * 1.3
                    )
                )

        except KeyboardInterrupt:

            print(
                "Worker stopped manually."
            )

            break

        except Exception as exc:

            print(
                f"❌ Unexpected worker error: "
                f"{type(exc).__name__}: "
                f"{exc}"
            )

            traceback.print_exc()

            time.sleep(
                10
            )


if __name__ == "__main__":
    main()
